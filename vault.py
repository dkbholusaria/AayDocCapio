import os
import re
import sys
import csv
import json
import uuid
import datetime
from base64 import urlsafe_b64encode
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

_PAN_RE = re.compile(r'^[A-Z]{3}[PCHFATBLJG][A-Z][0-9]{4}[A-Z]$')
_DOB_RE = re.compile(r'^(\d{2})-(\d{2})-(\d{4})$')

# Accepted import formats → normalised DD-MM-YYYY
_DOB_FORMATS = [
    '%d-%m-%Y',   # 06-07-1974  (already correct)
    '%d/%m/%Y',   # 06/07/1974
    '%d.%m.%Y',   # 06.07.1974
    '%Y-%m-%d',   # 1974-07-06  (ISO)
    '%Y/%m/%d',   # 1974/07/06
    '%d-%m-%y',   # 06-07-74
    '%d/%m/%y',   # 06/07/74
    '%d.%m.%y',   # 06.07.74
    '%d %m %Y',   # 06 07 1974
    '%d %B %Y',   # 06 July 1974
    '%d-%B-%Y',   # 06-July-1974
]

def _normalise_dob(raw: str) -> str:
    """Convert any recognised date string to DD-MM-YYYY; return raw if unrecognised."""
    # Strip time component if present (e.g. "06/07/1974 00:00:00")
    s = raw.strip().split(' ')[0].split('T')[0]
    for fmt in _DOB_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).strftime('%d-%m-%Y')
        except ValueError:
            continue
    return raw.strip()  # let _validate_fields report the error with the original value

def _validate_fields(name: str, pan: str, dob: str, password: str):
    name = name.strip()
    pan  = pan.strip().upper()
    dob  = dob.strip()

    if not name:
        raise ValueError("Full Name is required.")
    if len(name) < 2:
        raise ValueError("Full Name must be at least 2 characters.")

    if not pan:
        raise ValueError("PAN Number is required.")
    if not _PAN_RE.match(pan):
        raise ValueError(
            "Invalid PAN format.\n\n"
            "Format: AAA · T · N · 0001 · Z\n"
            "  · Characters 1–3 : any letters (A–Z)\n"
            "  · Character 4    : P/C/H/F/A/T/B/L/J/G (taxpayer type)\n"
            "  · Character 5    : any letter (A–Z)\n"
            "  · Characters 6–9 : 4 digits\n"
            "  · Character 10   : any letter (A–Z)\n\n"
            "Example: AAAPT0001A")

    if not dob:
        raise ValueError("Date of Birth is required.")
    m = _DOB_RE.match(dob)
    if not m:
        raise ValueError("Date of Birth must be in DD-MM-YYYY format.\nExample: 01-01-1980")
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        dt = datetime.date(year, month, day)
    except ValueError:
        raise ValueError(f"'{dob}' is not a valid date. Please use DD-MM-YYYY.")
    today = datetime.date.today()
    if dt >= today:
        raise ValueError("Date of Birth cannot be today or a future date.")
    if year < 1900:
        raise ValueError("Date of Birth year seems incorrect (before 1900).")

    if not password:
        raise ValueError("Portal Password is required.")
    if len(password) < 4:
        raise ValueError("Password is too short (minimum 4 characters).")

class VaultManager:
    """
    Manages secure CRUD operations for Standalone Tax Downloader.
    Stores details in an encrypted local JSON vault.
    """
    def __init__(self, vault_path=None, master_password="automated_tax_app_key"):
        if vault_path is None:
            # When frozen by PyInstaller use folder next to .exe, not _MEIPASS
            if getattr(sys, "frozen", False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            self.vault_file = os.path.join(base_dir, "tax_vault.json")
        else:
            self.vault_file = vault_path
        
        # Derive cryptographic key
        salt = b'secure_tax_salt'
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        self._key = urlsafe_b64encode(kdf.derive(master_password.encode()))
        self._cipher = Fernet(self._key)
        
        # Initialize and migrate/ensure schema
        self._ensure_vault()

    def _ensure_vault(self):
        dir_name = os.path.dirname(self.vault_file)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        if not os.path.exists(self.vault_file):
            self._save_raw({"assessees": [], "settings": {}})
        else:
            data = self._get_raw()
            updated = False
            if "assessees" not in data:
                data["assessees"] = []
                updated = True
            if "settings" not in data:
                data["settings"] = {}
                updated = True

            # Migrate old entries if needed (add uuid, clean up keys)
            for entry in data["assessees"]:
                if "id" not in entry:
                    entry["id"] = str(uuid.uuid4())
                    updated = True
                if "pan" in entry:
                    entry["pan"] = entry["pan"].strip().upper()
                if "email" not in entry:
                    entry["email"] = ""
                    updated = True
                if "cc" not in entry:
                    entry["cc"] = ""
                    updated = True

            if updated:
                self._save_raw(data)

    def _get_raw(self):
        try:
            with open(self.vault_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {"assessees": [], "settings": {}}

    def _save_raw(self, data):
        with open(self.vault_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4)

    def encrypt_password(self, password: str) -> str:
        if not password:
            return ""
        return self._cipher.encrypt(password.encode('utf-8')).decode('utf-8')

    def decrypt_password(self, encrypted_pwd: str) -> str:
        if not encrypted_pwd:
            return ""
        try:
            return self._cipher.decrypt(encrypted_pwd.encode('utf-8')).decode('utf-8')
        except Exception:
            return ""

    # --- Assessee CRUD Operations ---

    def get_all_assessees(self):
        """Returns all assessees with decrypted passwords for utility."""
        raw_data = self._get_raw()
        assessees = []
        for a in raw_data.get("assessees", []):
            decrypted = a.copy()
            decrypted["password"] = self.decrypt_password(a.get("password_enc", ""))
            assessees.append(decrypted)
        return assessees

    def add_update_assessee(self, name: str, pan: str, dob: str, password: str,
                            assessee_id: str = None, email: str = "", cc: str = "") -> str:
        """Adds or updates a single assessee."""
        raw_data = self._get_raw()
        _validate_fields(name, pan, dob, password)
        pan = pan.strip().upper()

        password_enc = self.encrypt_password(password)

        found = False
        if assessee_id:
            # Update by ID
            for i, a in enumerate(raw_data["assessees"]):
                if a.get("id") == assessee_id:
                    raw_data["assessees"][i] = {
                        "id": assessee_id,
                        "name": name.strip(),
                        "pan": pan,
                        "dob": dob.strip(),
                        "password_enc": password_enc,
                        "email": email.strip(),
                        "cc": cc.strip(),
                    }
                    found = True
                    break
        else:
            # Check if PAN already exists to update it, or treat as new
            for i, a in enumerate(raw_data["assessees"]):
                if a.get("pan") == pan:
                    assessee_id = a.get("id") or str(uuid.uuid4())
                    raw_data["assessees"][i] = {
                        "id": assessee_id,
                        "name": name.strip(),
                        "pan": pan,
                        "dob": dob.strip(),
                        "password_enc": password_enc,
                        "email": email.strip(),
                        "cc": cc.strip(),
                    }
                    found = True
                    break

        if not found:
            new_id = str(uuid.uuid4())
            raw_data["assessees"].append({
                "id": new_id,
                "name": name.strip(),
                "pan": pan,
                "dob": dob.strip(),
                "password_enc": password_enc,
                "email": email.strip(),
                "cc": cc.strip(),
            })
            assessee_id = new_id

        self._save_raw(raw_data)
        return assessee_id

    def update_assessee_email(self, pan: str, email: str, cc: str = "") -> bool:
        """Update only the email and cc fields for a client by PAN. Returns True if found."""
        raw_data = self._get_raw()
        pan = pan.strip().upper()
        for entry in raw_data["assessees"]:
            if entry.get("pan") == pan:
                entry["email"] = email.strip()
                entry["cc"] = cc.strip()
                self._save_raw(raw_data)
                return True
        return False

    def delete_assessee(self, assessee_id: str):
        """Deletes an assessee by ID."""
        raw_data = self._get_raw()
        raw_data["assessees"] = [a for a in raw_data["assessees"] if a.get("id") != assessee_id]
        self._save_raw(raw_data)

    # --- Download History ---

    def record_download(self, pan: str, ay_label: str, status: str, path: str):
        """Record the last download status + path for a client/AY pair."""
        raw_data = self._get_raw()
        pan = pan.strip().upper()
        hist = raw_data.setdefault("download_history", {})
        hist.setdefault(pan, {})[ay_label] = {
            "status": status,
            "path": path,
            "ts": datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S"),
        }
        self._save_raw(raw_data)

    def get_download_history(self, ay_label: str) -> dict:
        """Return {pan: {"status":..., "path":..., "ts":...}} for the given AY label."""
        raw_data = self._get_raw()
        hist = raw_data.get("download_history", {})
        result = {}
        for pan, ay_map in hist.items():
            if ay_label in ay_map:
                result[pan] = ay_map[ay_label]
        return result

    # --- Bulk Import / Export ---

    def import_bulk(self, file_path: str) -> tuple:
        """
        Imports assessees from an Excel (.xlsx) or CSV (.csv) file.
        Expects columns: Name, PAN, DOB, Password
        Returns: (success_count, error_messages_list)
        """
        if not os.path.exists(file_path):
            return 0, [f"File {file_path} does not exist."]
        
        try:
            if file_path.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                raw_rows = list(ws.iter_rows(values_only=True))
                if not raw_rows:
                    return 0, ["File is empty."]
                headers = [str(c).strip().lower() if c is not None else "" for c in raw_rows[0]]
                data_rows = raw_rows[1:]
            elif file_path.endswith('.csv'):
                with open(file_path, newline='', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    raw_rows = list(reader)
                if not raw_rows:
                    return 0, ["File is empty."]
                headers = [c.strip().lower() for c in raw_rows[0]]
                data_rows = raw_rows[1:]
            else:
                return 0, ["Unsupported file format. Please use Excel (.xlsx) or CSV (.csv)."]
        except Exception as e:
            return 0, [f"Failed to read file: {str(e)}"]

        required_cols = {'name', 'pan', 'dob', 'password'}
        missing = required_cols - set(headers)
        if missing:
            return 0, [f"Missing required columns: {', '.join(missing)}. Headers must include: Name, PAN, DOB, Password."]

        col = {h: i for i, h in enumerate(headers)}
        added_count = 0
        updated_count = 0
        errors = []

        existing_pans = {a.get("pan") for a in self.get_all_assessees()}

        for idx, row in enumerate(data_rows):
            row_num = idx + 2
            try:
                def _cell(key):
                    v = row[col[key]] if key in col and col[key] < len(row) else None
                    return str(v).strip() if v is not None and str(v).strip() not in ("", "None") else ""

                name = _cell('name')
                pan = _cell('pan').upper()

                dob_val = row[col['dob']] if col['dob'] < len(row) else None
                if dob_val is None or str(dob_val).strip() in ("", "None"):
                    dob = ""
                elif isinstance(dob_val, (datetime.datetime, datetime.date)):
                    dob = dob_val.strftime('%d-%m-%Y')
                else:
                    dob = _normalise_dob(str(dob_val).strip())

                password = _cell('password')
                email = _cell('email') if 'email' in col else ""
                cc = _cell('cc') if 'cc' in col else ""

                if not name or not pan or not dob or not password:
                    errors.append(f"Row {row_num}: Missing values in Name, PAN, DOB, or Password.")
                    continue

                if len(pan) != 10:
                    errors.append(f"Row {row_num}: Invalid PAN length (must be 10 characters).")
                    continue

                is_existing = pan in existing_pans
                self.add_update_assessee(name, pan, dob, password, email=email, cc=cc)
                if is_existing:
                    updated_count += 1
                else:
                    existing_pans.add(pan)
                    added_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: Error importing entry: {str(e)}")

        return added_count, updated_count, errors

    def generate_template(self, file_path: str):
        """Generates an Excel import template with sample columns."""
        headers = ["Name", "PAN", "DOB", "Password", "Email", "CC"]
        sample = ["John Doe", "AAAPT0001A", "01-01-1980", "YourPortalPassword",
                  "client@example.com", "spouse@example.com;accountant@firm.com"]
        if file_path.endswith('.csv'):
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                csv.writer(f).writerows([headers, sample])
        else:
            from openpyxl import Workbook
            target = file_path if file_path.endswith('.xlsx') else file_path + ".xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            ws.append(sample)
            wb.save(target)

    def export_data(self, file_path: str):
        """Exports all saved assessees (with decrypted passwords) to Excel or CSV."""
        assessees = self.get_all_assessees()
        headers = ["Name", "PAN", "DOB", "Password", "Email", "CC"]
        rows = [[a.get("name", ""), a.get("pan", ""), a.get("dob", ""), a.get("password", ""),
                 a.get("email", ""), a.get("cc", "")]
                for a in assessees]
        if file_path.endswith('.csv'):
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(rows)
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            for row in rows:
                ws.append(row)
            wb.save(file_path)

    # --- Settings Management ---

    def get_setting(self, key: str, default=None):
        raw_data = self._get_raw()
        return raw_data.get("settings", {}).get(key, default)

    def update_setting(self, key: str, value):
        raw_data = self._get_raw()
        raw_data["settings"][key] = value
        self._save_raw(raw_data)

    # --- Email / SMTP Settings ---

    _EMAIL_DEFAULTS = {
        "smtp_host": "",
        "smtp_port": "587",
        "smtp_user": "",
        "smtp_from": "",            # optional Send-As address; uses smtp_user when blank
        "smtp_password_enc": "",
        "smtp_use_tls": True,       # legacy — kept for back-compat; use smtp_encryption
        "smtp_encryption": "STARTTLS",  # "STARTTLS" | "SSL/TLS" | "None"
        "firm_name": "",
        "bcc_addresses": "",
        "email_subject_tpl": "[Action Required] Your Annual Income Tax Documents for {ay} | {client_name}",
        "email_body_tpl": (
            "<p>Dear {client_name},</p>"
            "<p>We hope this message finds you well.</p>"
            "<p>Please find attached your Income Tax documents for <b>{ay}</b>, "
            "as downloaded from the Income Tax Department's e-Filing portal:</p>"
            "<p>{documents}</p>"
            "<p>Kindly review the attached documents at your earliest convenience. "
            "If you notice any discrepancies or have any queries, please do not "
            "hesitate to contact us and we will be happy to assist you.</p>"
            "<p>Please note that these documents are sourced directly from the "
            "IT Department portal and are provided for your reference and records.</p>"
            "<p>Warm regards,<br><b>{firm_name}</b></p>"
        ),
    }

    def get_email_settings(self) -> dict:
        raw = self._get_raw().get("settings", {})
        cfg = {}
        for k, default in self._EMAIL_DEFAULTS.items():
            cfg[k] = raw.get(k, default)
        # Migrate legacy smtp_use_tls → smtp_encryption
        if "smtp_encryption" not in raw and "smtp_use_tls" in raw:
            cfg["smtp_encryption"] = "STARTTLS" if raw["smtp_use_tls"] else "None"
        # Migrate plain-text body template → HTML
        body = cfg.get("email_body_tpl", "")
        if body and not body.lstrip().startswith("<"):
            import html as _html
            cfg["email_body_tpl"] = (
                "<p style='white-space:pre-wrap'>"
                + _html.escape(body).replace("\n", "<br>")
                + "</p>"
            )
        # Decrypt password for caller
        cfg["smtp_password"] = self.decrypt_password(cfg.get("smtp_password_enc", ""))
        return cfg

    def save_email_settings(self, cfg: dict):
        """Save SMTP settings. Pass plain-text 'smtp_password'; it will be encrypted."""
        raw_data = self._get_raw()
        s = raw_data.setdefault("settings", {})
        for k in self._EMAIL_DEFAULTS:
            if k == "smtp_password_enc":
                continue
            if k in cfg:
                s[k] = cfg[k]
        if "smtp_password" in cfg and cfg["smtp_password"]:
            s["smtp_password_enc"] = self.encrypt_password(cfg["smtp_password"])
        elif "smtp_password" in cfg and not cfg["smtp_password"]:
            # Blank password clears the stored encrypted value
            s["smtp_password_enc"] = ""
        self._save_raw(raw_data)

    _ALL_DOCS = ["26as_pdf", "26as_xlsx", "ais_pdf", "ais_xlsx", "tis_pdf"]

    def _make_legacy_template(self, cfg: dict) -> dict:
        return {
            "name":    "Legacy Template",
            "subject": cfg.get("email_subject_tpl", self._EMAIL_DEFAULTS["email_subject_tpl"]),
            "body":    cfg.get("email_body_tpl",    self._EMAIL_DEFAULTS["email_body_tpl"]),
            "docs":    {k: True for k in self._ALL_DOCS},
        }

    def get_email_templates(self) -> list:
        """Return list of template dicts. Auto-creates 'Legacy Template' from existing settings if none exist."""
        raw_data = self._get_raw()
        s = raw_data.get("settings", {})
        templates = s.get("email_templates")
        if not templates:
            # First run — migrate existing subject/body into Legacy Template
            cfg = self.get_email_settings()
            templates = [self._make_legacy_template(cfg)]
            s = raw_data.setdefault("settings", {})
            s["email_templates"]  = templates
            s["active_template"]  = "Legacy Template"
            self._save_raw(raw_data)
        return templates

    def save_email_templates(self, templates: list):
        raw_data = self._get_raw()
        raw_data.setdefault("settings", {})["email_templates"] = templates
        self._save_raw(raw_data)

    def get_active_template_name(self) -> str:
        s = self._get_raw().get("settings", {})
        templates = s.get("email_templates", [])
        name = s.get("active_template", "")
        # Fall back to first template if stored name no longer exists
        if templates and not any(t["name"] == name for t in templates):
            name = templates[0]["name"]
        return name

    def set_active_template(self, name: str):
        raw_data = self._get_raw()
        raw_data.setdefault("settings", {})["active_template"] = name
        self._save_raw(raw_data)

    def get_active_template(self) -> dict | None:
        templates = self.get_email_templates()
        name = self.get_active_template_name()
        for t in templates:
            if t["name"] == name:
                return t
        return templates[0] if templates else None
