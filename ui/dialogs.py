import os, json, datetime, threading
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QCheckBox, QRadioButton, QLineEdit, QScrollArea, QWidget,
    QFrame, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QProgressBar, QSizePolicy, QFileDialog, QMessageBox,
    QTextEdit, QSpinBox, QComboBox, QFontComboBox, QTabWidget,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import (
    QFont, QColor, QPixmap, QPainter, QBrush, QPen,
    QTextCharFormat, QIcon,
)

from ui._theme import _t
from ui.helpers import _btn, _lbl, _status_style, _UI_FONT, _icon_path
from config import _open_path, _log_open
from themes import MONO_FONT_NAME as _MONO_FONT


# ── Manage Years Dialog ───────────────────────────────────────────────────────

class ManageYearsDialog(QDialog):
    def __init__(self, parent, json_path: str, on_save):
        super().__init__(parent)
        self.setWindowTitle("Manage Assessment / Tax Years")
        self.setFixedSize(500, 560)
        self.setModal(True)
        self._json_path = json_path
        self._on_save = on_save
        self._checkboxes = []

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                entries = json.load(f)
        except Exception:
            entries = []

        self._build_ui(entries)

    def _build_ui(self, entries):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}"
            f"QRadioButton{{color:{t.text_primary};background:transparent;spacing:6px;}}"
            f"QRadioButton::indicator{{width:14px;height:14px;border:1.5px solid {t.border};"
            f"border-radius:7px;background:{t.bg_checkbox};}}"
            f"QRadioButton::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
            f"QScrollArea{{background:{t.bg_input};border:1px solid {t.border};border-radius:6px;}}"
            f"QScrollArea > QWidget > QWidget{{background:{t.bg_input};}}"
        )
        main = QVBoxLayout(self)
        main.setContentsMargins(20, 16, 20, 16)
        main.setSpacing(8)

        # Header: logo left, title+subtitle stacked right
        from config import _bundled_dir
        hdr_row = QHBoxLayout()
        hdr_row.setSpacing(12)
        hdr_row.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        _icon_lbl = QLabel()
        _icon_lbl.setStyleSheet("background:transparent;")
        _ip = os.path.join(_bundled_dir(), "resources", "app_icon.png")
        if os.path.isfile(_ip):
            _icon_lbl.setPixmap(QPixmap(_ip).scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio,
                                                     Qt.TransformationMode.SmoothTransformation))
        hdr_row.addWidget(_icon_lbl)
        txt_col = QVBoxLayout()
        txt_col.setSpacing(2)
        txt_col.addWidget(_lbl("Manage Assessment / Tax Years", 13, bold=True))
        txt_col.addWidget(_lbl("Toggle enabled/disabled or add new years.", 10, color=t.text_muted))
        hdr_row.addLayout(txt_col)
        hdr_row.addStretch()
        main.addLayout(hdr_row)
        main.addWidget(_lbl("Existing Entries", 11, bold=True, color=t.text_muted))

        scroll = QScrollArea()
        scroll.setFixedHeight(180)
        scroll.setWidgetResizable(True)
        inner = QWidget()
        inner.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        inner.setStyleSheet(f"QWidget{{background:{t.bg_input};}}")
        self._list_layout = QVBoxLayout(inner)
        self._list_layout.setSpacing(2)
        self._list_layout.addStretch()
        scroll.setWidget(inner)
        main.addWidget(scroll)

        for e in entries:
            self._add_row(e)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep)

        main.addWidget(_lbl("Add New Year", 11, bold=True, color=t.text_muted))

        type_row = QHBoxLayout()
        type_row.addWidget(_lbl("Type:", 11))
        self._type_ay = QRadioButton("AY (Assessment Year)"); self._type_ay.setChecked(True)
        self._type_ty = QRadioButton("TY (Tax Year)")
        self._type_ay.toggled.connect(self._auto_fy)
        type_row.addWidget(self._type_ay); type_row.addWidget(self._type_ty)
        type_row.addStretch(); main.addLayout(type_row)

        yr_row = QHBoxLayout()
        yr_row.addWidget(_lbl("Year:", 11))
        self._year_edit = QLineEdit(); self._year_edit.setPlaceholderText("e.g. 2027-28")
        self._year_edit.setFixedWidth(120); self._year_edit.textChanged.connect(self._auto_fy)
        yr_row.addWidget(self._year_edit)
        yr_row.addWidget(_lbl("FY:", 11))
        self._fy_edit = QLineEdit(); self._fy_edit.setPlaceholderText("auto-filled")
        self._fy_edit.setFixedWidth(120)
        yr_row.addWidget(self._fy_edit)
        yr_row.addWidget(_lbl("(editable)", 10, color=t.text_muted))
        yr_row.addStretch(); main.addLayout(yr_row)

        add_btn = _btn("Add to List", "outline", height=32, min_width=130, icon="btn_add_list.png")
        add_btn.clicked.connect(self._add_entry)
        main.addWidget(add_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        sep2 = QFrame(); sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep2)

        btns_row = QHBoxLayout()
        save_btn = _btn("Save & Close", "primary", height=36, icon="btn_save_close.png")
        save_btn.clicked.connect(self._save)
        cancel_btn = _btn("Cancel", "secondary", height=36, icon="btn_cancel.png")
        cancel_btn.clicked.connect(self.reject)
        btns_row.addWidget(save_btn); btns_row.addWidget(cancel_btn)
        main.addLayout(btns_row)

    def _add_row(self, entry):
        t = _t()
        cb = QCheckBox(entry["label"])
        cb.setChecked(entry.get("enabled", True))
        cb.setStyleSheet(
            f"QCheckBox{{color:{t.text_primary};background:transparent;spacing:6px;}}"
            f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
            f"border-radius:3px;background:{t.bg_checkbox};}}"
            f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
        )
        self._checkboxes.append((entry, cb))
        self._list_layout.insertWidget(self._list_layout.count() - 1, cb)

    def _auto_fy(self):
        import re
        year = self._year_edit.text().strip()
        m = re.match(r"(\d{4})-(\d{2}|\d{4})$", year)
        if not m:
            return
        y1 = int(m.group(1)); suffix = m.group(2)
        y2 = int(str(y1)[:2] + suffix) if len(suffix) == 2 else int(suffix)
        fy = f"{y1}-{str(y2)[-2:]}" if self._type_ty.isChecked() else f"{y1-1}-{str(y2-1)[-2:]}"
        self._fy_edit.setText(fy)

    def _add_entry(self):
        year_type = "TY" if self._type_ty.isChecked() else "AY"
        year = self._year_edit.text().strip()
        fy = self._fy_edit.text().strip()
        if not year or not fy:
            QMessageBox.warning(self, "Missing Fields", "Please fill in Year and FY.")
            return
        label = f"{year_type} {year} (FY {fy})"
        if any(e["label"] == label for e, _ in self._checkboxes):
            QMessageBox.warning(self, "Duplicate", f'"{label}" already exists.')
            return
        year_obj = {"TY": year, "FY": fy} if year_type == "TY" else {"AY": year, "FY": fy}
        self._add_row({"label": label, "enabled": True, "year": year_obj})
        self._year_edit.clear(); self._fy_edit.clear()

    def _save(self):
        final = [{**e, "enabled": cb.isChecked()} for e, cb in self._checkboxes]
        def _sort_key(e):
            y = e.get("year", {})
            label_year = y.get("AY") or y.get("TY") or y.get("FY") or "0000-00"
            try:
                return (0 if not e.get("enabled", True) else 1, -int(label_year[:4]))
            except ValueError:
                return (1, 0)
        final.sort(key=_sort_key)
        try:
            with open(self._json_path, "w", encoding="utf-8") as f:
                json.dump(final, f, indent=2, ensure_ascii=False)
            self._on_save()
            self.accept()
        except Exception as ex:
            QMessageBox.critical(self, "Save Error", str(ex))


# ── Batch Progress Dialog ─────────────────────────────────────────────────────

class BatchProgressDialog(QDialog):
    """
    Live progress popup shown during any batch run.
    Columns: Name | PAN | Status | Save Path (clickable link)
    Status and path updates arrive from the worker thread via Qt signals.
    """
    _update_signal = pyqtSignal(str, str)
    _path_signal   = pyqtSignal(str, str)
    _resume_signal = pyqtSignal(list)

    _COL_NAME   = 0
    _COL_PAN    = 1
    _COL_STATUS = 2
    _COL_PATH   = 3

    def __init__(self, targets: list, mode: str, ay: str = "",
                 stop_callback=None, resume_callback=None, skip_callback=None,
                 tray_callback=None, output_dir: str = "", parent=None):
        super().__init__(parent)
        self._stop_callback   = stop_callback
        self._resume_callback = resume_callback
        self._skip_callback   = skip_callback
        self._tray_callback   = tray_callback
        self._output_dir      = output_dir
        self._mode            = mode
        self._ay              = ay
        self._targets         = targets
        self._pan_to_path     = {}

        mode_label = {
            "26as":        "Downloading 26AS",
            "request_ais": "Requesting AIS Generation",
            "ais_tis":     "Downloading AIS / TIS",
        }.get(mode, "Batch Run")
        self._mode_label = mode_label

        self.setWindowTitle(f"{mode_label} — Batch Progress")
        self.setMinimumSize(900, 500)
        self.resize(1060, min(160 + len(targets) * 42, 720))
        self.setSizeGripEnabled(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        self._pan_to_row   = {}
        self._counted_pans: set = set()   # B-07: track which PANs already incremented done count
        self._done_count   = 0
        self._total        = len(targets)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        # ── Title bar ─────────────────────────────────────────────────────────
        ay_tag = (f" &nbsp;·&nbsp; <span style='color:{_bt.accent}'>{ay}</span>") if ay else ""
        title = QLabel(f"<b>{mode_label}</b> — {len(targets)} client(s){ay_tag}")
        title.setStyleSheet(f"font-size:14px; color:{_bt.text_primary}; background:transparent;")
        layout.addWidget(title)

        # ── Table ─────────────────────────────────────────────────────────────
        self._table = QTableWidget(len(targets), 4)
        self._table.setHorizontalHeaderLabels(["Name", "PAN", "Status", "Save Path"])

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_NAME,   QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN,    QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_STATUS, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PATH,   QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(self._COL_NAME,   180)
        self._table.setColumnWidth(self._COL_PAN,    120)
        self._table.setColumnWidth(self._COL_STATUS, 260)

        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setShowGrid(True)
        self._table.setAlternatingRowColors(False)
        self._table.setWordWrap(False)
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}"
            f"QPushButton{{border:none;background:transparent;font-size:14px;}}"
            f"QPushButton:hover{{background:{_bt.bg_table_alt};border-radius:4px;}}")
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        for row, tgt in enumerate(targets):
            pan  = tgt.get("pan", "")
            name = tgt.get("name", "—")
            self._pan_to_row[pan] = row
            self._table.setRowHeight(row, 40)

            name_item = QTableWidgetItem(name)
            name_item.setForeground(QColor(_bt.text_primary))
            self._table.setItem(row, self._COL_NAME, name_item)

            pan_item = QTableWidgetItem(pan)
            pan_item.setForeground(QColor(_bt.text_muted))
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_PAN, pan_item)

            self._set_status_item(row, "⬜ Waiting")

            path_lbl = QLabel("—")
            path_lbl.setStyleSheet(
                f"color:{_bt.text_muted};font-size:11px;padding:0 8px;background:transparent;")
            path_lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            path_lbl.setWordWrap(False)
            path_lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            path_lbl.setOpenExternalLinks(False)
            path_lbl.linkActivated.connect(self._open_row_path)
            self._table.setCellWidget(row, self._COL_PATH, path_lbl)

        layout.addWidget(self._table, stretch=1)

        # ── Progress bar ──────────────────────────────────────────────────────
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, len(targets))
        self._progress_bar.setValue(0)
        self._progress_bar.setFixedHeight(18)
        self._progress_bar.setTextVisible(True)
        self._progress_bar.setFormat(f"0 / {len(targets)} done")
        self._progress_bar.setStyleSheet(
            f"QProgressBar{{border:1px solid {_bt.border};border-radius:9px;"
            f"background:{_bt.scrollbar_handle};text-align:center;font-size:11px;"
            f"font-weight:600;color:{_bt.accent_text};}}"
            f"QProgressBar::chunk{{background:#16A34A;border-radius:9px;}}")
        layout.addWidget(self._progress_bar)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = QHBoxLayout()
        footer.setSpacing(8)
        footer.setContentsMargins(0, 0, 0, 0)

        loc_cap = QLabel("📁")
        loc_cap.setStyleSheet("font-size:13px;background:transparent;")
        loc_cap.setFixedWidth(18)
        footer.addWidget(loc_cap)

        self._loc_val = QLabel(output_dir or "—")
        self._loc_val.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        self._loc_val.setWordWrap(False)
        self._loc_val.setMinimumWidth(0)
        self._loc_val.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self._loc_val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        footer.addWidget(self._loc_val, stretch=1)

        self._open_folder_btn = QPushButton("📂  Open Folder")
        self._open_folder_btn.setFixedHeight(32)
        self._open_folder_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._open_folder_btn.clicked.connect(self._open_output_dir)
        footer.addWidget(self._open_folder_btn)

        self._report_btn = QPushButton("📂  Open Report")
        self._report_btn.setFixedHeight(32)
        self._report_btn.setEnabled(False)
        self._report_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;padding:0 12px;}}"
            f"QPushButton:enabled:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._report_btn.clicked.connect(self._open_last_report)
        self._last_report_path = ""
        footer.addWidget(self._report_btn)

        self._skip_btn = QPushButton("⏭  Skip")
        self._skip_btn.setFixedHeight(32)
        self._skip_btn.setMinimumWidth(80)
        self._skip_btn.setToolTip("Skip the currently-downloading client and move to the next")
        self._skip_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._skip_btn.clicked.connect(self._on_skip_clicked)
        footer.addWidget(self._skip_btn)

        self._tray_btn = QPushButton("⬇  Tray")
        self._tray_btn.setFixedHeight(32)
        self._tray_btn.setMinimumWidth(80)
        self._tray_btn.setToolTip("Hide to system tray — click the tray icon to restore")
        self._tray_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._tray_btn.setVisible(bool(self._tray_callback))
        self._tray_btn.clicked.connect(self._on_tray_clicked)
        footer.addWidget(self._tray_btn)

        self._stop_btn = QPushButton("⏹  Stop")
        self._stop_btn.setFixedHeight(32)
        self._stop_btn.setMinimumWidth(90)
        self._stop_btn.setStyleSheet(
            "QPushButton{background:#EF4444;color:#FFFFFF;border:none;"
            "border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:#DC2626;}"
            "QPushButton:disabled{background:#E2E8F0;color:#94A3B8;}")
        self._stop_btn.clicked.connect(self._on_stop_clicked)
        footer.addWidget(self._stop_btn)

        self._resume_btn = QPushButton("▶  Resume")
        self._resume_btn.setFixedHeight(32)
        self._resume_btn.setMinimumWidth(100)
        self._resume_btn.setVisible(False)
        self._resume_btn.setStyleSheet(
            "QPushButton{background:#16A34A;color:#FFFFFF;border:none;"
            "border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:#15803D;}")
        self._resume_btn.clicked.connect(self._on_resume_clicked)
        footer.addWidget(self._resume_btn)

        self._close_btn = QPushButton("Close")
        self._close_btn.setFixedSize(80, 32)
        self._close_btn.setEnabled(False)
        self._close_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.border};color:{_bt.text_muted};border:none;"
            f"border-radius:6px;font-size:12px;}}"
            f"QPushButton:enabled{{background:{_bt.accent};color:{_bt.accent_text};}}"
            f"QPushButton:enabled:hover{{background:{_bt.accent_hover};}}")
        self._close_btn.clicked.connect(self.accept)
        footer.addWidget(self._close_btn)

        layout.addLayout(footer)

        self._rows_data  = {}

        for tgt in targets:
            self._rows_data[tgt.get("pan", "")] = {
                "name": tgt.get("name", ""), "path": "", "status": "Waiting", "ts": ""}

        self._update_signal.connect(self._on_update)
        self._path_signal.connect(self._on_path_update)
        self._table.cellDoubleClicked.connect(self._on_row_double_clicked)

    # ── internal helpers ──────────────────────────────────────────────────────

    def _on_row_double_clicked(self, row: int, _col: int):
        """Show full status text for the double-clicked row."""
        status_item = self._table.item(row, self._COL_STATUS)
        name_item   = self._table.item(row, self._COL_NAME)
        if not status_item:
            return
        status = status_item.text()
        name   = name_item.text() if name_item else ""
        # Find timestamp from rows_data
        pan_item = self._table.item(row, self._COL_PAN)
        pan = pan_item.text() if pan_item else ""
        ts  = self._rows_data.get(pan, {}).get("ts", "")
        msg = f"{name}  ({pan})\n"
        if ts:
            msg += f"{ts}\n"
        msg += f"\n{status}"
        QMessageBox.information(self, "Status Detail", msg)

    def _open_output_dir(self):
        _log_open(f"[OpenFolder] Button clicked: {self._output_dir!r}")
        _open_path(self._output_dir)

    def _open_row_path(self, url: str):
        _log_open(f"[OpenFolder] Row link clicked: {url!r}")
        _open_path(url)

    def _set_status_item(self, row: int, text: str):
        _bt = _t()
        _, light_fg, dark_fg = _status_style(text)
        fg = dark_fg if getattr(_bt, "name", "").lower() != "light" else light_fg
        item = QTableWidgetItem(text)
        item.setForeground(QColor(fg))
        item.setFont(QFont(_UI_FONT, 10))
        item.setToolTip(text)
        self._table.setItem(row, self._COL_STATUS, item)

    def _on_update(self, pan: str, status: str):
        row = self._pan_to_row.get(pan)
        if row is None:
            return
        self._set_status_item(row, status)
        if pan in self._rows_data:
            self._rows_data[pan]["status"] = status
            self._rows_data[pan]["ts"] = datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S")
        terminal = ("✅", "❌", "🕐", "⬜", "⏹", "⚠")
        if any(status.startswith(p) for p in terminal):
            if pan not in self._counted_pans:
                self._counted_pans.add(pan)
                self._done_count += 1
                self._progress_bar.setValue(self._done_count)
                self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")
        if self._done_count >= self._total:
            self._close_btn.setEnabled(True)
            self._report_btn.setEnabled(True)
            self._progress_bar.setFormat(f"All {self._total} done")

    def _on_path_update(self, pan: str, folder: str):
        row = self._pan_to_row.get(pan)
        if row is None:
            return
        self._pan_to_path[pan] = folder
        if pan in self._rows_data:
            self._rows_data[pan]["path"] = folder
        lbl = self._table.cellWidget(row, self._COL_PATH)
        if isinstance(lbl, QLabel):
            lbl.setText(
                f'<a href="{folder}" style="color:#2563EB;text-decoration:underline;">'
                f'{folder}</a>')
            lbl.setToolTip(folder)
            lbl.setStyleSheet("font-size:11px;padding:0 8px;background:transparent;")

    def _on_skip_clicked(self):
        if self._skip_callback:
            self._skip_callback()
        self._skip_btn.setEnabled(False)
        self._skip_btn.setText("⏭  Skipping...")

    def _on_tray_clicked(self):
        if self._tray_callback:
            self._tray_callback()

    def _on_stop_clicked(self):
        if self._stop_callback:
            self._stop_callback()
        self._stop_btn.setEnabled(False)
        self._stop_btn.setText("⏹  Stopping...")

    # ── Excel report ──────────────────────────────────────────────────────────

    def _export_report(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"BatchReport_{self._ay.replace(' ','_')}_{timestamp}.xlsx"
        default_path = os.path.join(self._output_dir or os.path.expanduser("~"), default_name)

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Download Report", default_path, "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Download Report"

        hdr_fill  = PatternFill("solid", fgColor="0F172A")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        link_font = Font(color="2563EB", underline="single", size=10)
        body_font = Font(size=10)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        thin      = Side(style="thin", color="CBD5E1")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers    = ["#", "Client Name", "Save Folder", "Status", "Timestamp"]
        col_widths = [5, 30, 60, 40, 22]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[1].height = 22

        for seq, tgt in enumerate(self._targets, start=1):
            pan     = tgt.get("pan", "")
            data    = self._rows_data.get(pan, {})
            row_num = seq + 1
            folder  = data.get("path", "")
            status  = data.get("status", "—")
            name    = data.get("name", tgt.get("name", ""))
            row_ts  = data.get("ts", "")

            ws.cell(row=row_num, column=1, value=seq).alignment = center
            ws.cell(row=row_num, column=2, value=name).alignment = left

            if folder and os.path.exists(folder):
                cell = ws.cell(row=row_num, column=3, value=folder)
                cell.hyperlink = folder
                cell.font      = link_font
                cell.alignment = left
            else:
                ws.cell(row=row_num, column=3, value=folder or "—").alignment = left

            import re as _re
            status_clean = _re.sub(r'[^\x00-\x7F✅❌🕐⬜⏹⏳]+', '', status).strip()
            ws.cell(row=row_num, column=4, value=status_clean).alignment = left
            ws.cell(row=row_num, column=5, value=row_ts or "—").alignment = center

            for col_idx in range(1, 6):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border
                if not cell.font or cell.font == Font():
                    cell.font = body_font
            ws.row_dimensions[row_num].height = 18

        ws.freeze_panes = "A2"

        try:
            wb.save(path)
            _open_path(path)
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", str(e))

    # ── public API ────────────────────────────────────────────────────────────

    def set_status(self, pan: str, status: str):
        """Thread-safe status update."""
        self._update_signal.emit(pan, status)

    def set_client_path(self, pan: str, folder: str):
        """Thread-safe path update — call once the client folder is known."""
        self._path_signal.emit(pan, folder)

    def batch_finished(self, aborted: bool = False):
        """Enable Close/Report and hide Stop. If aborted, sweeps non-terminal rows to ⏹ Stopped."""
        terminal = ("✅", "❌", "🕐", "⏹", "⬜ Skipped")
        if aborted:
            for pan, row in self._pan_to_row.items():
                item = self._table.item(row, self._COL_STATUS)
                current = item.text() if item else ""
                if not any(current.startswith(t) for t in terminal):
                    self._set_status_item(row, "⏹ Stopped")
                    if pan in self._rows_data:
                        self._rows_data[pan]["status"] = "⏹ Stopped"
        self._skip_btn.setVisible(False)
        self._tray_btn.setVisible(False)
        self._stop_btn.setVisible(False)
        self._resume_btn.setVisible(aborted)
        self._close_btn.setEnabled(True)
        self._report_btn.setEnabled(True)
        n = self._done_count
        self._progress_bar.setValue(n)
        label = "Stopped" if aborted else "All done"
        self._progress_bar.setFormat(f"{label} — {n} / {self._total} processed")
        # F-37: auto-save report to output folder
        self._auto_save_report()

    def _open_last_report(self):
        if self._last_report_path and os.path.exists(self._last_report_path):
            _open_path(self._last_report_path)
        else:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "No Report", "No report has been generated yet.")

    def _auto_save_report(self):
        """F-37: silently save a timestamped report to _app_dir()/BatchReports/ after every run."""
        try:
            from config import _app_dir
            reports_dir = os.path.join(_app_dir(), "BatchReports")
            os.makedirs(reports_dir, exist_ok=True)
            import openpyxl, re as _re
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            timestamp    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            report_name  = f"BatchReport_{self._ay.replace(' ', '_')}_{timestamp}.xlsx"
            path         = os.path.join(reports_dir, report_name)
            hdr_fill  = PatternFill("solid", fgColor="0F172A")
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            link_font = Font(color="2563EB", underline="single", size=10)
            body_font = Font(size=10)
            center    = Alignment(horizontal="center", vertical="center")
            left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
            thin      = Side(style="thin", color="CBD5E1")
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Download Report"
            headers    = ["#", "Client Name", "Save Folder", "Status", "Timestamp"]
            col_widths = [5, 30, 60, 40, 22]
            for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = center; cell.border = border
                ws.column_dimensions[get_column_letter(col_idx)].width = w
            ws.row_dimensions[1].height = 22
            for seq, tgt in enumerate(self._targets, start=1):
                pan     = tgt.get("pan", "")
                data    = self._rows_data.get(pan, {})
                row_num = seq + 1
                folder  = data.get("path", "")
                status  = data.get("status", "—")
                name    = data.get("name", tgt.get("name", ""))
                row_ts  = data.get("ts", "")
                ws.cell(row=row_num, column=1, value=seq).alignment = center
                ws.cell(row=row_num, column=2, value=name).alignment = left
                if folder and os.path.exists(folder):
                    cell = ws.cell(row=row_num, column=3, value=folder)
                    cell.hyperlink = folder; cell.font = link_font; cell.alignment = left
                else:
                    ws.cell(row=row_num, column=3, value=folder or "—").alignment = left
                status_clean = _re.sub(r'[^\x00-\x7F✅❌🕐⬜⏹⏳]+', '', status).strip()
                ws.cell(row=row_num, column=4, value=status_clean).alignment = left
                ws.cell(row=row_num, column=5, value=row_ts or "—").alignment = center
                for col_idx in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    if not cell.font or cell.font == Font():
                        cell.font = body_font
                ws.row_dimensions[row_num].height = 18
            ws.freeze_panes = "A2"
            wb.save(path)
            self._last_report_path = path
        except Exception:
            pass   # Never crash the UI due to report saving

    def batch_resumed(self):
        """Reset UI back to running state after Resume is clicked."""
        self._resume_btn.setVisible(False)
        self._skip_btn.setText("⏭  Skip")
        self._skip_btn.setEnabled(True)
        self._skip_btn.setVisible(True)
        self._tray_btn.setVisible(bool(self._tray_callback))
        self._stop_btn.setText("⏹  Stop")
        self._stop_btn.setEnabled(True)
        self._stop_btn.setVisible(True)
        self._close_btn.setEnabled(False)
        self._report_btn.setEnabled(False)
        self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")

    def client_started(self):
        """Re-enable the Skip button when a new client begins downloading."""
        self._skip_btn.setText("⏭  Skip")
        self._skip_btn.setEnabled(True)

    def _on_resume_clicked(self):
        remaining = []
        for tgt in self._targets:
            pan = tgt.get("pan", "")
            status = (self._rows_data.get(pan) or {}).get("status", "")
            if status.startswith("⏹"):
                remaining.append(tgt)
                row = self._pan_to_row.get(pan)
                if row is not None:
                    self._set_status_item(row, "⬜ Waiting")
                    self._rows_data[pan]["status"] = "⬜ Waiting"
        if remaining and self._resume_callback:
            self._resume_callback(remaining)


# ── SMTP provider presets ─────────────────────────────────────────────────────

_SMTP_PRESETS = [
    {
        "name": "Gmail",
        "icon": "G", "icon_color": "#EA4335", "icon_file": "email_gmail.png",
        "host": "smtp.gmail.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Gmail requires an App Password — not your Google account password.<br>"
            "Go to: <a href='https://myaccount.google.com/apppasswords' style='color:#2563EB;'>myaccount.google.com → Security → 2-Step Verification → App Passwords</a> → Mail."
        ),
    },
    {
        "name": "Outlook.com",
        "icon": "O", "icon_color": "#0078D4", "icon_file": "email_outlook.png",
        "host": "smtp-mail.outlook.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Use your Outlook.com / Hotmail password.<br>"
            "If MFA is on, create an App Password at <a href='https://account.microsoft.com/security' style='color:#2563EB;'>account.microsoft.com → Security</a>."
        ),
    },
    {
        "name": "Office 365",
        "icon": "365", "icon_color": "#D83B01", "icon_file": "email_office365.png",
        "host": "smtp.office365.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "⚠ MFA enabled? Your regular password will NOT work — use an App Password instead.<br><br>"
            "To create an App Password:<br>"
            "1. Go to <a href='https://mysignins.microsoft.com/security-info' style='color:#2563EB;'>mysignins.microsoft.com/security-info</a><br>"
            "2. Click '+ Add sign-in method' → choose 'App password' → Next<br>"
            "3. Enter a name (e.g. AayDocCapio) → copy the generated password → paste it here<br><br>"
            "No MFA? Enable Authenticated SMTP in <a href='https://admin.microsoft.com' style='color:#2563EB;'>Microsoft 365 Admin Centre</a>:<br>"
            "Users → [your user] → Mail → Manage email apps → tick Authenticated SMTP."
        ),
    },
    {
        "name": "Exchange",
        "icon": "Ex", "icon_color": "#0F6CBD", "icon_file": "email_exchange.png",
        "host": "",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Enter your organisation's Exchange SMTP server address.<br>"
            "Typical format: <b>mail.yourcompany.com</b> — ask your IT admin if unsure."
        ),
    },
    {
        "name": "Yahoo",
        "icon": "Y!", "icon_color": "#6001D2", "icon_file": "email_yahoo.png",
        "host": "smtp.mail.yahoo.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Yahoo requires an App Password.<br>"
            "Go to: <a href='https://login.yahoo.com/account/security' style='color:#2563EB;'>Yahoo Account Security</a> → Generate app password → select 'Other app'."
        ),
    },
    {
        "name": "iCloud",
        "icon": "iC", "icon_color": "#3478F6", "icon_file": "email_icloud.png",
        "host": "smtp.mail.me.com",
        "port": 587,
        "encryption": "STARTTLS",
        "help": (
            "Use an App-Specific Password from <a href='https://appleid.apple.com' style='color:#2563EB;'>appleid.apple.com</a><br>"
            "→ Sign-In and Security → App-Specific Passwords → Generate."
        ),
    },
    {
        "name": "Custom",
        "icon": "⚙", "icon_color": "#64748B", "icon_file": "email_custom.png",
        "host": None,
        "port": None,
        "encryption": None,
        "help": "",
    },
]

# Map known SMTP hosts → preset name for auto-highlight on load
_HOST_TO_PRESET = {p["host"]: p["name"] for p in _SMTP_PRESETS if p["host"]}

# Pixmap cache — built once per process, reused across dialog opens
_TILE_PIXMAP_CACHE: dict[str, QPixmap] = {}


# ── Selective Export / Import Dialogs ────────────────────────────────────────

class _SelectiveExportDialog(QDialog):
    """Choose what to include in an email-settings export."""

    def __init__(self, parent, tpl_names: list[str]):
        super().__init__(parent)
        self.setWindowTitle("Export Email Settings")
        self.setModal(True)
        self.setMinimumWidth(380)
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        lay = QVBoxLayout(self)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 16, 20, 16)

        lay.addWidget(_lbl("Select what to export:", bold=True))

        # SMTP section
        self._chk_smtp = QCheckBox("SMTP / Sender Settings")
        self._chk_smtp.setChecked(True)
        lay.addWidget(self._chk_smtp)

        # Templates section
        self._chk_tpls = QCheckBox("Email Templates")
        self._chk_tpls.setChecked(True)
        lay.addWidget(self._chk_tpls)

        # Template list — always visible, disabled when checkbox unchecked
        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self._tpl_list = QListWidget()
        self._tpl_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for name in tpl_names:
            item = QListWidgetItem(name)
            self._tpl_list.addItem(item)
            item.setSelected(True)
        self._tpl_list.setFixedHeight(min(max(len(tpl_names), 2), 8) * 28 + 4)
        lay.addWidget(self._tpl_list)

        def _toggle_tpl_list(enabled: bool):
            self._tpl_list.setEnabled(enabled)
            if enabled:
                for i in range(self._tpl_list.count()):
                    self._tpl_list.item(i).setSelected(True)
            else:
                self._tpl_list.clearSelection()
        self._chk_tpls.toggled.connect(_toggle_tpl_list)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        cancel = _btn("Cancel", "outline", height=34)
        cancel.clicked.connect(self.reject)
        ok = _btn("Export", "primary", height=34)
        ok.clicked.connect(self._on_ok)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(ok)
        lay.addLayout(btn_row)
        self.adjustSize()

    def _on_ok(self):
        if not self._chk_smtp.isChecked() and not self._chk_tpls.isChecked():
            QMessageBox.warning(self, "Nothing Selected", "Please select at least one item to export.")
            return
        if self._chk_tpls.isChecked() and not self._tpl_list.selectedItems():
            QMessageBox.warning(self, "No Templates", "Please select at least one template.")
            return
        self.accept()

    def result_choices(self) -> tuple[bool, bool, set[str]]:
        include_smtp = self._chk_smtp.isChecked()
        include_tpls = self._chk_tpls.isChecked()
        selected = {item.text() for item in self._tpl_list.selectedItems()}
        return include_smtp, include_tpls, selected


class _SelectiveImportDialog(QDialog):
    """Choose what to apply from an email-settings import file."""

    def __init__(self, parent, has_smtp: bool, tpl_names: list[str]):
        super().__init__(parent)
        self.setWindowTitle("Import Email Settings")
        self.setModal(True)
        self.setMinimumWidth(380)
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum)
        lay = QVBoxLayout(self)
        lay.setSpacing(10)
        lay.setContentsMargins(20, 16, 20, 16)

        lay.addWidget(_lbl("Select what to import:", bold=True))

        self._chk_smtp = QCheckBox("SMTP / Sender Settings")
        self._chk_smtp.setChecked(has_smtp)
        self._chk_smtp.setEnabled(has_smtp)
        if not has_smtp:
            self._chk_smtp.setText("SMTP / Sender Settings (not in file)")
        lay.addWidget(self._chk_smtp)

        has_tpls = bool(tpl_names)
        self._chk_tpls = QCheckBox("Email Templates")
        self._chk_tpls.setChecked(has_tpls)
        self._chk_tpls.setEnabled(has_tpls)
        if not has_tpls:
            self._chk_tpls.setText("Email Templates (not in file)")
        lay.addWidget(self._chk_tpls)

        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self._tpl_list = QListWidget()
        self._tpl_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        for name in tpl_names:
            item = QListWidgetItem(name)
            self._tpl_list.addItem(item)
            item.setSelected(True)
        self._tpl_list.setFixedHeight(min(max(len(tpl_names), 2), 8) * 28 + 4)
        self._tpl_list.setEnabled(has_tpls)
        lay.addWidget(self._tpl_list)

        if has_tpls:
            def _toggle_tpl_list(enabled: bool):
                self._tpl_list.setEnabled(enabled)
                if enabled:
                    for i in range(self._tpl_list.count()):
                        self._tpl_list.item(i).setSelected(True)
                else:
                    self._tpl_list.clearSelection()
            self._chk_tpls.toggled.connect(_toggle_tpl_list)

        note = _lbl("Existing templates with the same name will be overwritten.")
        note.setWordWrap(True)
        lay.addWidget(note)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        cancel = _btn("Cancel", "outline", height=34)
        cancel.clicked.connect(self.reject)
        ok = _btn("Import", "primary", height=34)
        ok.clicked.connect(self._on_ok)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(ok)
        lay.addLayout(btn_row)
        self.adjustSize()

    def _on_ok(self):
        if not self._chk_smtp.isChecked() and not self._chk_tpls.isChecked():
            QMessageBox.warning(self, "Nothing Selected", "Please select at least one item to import.")
            return
        if self._chk_tpls.isChecked() and self._tpl_list.isEnabled() and not self._tpl_list.selectedItems():
            QMessageBox.warning(self, "No Templates", "Please select at least one template.")
            return
        self.accept()

    def result_choices(self) -> tuple[bool, bool, set[str]]:
        import_smtp = self._chk_smtp.isChecked()
        import_tpls = self._chk_tpls.isChecked()
        selected = {item.text() for item in self._tpl_list.selectedItems()}
        return import_smtp, import_tpls, selected


# ── SMTP Settings Dialog ──────────────────────────────────────────────────────

class SmtpSettingsDialog(QDialog):
    """Configure SMTP credentials, firm name, BCC addresses, and email templates."""

    _test_result = pyqtSignal(bool, str)   # (success, message)

    def __init__(self, parent, vault):
        super().__init__(parent)
        self._vault = vault
        self._tile_btns: dict[str, QPushButton] = {}
        self._selected_preset: str | None = None
        self._test_result.connect(self._on_test_result)
        self.setWindowTitle("Email Settings")
        self.setMinimumSize(1020, 620)
        self.setModal(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowSystemMenuHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        self._build_ui()

    # ── build ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        t = _t()
        field_ss = (
            f"QLineEdit{{border:1px solid {t.border};border-radius:6px;padding:6px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QLineEdit:focus{{border-color:{t.border_focus};background:{t.bg_input_focus};}}"
        )
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;font-size:12px;}}"
            + field_ss +
            f"QTextEdit{{border:1px solid {t.border};border-radius:6px;padding:8px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QTextEdit:focus{{border-color:{t.border_focus};}}"
            f"QSpinBox{{border:1px solid {t.border};border-radius:6px;padding:4px 8px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox{{border:1px solid {t.border};border-radius:6px;padding:4px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox::drop-down{{border:none;width:20px;}}"
            f"QComboBox::down-arrow{{image:url(none);width:0;height:0;"
            f"border-left:4px solid transparent;border-right:4px solid transparent;"
            f"border-top:5px solid {t.text_primary};}}"
            f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
            f"selection-background-color:{t.accent};}}"
        )

        cfg = self._vault.get_email_settings()

        # ── outer: header bar (fixed) + tab widget (stretchy) + footer bar (fixed)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # ── Header bar ────────────────────────────────────────────────────────
        hdr_widget = QWidget()
        hdr_widget.setStyleSheet(
            f"QWidget{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};}}"
        )
        hdr_lay = QVBoxLayout(hdr_widget)
        hdr_lay.setContentsMargins(16, 14, 16, 12)
        hdr_lay.setSpacing(0)

        title_row = QHBoxLayout()
        title_row.addWidget(_lbl("Email Settings", 15, bold=True))
        title_row.addStretch()
        help_btn = QPushButton("? Help")
        help_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        help_btn.setStyleSheet(
            f"QPushButton{{background:transparent;color:{t.accent};border:none;"
            f"font-size:12px;font-weight:bold;padding:2px 6px;}}"
            f"QPushButton:hover{{text-decoration:underline;}}"
        )
        help_btn.clicked.connect(self._show_help)
        title_row.addWidget(help_btn)
        hdr_lay.addLayout(title_row)
        hdr_lay.addSpacing(3)
        hdr_lay.addWidget(_lbl(
            "Configure SMTP server and email template for mailing tax documents to clients.",
            10, color=t.text_muted))

        sep_top = QFrame()
        sep_top.setFrameShape(QFrame.Shape.HLine)
        sep_top.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")

        outer.addWidget(hdr_widget)
        outer.addWidget(sep_top)

        # ── Shared scroll area stylesheet ─────────────────────────────────────
        _scroll_ss = (
            f"QScrollArea{{background:{t.bg_window};border:none;}}"
            f"QScrollBar:vertical{{width:8px;background:{t.bg_table_alt};}}"
            f"QScrollBar::handle:vertical{{background:{t.border};border-radius:4px;min-height:24px;}}"
            f"QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{{height:0px;}}"
        )

        # ── Tab widget ────────────────────────────────────────────────────────
        tab = QTabWidget()
        tab.setStyleSheet(
            f"QTabWidget::pane{{border:1px solid {t.border};border-top:none;}}"
            f"QTabBar::tab{{background:{t.bg_table_alt};color:{t.text_muted};"
            f"padding:8px 20px;border:1px solid {t.border};border-bottom:none;"
            f"border-radius:6px 6px 0 0;margin-right:2px;font-size:12px;}}"
            f"QTabBar::tab:selected{{background:{t.accent};color:white;"
            f"border-color:{t.accent};font-weight:600;}}"
            f"QTabBar::tab:hover:!selected{{background:{t.bg_input};}}"
        )
        outer.addWidget(tab, stretch=1)

        # ── Tab 1 — SMTP / Sender ─────────────────────────────────────────────
        tab1_scroll = QScrollArea()
        tab1_scroll.setWidgetResizable(True)
        tab1_scroll.setFrameShape(QFrame.Shape.NoFrame)
        tab1_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        tab1_scroll.setStyleSheet(_scroll_ss)
        tab1_content = QWidget()
        tab1_content.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        tab1_main = QVBoxLayout(tab1_content)
        tab1_main.setContentsMargins(16, 16, 16, 16)
        tab1_main.setSpacing(0)
        tab1_scroll.setWidget(tab1_content)
        # ── Tab 1 — Templates ─────────────────────────────────────────────────
        tab2_widget = QWidget()
        tab2_widget.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        tab2_outer = QHBoxLayout(tab2_widget)
        tab2_outer.setContentsMargins(0, 0, 0, 0)
        tab2_outer.setSpacing(0)
        _tpl_icon_p = _icon_path("menu_template.png")
        _tpl_icon = QIcon(QPixmap(_tpl_icon_p).scaled(
            16, 16, Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation)) if _tpl_icon_p else QIcon()
        tab.addTab(tab2_widget, _tpl_icon, "Templates")

        # ── Tab 2 — SMTP / Sender ─────────────────────────────────────────────
        tab.addTab(tab1_scroll, "SMTP / Sender")

        def _flbl(text):
            l = QLabel(text)
            l.setStyleSheet(
                f"font-size:11px;font-weight:600;color:{t.text_muted};margin-bottom:3px;")
            return l

        # ─────────────────────────────────────────────────────────────────────
        # Tab 1 content
        # ─────────────────────────────────────────────────────────────────────

        # Provider picker
        tab1_main.addWidget(_lbl("Select your email provider", 12, bold=True))
        tab1_main.addSpacing(10)

        tile_row = QHBoxLayout()
        tile_row.setSpacing(8)
        for preset in _SMTP_PRESETS:
            btn = self._make_tile(preset)
            self._tile_btns[preset["name"]] = btn
            tile_row.addWidget(btn)
        tile_row.addStretch()
        tab1_main.addLayout(tile_row)
        tab1_main.addSpacing(10)

        self._help_note = QLabel("")
        self._help_note.setWordWrap(True)
        self._help_note.setOpenExternalLinks(True)
        self._help_note.setTextFormat(Qt.TextFormat.RichText)
        self._help_note.setStyleSheet(
            f"background:#EFF6FF;color:#1E3A5F;"
            f"border-left:4px solid #2563EB;border-top:1px solid #BFDBFE;"
            f"border-right:1px solid #BFDBFE;border-bottom:1px solid #BFDBFE;"
            f"border-radius:0 6px 6px 0;padding:10px 14px;"
            f"font-size:11px;line-height:160%;")
        self._help_note.hide()
        tab1_main.addWidget(self._help_note)
        tab1_main.addSpacing(20)

        # SMTP Server / Port / Encryption
        row_host = QHBoxLayout(); row_host.setSpacing(10)

        host_col = QVBoxLayout(); host_col.setSpacing(4)
        host_col.addWidget(_flbl("SMTP Server"))
        self._host = QLineEdit(cfg.get("smtp_host", ""))
        self._host.setPlaceholderText("smtp.gmail.com")
        self._host.setFixedHeight(34)
        host_col.addWidget(self._host)
        row_host.addLayout(host_col, stretch=1)

        port_col = QVBoxLayout(); port_col.setSpacing(4)
        port_col.addWidget(_flbl("Port"))
        self._port = QSpinBox()
        self._port.setRange(1, 65535)
        self._port.setValue(int(cfg.get("smtp_port", 587)))
        self._port.setFixedHeight(34)
        self._port.setFixedWidth(90)
        port_col.addWidget(self._port)
        row_host.addLayout(port_col)

        enc_col = QVBoxLayout(); enc_col.setSpacing(4)
        enc_col.addWidget(_flbl("Encryption"))
        self._enc = QComboBox()
        self._enc.addItems(["STARTTLS", "SSL/TLS", "None"])
        saved_enc = cfg.get("smtp_encryption", "STARTTLS")
        self._enc.setCurrentIndex(max(0, self._enc.findText(saved_enc)))
        self._enc.setFixedHeight(34)
        self._enc.setFixedWidth(110)
        enc_col.addWidget(self._enc)
        row_host.addLayout(enc_col)
        tab1_main.addLayout(row_host)
        tab1_main.addSpacing(12)

        # Username / Password
        user_pwd_row = QHBoxLayout(); user_pwd_row.setSpacing(12)

        user_col = QVBoxLayout(); user_col.setSpacing(4)
        user_col.addWidget(_flbl("Username / Email"))
        self._user = QLineEdit(cfg.get("smtp_user", ""))
        self._user.setPlaceholderText("you@gmail.com")
        self._user.setFixedHeight(34)
        user_col.addWidget(self._user)
        user_pwd_row.addLayout(user_col, stretch=35)

        pwd_col = QVBoxLayout(); pwd_col.setSpacing(4)
        pwd_col.addWidget(_flbl("Password / App Password"))
        pwd_field_row = QHBoxLayout(); pwd_field_row.setSpacing(8)
        self._pwd = QLineEdit(cfg.get("smtp_password", ""))
        self._pwd.setEchoMode(QLineEdit.EchoMode.Password)
        self._pwd.setPlaceholderText("Enter password")
        self._pwd.setFixedHeight(34)
        pwd_field_row.addWidget(self._pwd)
        show_pwd_cb = QCheckBox("Show password")
        show_pwd_cb.setStyleSheet(
            f"QCheckBox{{font-size:11px;color:{t.text_muted};}}"
            f"QCheckBox::indicator{{width:14px;height:14px;}}"
        )
        show_pwd_cb.toggled.connect(
            lambda on: self._pwd.setEchoMode(
                QLineEdit.EchoMode.Normal if on else QLineEdit.EchoMode.Password))
        pwd_field_row.addWidget(show_pwd_cb)
        pwd_col.addLayout(pwd_field_row)
        user_pwd_row.addLayout(pwd_col, stretch=25)

        tab1_main.addLayout(user_pwd_row)
        tab1_main.addSpacing(10)

        # Send As / From + BCC
        from_bcc_row = QHBoxLayout(); from_bcc_row.setSpacing(12)

        from_col = QVBoxLayout(); from_col.setSpacing(4)
        from_col.addWidget(_flbl("Send As / From Address (optional — leave blank to use username)"))
        self._from = QLineEdit(cfg.get("smtp_from", ""))
        self._from.setPlaceholderText("income-tax@daksm.com")
        self._from.setFixedHeight(34)
        from_col.addWidget(self._from)
        from_bcc_row.addLayout(from_col, stretch=1)

        bcc_col = QVBoxLayout(); bcc_col.setSpacing(4)
        bcc_col.addWidget(_flbl("BCC Addresses  (separate multiple with ;)"))
        self._bcc = QLineEdit(cfg.get("bcc_addresses", ""))
        self._bcc.setPlaceholderText("partner@firm.com;team@firm.com")
        self._bcc.setFixedHeight(34)
        bcc_col.addWidget(self._bcc)
        from_bcc_row.addLayout(bcc_col, stretch=1)

        tab1_main.addLayout(from_bcc_row)
        tab1_main.addSpacing(14)

        tab1_main.addWidget(_flbl("Firm Name  (used in {firm_name} placeholder)"))
        tab1_main.addSpacing(3)
        self._firm = QLineEdit(cfg.get("firm_name", ""))
        self._firm.setPlaceholderText("AI Learrning Guru")
        self._firm.setFixedHeight(34)
        tab1_main.addWidget(self._firm)
        tab1_main.addStretch()

        # ─────────────────────────────────────────────────────────────────────
        # Tab 2 content
        # ─────────────────────────────────────────────────────────────────────

        _ph_chip_ss = (
            f"QPushButton{{background:{t.bg_table_alt};color:{t.accent};"
            f"border:1px solid {t.border};border-radius:5px;"
            f"font-size:10px;padding:0 6px;font-family:monospace;}}"
            f"QPushButton:hover{{background:{t.bg_input};border-color:{t.accent};}}"
        )
        _combo_ss = (
            f"QComboBox{{border:1px solid {t.border};border-radius:5px;padding:2px 6px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox::drop-down{{border:none;width:16px;}}"
            f"QComboBox::down-arrow{{image:url(none);width:0;height:0;"
            f"border-left:4px solid transparent;border-right:4px solid transparent;"
            f"border-top:5px solid {t.text_primary};}}"
            f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
            f"selection-background-color:{t.accent};}}"
        )

        # ── Left panel — template list ────────────────────────────────────────
        left_panel = QWidget()
        left_panel.setFixedWidth(190)
        left_panel.setStyleSheet(
            f"QWidget{{background:{t.bg_panel};border-right:1px solid {t.border};}}")
        left_v = QVBoxLayout(left_panel)
        left_v.setContentsMargins(10, 12, 10, 10)
        left_v.setSpacing(6)

        left_v.addWidget(_lbl("Templates", 12, bold=True))
        left_v.addSpacing(4)

        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self._tpl_list = QListWidget()
        self._tpl_list.setStyleSheet(
            f"QListWidget{{background:{t.bg_panel};border:none;outline:none;font-size:12px;}}"
            f"QListWidget::item{{padding:8px 10px;border-radius:5px;color:{t.text_primary};}}"
            f"QListWidget::item:selected{{background:{t.accent};color:white;}}"
            f"QListWidget::item:hover:!selected{{background:{t.bg_table_alt};}}"
        )
        self._tpl_list.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self._tpl_list.itemChanged.connect(self._on_tpl_renamed)
        self._tpl_list.currentRowChanged.connect(self._on_tpl_select)
        left_v.addWidget(self._tpl_list, stretch=1)

        hint_lbl = QLabel("Double-click to rename")
        hint_lbl.setStyleSheet(f"color:{t.text_muted};font-size:10px;background:transparent;")
        left_v.addWidget(hint_lbl)

        self._tpl_default_btn = _btn("★ Set as Default", "outline", height=28)
        self._tpl_default_btn.clicked.connect(self._tpl_set_default)
        left_v.addWidget(self._tpl_default_btn)

        btn_row = QHBoxLayout(); btn_row.setSpacing(6)
        self._tpl_add_btn = _btn("+ New", "outline", height=28, icon="btn_add_list.png")
        self._tpl_add_btn.clicked.connect(self._tpl_add)
        self._tpl_del_btn = _btn("Delete", "danger", height=28, icon="btn_delete.png")
        self._tpl_del_btn.clicked.connect(self._tpl_delete)
        btn_row.addWidget(self._tpl_add_btn)
        btn_row.addWidget(self._tpl_del_btn)
        left_v.addLayout(btn_row)

        tab2_outer.addWidget(left_panel)

        # ── Right panel — header + editor ─────────────────────────────────────
        right_container = QWidget()
        right_container.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        right_container_v = QVBoxLayout(right_container)
        right_container_v.setContentsMargins(0, 0, 0, 0)
        right_container_v.setSpacing(0)

        # Header strip showing current template name + default badge
        self._tpl_header = QWidget()
        self._tpl_header.setStyleSheet(
            f"QWidget{{background:{t.bg_table_alt};border-bottom:1px solid {t.border};}}")
        _hdr_lay = QHBoxLayout(self._tpl_header)
        _hdr_lay.setContentsMargins(16, 10, 16, 10)
        self._tpl_header_lbl = QLabel("")
        self._tpl_header_lbl.setStyleSheet(
            f"font-size:13px;font-weight:bold;color:{t.text_primary};background:transparent;")
        self._tpl_default_lbl = QLabel("Default")
        self._tpl_default_lbl.setStyleSheet(
            f"font-size:10px;color:{t.accent};background:transparent;"
            f"border:1px solid {t.accent};border-radius:4px;padding:1px 6px;")
        self._tpl_default_lbl.hide()
        _hdr_lay.addWidget(self._tpl_header_lbl)
        _hdr_lay.addSpacing(8)
        _hdr_lay.addWidget(self._tpl_default_lbl)
        _hdr_lay.addStretch()
        right_container_v.addWidget(self._tpl_header)

        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setFrameShape(QFrame.Shape.NoFrame)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        right_scroll.setStyleSheet(_scroll_ss)
        right_content = QWidget()
        right_content.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        right_v = QVBoxLayout(right_content)
        right_v.setContentsMargins(16, 14, 16, 14)
        right_v.setSpacing(0)
        right_scroll.setWidget(right_content)
        right_container_v.addWidget(right_scroll, stretch=1)
        tab2_outer.addWidget(right_container, stretch=1)

        # Template name — stored internally, edited via list item directly
        self._tpl_name = QLineEdit()
        self._tpl_name.hide()

        # Subject
        subj_lbl_row = QHBoxLayout(); subj_lbl_row.setSpacing(6)
        subj_lbl_row.addWidget(_flbl("Subject"))
        subj_lbl_row.addStretch()
        for ph in ["{client_name}", "{ay}", "{firm_name}"]:
            pb = QPushButton(ph)
            pb.setFixedHeight(22)
            pb.setCursor(Qt.CursorShape.PointingHandCursor)
            pb.setStyleSheet(_ph_chip_ss)
            pb.clicked.connect(lambda _, p=ph: self._subj.insert(p))
            subj_lbl_row.addWidget(pb)
        right_v.addLayout(subj_lbl_row)
        right_v.addSpacing(3)
        self._subj = QLineEdit()
        self._subj.setFixedHeight(34)
        right_v.addWidget(self._subj)
        right_v.addSpacing(14)

        # Body
        right_v.addWidget(_flbl("Body"))
        right_v.addSpacing(4)

        _fmt_btn_ss = (
            f"QPushButton{{background:{t.bg_table_alt};color:{t.text_primary};"
            f"border:1px solid {t.border};border-radius:5px;"
            f"font-size:12px;padding:0 8px;min-width:28px;height:28px;}}"
            f"QPushButton:hover{{background:{t.bg_input};border-color:{t.accent};}}"
            f"QPushButton:checked{{background:{t.accent};color:#fff;border-color:{t.accent};}}"
        )

        def _fmt_toggle(label: str, prop: str) -> QPushButton:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(28)
            btn.setStyleSheet(_fmt_btn_ss)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            def _apply(checked, p=prop):
                fmt = QTextCharFormat()
                if p == "bold":
                    fmt.setFontWeight(QFont.Weight.Bold if checked else QFont.Weight.Normal)
                elif p == "italic":
                    fmt.setFontItalic(checked)
                elif p == "underline":
                    fmt.setFontUnderline(checked)
                self._body.mergeCurrentCharFormat(fmt)
                self._body.setFocus()
            btn.toggled.connect(_apply)
            return btn

        fmt_bar = QHBoxLayout(); fmt_bar.setSpacing(6)
        self._font_combo = QFontComboBox()
        self._font_combo.setEditable(False)
        self._font_combo.setFixedHeight(26)
        self._font_combo.setFixedWidth(140)
        self._font_combo.setStyleSheet(_combo_ss)
        self._font_combo.setCurrentFont(QFont(_UI_FONT))
        self._font_combo.currentFontChanged.connect(lambda f: (
            self._body.setFocus(),
            self._body.mergeCurrentCharFormat(
                (lambda fmt: (fmt.setFontFamilies([f.family()]), fmt))
                (QTextCharFormat())[1]
            )
        ))
        fmt_bar.addWidget(self._font_combo)

        self._size_combo = QComboBox()
        self._size_combo.setFixedHeight(26)
        self._size_combo.setFixedWidth(48)
        self._size_combo.setStyleSheet(_combo_ss)
        for sz in ["8", "9", "10", "11", "12", "14", "16", "18"]:
            self._size_combo.addItem(sz)
        self._size_combo.setCurrentText("12")
        self._size_combo.currentTextChanged.connect(lambda sz: (
            self._body.setFocus(),
            self._body.mergeCurrentCharFormat(
                (lambda fmt: (fmt.setFontPointSize(float(sz)), fmt))
                (QTextCharFormat())[1]
            ) if sz.isdigit() else None
        ))
        fmt_bar.addWidget(self._size_combo)

        fmt_bar.addSpacing(2)
        self._fmt_bold      = _fmt_toggle("B", "bold")
        self._fmt_bold.setFont(QFont(_UI_FONT, 10, QFont.Weight.Bold))
        self._fmt_italic    = _fmt_toggle("I", "italic")
        self._fmt_italic.setFont(QFont(_UI_FONT, 10, -1, True))
        self._fmt_underline = _fmt_toggle("U", "underline")
        fmt_bar.addWidget(self._fmt_bold)
        fmt_bar.addWidget(self._fmt_italic)
        fmt_bar.addWidget(self._fmt_underline)
        fmt_bar.addSpacing(6)
        for ph in ["{client_name}", "{pan}", "{ay}", "{firm_name}", "{documents}"]:
            pb = QPushButton(ph)
            pb.setFixedHeight(26)
            pb.setCursor(Qt.CursorShape.PointingHandCursor)
            pb.setStyleSheet(_ph_chip_ss)
            pb.clicked.connect(lambda _, p=ph: self._body.insertPlainText(p))
            fmt_bar.addWidget(pb)
        fmt_bar.addStretch()
        right_v.addLayout(fmt_bar)
        right_v.addSpacing(4)

        self._body = QTextEdit()
        self._body.setAcceptRichText(True)
        self._body.setMinimumHeight(160)
        self._body.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._body.currentCharFormatChanged.connect(self._sync_fmt_buttons)
        right_v.addWidget(self._body, stretch=1)
        right_v.addSpacing(14)

        # Document type checkboxes
        right_v.addWidget(_flbl("Attach these document types"))
        right_v.addSpacing(6)
        _cb_ss = (
            f"QCheckBox{{color:{t.text_primary};font-size:12px;spacing:6px;}}"
            f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
            f"border-radius:3px;background:{t.bg_checkbox};}}"
            f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
        )
        self._doc_cbs = {}
        docs_grid = QHBoxLayout(); docs_grid.setSpacing(20)
        for label, key in [
            ("26AS PDF",   "26as_pdf"),
            ("26AS Excel", "26as_xlsx"),
            ("168 PDF",    "168_pdf"),
            ("168 Excel",  "168_xlsx"),
            ("AIS PDF",    "ais_pdf"),
            ("AIS Excel",  "ais_xlsx"),
            ("TIS",        "tis_pdf"),
        ]:
            cb = QCheckBox(label)
            cb.setChecked(True)
            cb.setStyleSheet(_cb_ss)
            self._doc_cbs[key] = cb
            docs_grid.addWidget(cb)
        docs_grid.addStretch()
        right_v.addLayout(docs_grid)

        # Load templates into list
        self._templates = self._vault.get_email_templates()
        self._loading_tpl = False
        for tpl in self._templates:
            item = QListWidgetItem(tpl["name"])
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self._tpl_list.addItem(item)

        # Select active template
        active_name = self._vault.get_active_template_name()
        for i, tpl in enumerate(self._templates):
            if tpl["name"] == active_name:
                self._tpl_list.setCurrentRow(i)
                break
        else:
            if self._templates:
                self._tpl_list.setCurrentRow(0)

        # Update delete button state and header
        self._tpl_del_btn.setEnabled(len(self._templates) > 1)
        self._refresh_tpl_header()

        # Auto-highlight tile if saved host matches a known preset
        saved_host = cfg.get("smtp_host", "")
        if saved_host in _HOST_TO_PRESET:
            self._highlight_tile(_HOST_TO_PRESET[saved_host])

        # ── Footer bar (fixed, outside tabs) ──────────────────────────────────
        sep_bot = QFrame()
        sep_bot.setFrameShape(QFrame.Shape.HLine)
        sep_bot.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        outer.addWidget(sep_bot)

        footer_widget = QWidget()
        footer_widget.setStyleSheet(f"QWidget{{background:{t.bg_window};}}")
        footer_lay = QHBoxLayout(footer_widget)
        footer_lay.setContentsMargins(16, 10, 16, 12)
        footer_lay.setSpacing(8)
        self._test_btn = _btn("Send Test Email", "outline", height=36, icon="btn_send_test.png")
        self._test_btn.clicked.connect(self._send_test)
        footer_lay.addWidget(self._test_btn)
        log_btn = _btn("View Log", "outline", height=36, icon="btn_view_log.png")
        export_btn = _btn("Export Settings", "outline", height=36, icon="menu_export.png")
        export_btn.setToolTip("Export SMTP settings and templates to a JSON file")
        export_btn.clicked.connect(self._export_settings)
        footer_lay.addWidget(export_btn)
        import_btn = _btn("Import Settings", "outline", height=36, icon="menu_import.png")
        import_btn.setToolTip("Import SMTP settings and templates from a JSON file")
        import_btn.clicked.connect(self._import_settings)
        footer_lay.addWidget(import_btn)
        log_btn.clicked.connect(self._open_log)
        footer_lay.addWidget(log_btn)
        footer_lay.addStretch()
        self._ctx_save_btn = _btn("Save SMTP Settings", "outline", height=36, icon="btn_save.png")
        footer_lay.addWidget(self._ctx_save_btn)
        self._save_close_btn = _btn("Cancel", "secondary", height=36, icon="btn_cancel.png")
        self._save_close_btn.clicked.connect(self._on_save_close_clicked)
        footer_lay.addWidget(self._save_close_btn)

        # Wire all fields to mark dirty (switches button to Save & Close)
        for field in (self._host, self._user, self._pwd, self._from, self._bcc, self._firm):
            field.textChanged.connect(self._mark_dirty)
        self._port.valueChanged.connect(self._mark_dirty)
        self._enc.currentIndexChanged.connect(self._mark_dirty)
        self._tpl_list.itemChanged.connect(self._mark_dirty)
        self._subj.textChanged.connect(self._mark_dirty)
        self._body.textChanged.connect(self._mark_dirty)

        def _on_tab_changed(i):
            try:
                self._ctx_save_btn.clicked.disconnect()
            except (RuntimeError, TypeError):
                pass
            if i == 0:
                self._ctx_save_btn.setText("Save Templates")
                self._ctx_save_btn.clicked.connect(self._save_templates_only)
            else:
                self._ctx_save_btn.setText("Save SMTP Settings")
                self._ctx_save_btn.clicked.connect(self._save_smtp_only)

        tab.currentChanged.connect(_on_tab_changed)
        _on_tab_changed(0)   # initialise for Templates tab
        outer.addWidget(footer_widget)

        self.resize(1100, 720)
        # Open on SMTP tab if not yet configured, else Templates
        if not cfg.get("smtp_host"):
            tab.setCurrentIndex(1)
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, lambda: tab1_scroll.verticalScrollBar().setValue(0))

    # ── tile factory ──────────────────────────────────────────────────────────

    def _make_tile(self, preset: dict) -> QPushButton:
        """Return an icon-only provider button with tooltip."""
        name = preset["name"]
        icon_text = preset["icon"]
        icon_color = preset["icon_color"]

        # Build pixmap once, cache for reuse
        if name not in _TILE_PIXMAP_CACHE:
            # Try loading PNG from resources/
            from config import _bundled_dir
            png_path = os.path.join(_bundled_dir(), "resources", "icons", preset.get("icon_file", ""))
            loaded = False
            if preset.get("icon_file") and os.path.isfile(png_path):
                src = QPixmap(png_path)
                if not src.isNull():
                    px = src.scaled(40, 40, Qt.AspectRatioMode.KeepAspectRatio,
                                    Qt.TransformationMode.SmoothTransformation)
                    _TILE_PIXMAP_CACHE[name] = px
                    loaded = True
            if not loaded:
                px = QPixmap(40, 40)
                px.fill(Qt.GlobalColor.transparent)
                painter = QPainter(px)
                painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                painter.setBrush(QBrush(QColor(icon_color)))
                painter.setPen(QPen(Qt.GlobalColor.transparent))
                painter.drawRoundedRect(0, 0, 40, 40, 8, 8)
                painter.setPen(QPen(QColor("white")))
                f = painter.font()
                f.setBold(True)
                f.setPixelSize(13 if len(icon_text) > 1 else 18)
                painter.setFont(f)
                painter.drawText(px.rect(), Qt.AlignmentFlag.AlignCenter, icon_text)
                painter.end()
                _TILE_PIXMAP_CACHE[name] = px
        px = _TILE_PIXMAP_CACHE[name]

        from PyQt6.QtGui import QIcon
        btn = QPushButton()
        btn.setFixedSize(52, 52)
        btn.setIcon(QIcon(px))
        btn.setIconSize(px.size())
        btn.setToolTip(name)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)

        self._set_tile_style(btn, selected=False)
        btn.clicked.connect(lambda _checked, n=name: self._apply_preset(n))
        return btn

    def _set_tile_style(self, btn: QPushButton, selected: bool):
        t = _t()
        if selected:
            btn.setStyleSheet(
                f"QPushButton{{background:{t.bg_input_focus};border:2px solid {t.accent};"
                f"border-radius:8px;color:{t.text_primary};}}"
                f"QPushButton:hover{{background:{t.bg_input_focus};}}"
            )
        else:
            btn.setStyleSheet(
                f"QPushButton{{background:{t.bg_table_alt};border:1px solid {t.border};"
                f"border-radius:8px;color:{t.text_primary};}}"
                f"QPushButton:hover{{background:{t.bg_input};border-color:{t.border_focus};}}"
            )

    def _highlight_tile(self, name: str):
        for n, btn in self._tile_btns.items():
            self._set_tile_style(btn, selected=(n == name))
        self._selected_preset = name

    # ── preset apply ─────────────────────────────────────────────────────────

    def _apply_preset(self, name: str):
        self._highlight_tile(name)
        preset = next((p for p in _SMTP_PRESETS if p["name"] == name), None)
        if not preset:
            return

        if preset["host"] is not None:    # None = Custom, leave fields alone
            self._host.setText(preset["host"])
        if preset["port"] is not None:
            self._port.setValue(preset["port"])
        if preset["encryption"] is not None:
            idx = self._enc.findText(preset["encryption"])
            if idx >= 0:
                self._enc.setCurrentIndex(idx)

        help_text = preset.get("help", "")
        if help_text:
            self._help_note.setText(help_text)
            self._help_note.show()
        else:
            self._help_note.hide()

    # ── format button sync ───────────────────────────────────────────────────

    def _sync_fmt_buttons(self, fmt: QTextCharFormat):
        for btn, attr in (
            (self._fmt_bold,      fmt.fontWeight() == QFont.Weight.Bold),
            (self._fmt_italic,    fmt.fontItalic()),
            (self._fmt_underline, fmt.fontUnderline()),
        ):
            btn.blockSignals(True)
            btn.setChecked(attr)
            btn.blockSignals(False)
        families = fmt.fontFamilies()
        if families:
            self._font_combo.blockSignals(True)
            self._font_combo.setCurrentFont(QFont(families[0]))
            self._font_combo.blockSignals(False)
        sz = fmt.fontPointSize()
        if sz > 0:
            self._size_combo.blockSignals(True)
            self._size_combo.setCurrentText(str(int(sz)))
            self._size_combo.blockSignals(False)

    # ── template list interactions ────────────────────────────────────────────

    def _current_tpl_dict(self) -> dict:
        """Read the right-panel editor into a template dict."""
        row = self._tpl_list.currentRow()
        item = self._tpl_list.item(row)
        name = item.text().strip() if item else "Untitled"
        return {
            "name":    name or "Untitled",
            "subject": self._subj.text().strip(),
            "body":    self._body.toHtml(),
            "docs":    {k: cb.isChecked() for k, cb in self._doc_cbs.items()},
        }

    def _on_tpl_renamed(self, item):
        """Sync list item rename back into self._templates."""
        row = self._tpl_list.row(item)
        if 0 <= row < len(self._templates):
            self._templates[row]["name"] = item.text().strip() or "Untitled"

    def _flush_current_tpl(self):
        """Save right-panel edits back into self._templates at the current row."""
        row = self._tpl_list.currentRow()
        if row < 0 or row >= len(self._templates):
            return
        self._templates[row] = self._current_tpl_dict()
        self._tpl_list.item(row).setText(self._templates[row]["name"])

    def _load_tpl_into_editor(self, tpl: dict):
        """Populate right-panel fields from a template dict."""
        self._loading_tpl = True
        self._subj.setText(tpl.get("subject", ""))
        raw_body = tpl.get("body", "")
        if raw_body.lstrip().startswith("<"):
            self._body.setHtml(raw_body)
        else:
            self._body.setPlainText(raw_body)
        docs = tpl.get("docs", {})
        for key, cb in self._doc_cbs.items():
            cb.setChecked(docs.get(key, True))
        self._loading_tpl = False

    def _on_tpl_select(self, row: int):
        if self._loading_tpl or row < 0 or row >= len(self._templates):
            return
        self._load_tpl_into_editor(self._templates[row])
        self._tpl_del_btn.setEnabled(len(self._templates) > 1)
        self._refresh_tpl_header()

    def _refresh_tpl_header(self):
        row = self._tpl_list.currentRow()
        if row < 0 or row >= len(self._templates):
            return
        name = self._templates[row]["name"]
        active_name = self._vault.get_active_template_name()
        self._tpl_header_lbl.setText(name)
        is_default = (name == active_name)
        self._tpl_default_lbl.setVisible(is_default)
        self._tpl_default_btn.setEnabled(not is_default)

    def _tpl_set_default(self):
        self._flush_current_tpl()
        row = self._tpl_list.currentRow()
        if row < 0 or row >= len(self._templates):
            return
        self._vault.set_active_template(self._templates[row]["name"])
        self._refresh_tpl_header()

    def _tpl_add(self):
        self._flush_current_tpl()

        # Ask: blank or copy from existing?
        dlg = QDialog(self)
        dlg.setWindowTitle("New Template")
        dlg.setFixedWidth(360)
        dlg.setModal(True)
        t = _t()
        dlg.setStyleSheet(f"QDialog{{background:{t.bg_window};}}"
                          f"QLabel{{color:{t.text_primary};background:transparent;}}"
                          f"QRadioButton{{color:{t.text_primary};font-size:12px;spacing:6px;}}"
                          f"QRadioButton::indicator{{width:14px;height:14px;border:1.5px solid {t.border};"
                          f"border-radius:7px;background:{t.bg_checkbox};}}"
                          f"QRadioButton::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
                          f"QComboBox{{border:1px solid {t.border};border-radius:5px;padding:4px 8px;"
                          f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
                          f"QComboBox::drop-down{{border:none;width:18px;}}"
                          f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
                          f"selection-background-color:{t.accent};}}")
        v = QVBoxLayout(dlg)
        v.setContentsMargins(20, 16, 20, 16)
        v.setSpacing(10)
        v.addWidget(_lbl("Create new template from:", 12, bold=True))

        rb_blank = QRadioButton("Start blank")
        rb_blank.setChecked(True)
        rb_copy  = QRadioButton("Copy from existing template:")
        v.addWidget(rb_blank)

        copy_row = QHBoxLayout()
        copy_row.addWidget(rb_copy)
        copy_combo = QComboBox()
        for tpl in self._templates:
            copy_combo.addItem(tpl["name"])
        copy_combo.setEnabled(False)
        rb_blank.toggled.connect(lambda checked: copy_combo.setEnabled(not checked))
        copy_row.addWidget(copy_combo, stretch=1)
        v.addLayout(copy_row)

        v.addSpacing(4)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel_btn = _btn("Cancel", "secondary", height=32)
        cancel_btn.clicked.connect(dlg.reject)
        ok_btn = _btn("Create", "primary", height=32)
        ok_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(ok_btn)
        v.addLayout(btn_row)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        if rb_copy.isChecked():
            src_name = copy_combo.currentText()
            src = next((t for t in self._templates if t["name"] == src_name), None)
            new_tpl = {
                "name":    f"Copy of {src_name}",
                "subject": src.get("subject", "") if src else "",
                "body":    src.get("body", "") if src else "",
                "docs":    dict(src.get("docs", {})) if src else {k: True for k in self._doc_cbs},
            }
        else:
            new_tpl = {
                "name":    "New Template",
                "subject": "",
                "body":    "",
                "docs":    {k: True for k in self._doc_cbs},
            }

        self._templates.append(new_tpl)
        from PyQt6.QtWidgets import QListWidgetItem
        item = QListWidgetItem(new_tpl["name"])
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        self._tpl_list.addItem(item)
        new_row = len(self._templates) - 1
        self._tpl_list.setCurrentRow(new_row)
        self._tpl_list.editItem(self._tpl_list.item(new_row))
        self._tpl_del_btn.setEnabled(True)

    def _tpl_delete(self):
        row = self._tpl_list.currentRow()
        if row < 0 or len(self._templates) <= 1:
            return
        self._templates.pop(row)
        self._tpl_list.takeItem(row)
        new_row = min(row, len(self._templates) - 1)
        self._tpl_list.setCurrentRow(new_row)
        self._tpl_del_btn.setEnabled(len(self._templates) > 1)

    def _export_settings(self):
        """Export SMTP settings and/or selected templates to a JSON file."""
        tpl_names = [t["name"] for t in self._templates]
        dlg = _SelectiveExportDialog(self, tpl_names)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        include_smtp, include_tpls, selected_tpls = dlg.result_choices()

        data: dict = {}

        if include_smtp:
            cfg = self._collect()
            raw_pwd = cfg.pop("smtp_password", "") or ""
            if raw_pwd:
                try:
                    from cryptography.fernet import Fernet
                    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
                    from cryptography.hazmat.primitives import hashes
                    from base64 import urlsafe_b64encode
                    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=b"AayDocCapio-export-v1", iterations=100000)
                    key = urlsafe_b64encode(kdf.derive(b"AayDocCapio-portable-export-key"))
                    cfg["smtp_password_exported"] = Fernet(key).encrypt(raw_pwd.encode()).decode()
                except Exception:
                    pass
            data["smtp"] = cfg

        if include_tpls:
            tpls_to_export = [t for t in self._templates if t["name"] in selected_tpls]
            data["templates"] = tpls_to_export
            active = self._vault.get_active_template_name()
            if active in selected_tpls:
                data["active_template"] = active

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Email Settings", "AayDocCapio_EmailSettings.json",
            "JSON Files (*.json)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "Exported", "Settings exported successfully.")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", str(e))

    def _import_settings(self):
        """Import SMTP settings and/or templates from a JSON file (selective)."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Email Settings", "", "JSON Files (*.json)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            QMessageBox.warning(self, "Import Failed", f"Could not read file:\n{e}")
            return

        has_smtp = bool(data.get("smtp"))
        file_tpls = data.get("templates", [])
        file_tpl_names = [t["name"] for t in file_tpls]

        dlg = _SelectiveImportDialog(self, has_smtp, file_tpl_names)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        import_smtp, import_tpls, selected_tpls = dlg.result_choices()

        if import_smtp and has_smtp:
            smtp = data["smtp"]
            if smtp.get("smtp_host"): self._host.setText(smtp["smtp_host"])
            if smtp.get("smtp_port"): self._port.setValue(int(smtp["smtp_port"]))
            if smtp.get("smtp_user"): self._user.setText(smtp["smtp_user"])
            if smtp.get("smtp_from"): self._from.setText(smtp["smtp_from"])
            if smtp.get("bcc_addresses"): self._bcc.setText(smtp["bcc_addresses"])
            if smtp.get("firm_name"): self._firm.setText(smtp["firm_name"])
            enc = smtp.get("smtp_encryption", "")
            if enc:
                idx = self._enc.findText(enc)
                if idx >= 0: self._enc.setCurrentIndex(idx)
            if smtp.get("smtp_password_exported"):
                try:
                    from cryptography.fernet import Fernet
                    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
                    from cryptography.hazmat.primitives import hashes
                    from base64 import urlsafe_b64encode
                    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=b"AayDocCapio-export-v1", iterations=100000)
                    key = urlsafe_b64encode(kdf.derive(b"AayDocCapio-portable-export-key"))
                    pwd = Fernet(key).decrypt(smtp["smtp_password_exported"].encode()).decode()
                    self._pwd.setText(pwd)
                except Exception:
                    pass

        if import_tpls and file_tpls:
            from PyQt6.QtWidgets import QListWidgetItem
            tpls_to_import = [t for t in file_tpls if t["name"] in selected_tpls]
            existing_names = {t["name"] for t in self._templates}
            for tpl in tpls_to_import:
                if tpl["name"] in existing_names:
                    # overwrite existing template with same name
                    for i, et in enumerate(self._templates):
                        if et["name"] == tpl["name"]:
                            self._templates[i] = tpl
                            break
                else:
                    self._templates.append(tpl)
            self._tpl_list.clear()
            for tpl in self._templates:
                item = QListWidgetItem(tpl["name"])
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                self._tpl_list.addItem(item)
            active = data.get("active_template", "")
            matched = False
            for i, tpl in enumerate(self._templates):
                if tpl["name"] == active:
                    self._tpl_list.setCurrentRow(i)
                    matched = True
                    break
            if not matched:
                self._tpl_list.setCurrentRow(0)

        QMessageBox.information(self, "Imported",
            "Settings imported successfully. Please verify and click Save.")

    def _mark_dirty(self):
        """Switch the Cancel button to Save & Close once any change is made."""
        self._save_close_btn.setText("Save && Close")
        self._save_close_btn.setProperty("style_variant", "primary")
        # Re-apply stylesheet via _btn helper logic — simplest is to swap icon/style directly
        t = _t()
        self._save_close_btn.setStyleSheet(
            f"QPushButton{{background:{t.accent};color:#FFFFFF;border:none;"
            f"border-radius:8px;font-size:13px;font-weight:600;padding:0 18px;}}"
            f"QPushButton:hover{{background:{t.accent_hover};}}")

    def _on_save_close_clicked(self):
        if self._save_close_btn.text() == "Cancel":
            self.reject()
        else:
            self._save()

    def _save_smtp_only(self):
        """Save SMTP/Sender settings without closing."""
        cfg = self._collect()
        if not cfg["smtp_host"]:
            QMessageBox.warning(self, "Missing Field", "Please enter an SMTP server address.")
            return
        if not cfg["smtp_user"]:
            QMessageBox.warning(self, "Missing Field", "Please enter a username / email.")
            return
        self._vault.save_email_settings(cfg)
        QMessageBox.information(self, "Saved", "SMTP settings saved.")

    def _save_templates_only(self):
        """Flush and persist templates without closing the dialog."""
        self._flush_current_tpl()
        names = [t["name"] for t in self._templates]
        if len(names) != len(set(names)):
            QMessageBox.warning(self, "Duplicate Template Name",
                                "Each template must have a unique name.")
            return
        self._vault.save_email_templates(self._templates)
        row = self._tpl_list.currentRow()
        if 0 <= row < len(self._templates):
            self._vault.set_active_template(self._templates[row]["name"])
        QMessageBox.information(self, "Saved", "Templates saved successfully.")

    # ── collect / save / test ─────────────────────────────────────────────────

    def _collect(self) -> dict:
        enc = self._enc.currentText()
        return {
            "smtp_host":       self._host.text().strip(),
            "smtp_port":       str(self._port.value()),
            "smtp_user":       self._user.text().strip(),
            "smtp_from":       self._from.text().strip(),
            "smtp_password":   self._pwd.text(),
            "smtp_encryption": enc,
            "smtp_use_tls":    enc == "STARTTLS",
            "firm_name":       self._firm.text().strip(),
            "bcc_addresses":   self._bcc.text().strip(),
        }

    def _save(self):
        cfg = self._collect()
        if not cfg["smtp_host"]:
            QMessageBox.warning(self, "Missing Field", "Please enter an SMTP server address.")
            return
        if not cfg["smtp_user"]:
            QMessageBox.warning(self, "Missing Field", "Please enter a username / email.")
            return
        # Flush current template editor into list
        self._flush_current_tpl()
        # Validate template names are unique and non-empty
        names = [t["name"] for t in self._templates]
        if len(names) != len(set(names)):
            QMessageBox.warning(self, "Duplicate Template Name",
                                "Each template must have a unique name.")
            return
        # Save SMTP settings
        # Also sync active template's subject/body into legacy keys for emailer compat
        active = self._templates[self._tpl_list.currentRow()] if self._templates else {}
        cfg["email_subject_tpl"] = active.get("subject", "")
        cfg["email_body_tpl"]    = active.get("body", "")
        self._vault.save_email_settings(cfg)
        self._vault.save_email_templates(self._templates)
        active_name = active.get("name", "")
        if active_name:
            self._vault.set_active_template(active_name)
        self.accept()

    def _send_test(self):
        cfg = self._collect()
        to = cfg["smtp_user"]
        if not cfg["smtp_host"] or not to:
            QMessageBox.warning(self, "Incomplete",
                                "Fill in the server address and username before sending a test.")
            return

        # Disable button and show spinner while connecting
        self._test_btn.setEnabled(False)
        self._test_btn.setText("⏳ Connecting…")

        def _worker():
            from automation.emailer import send_email
            try:
                send_email(cfg, to,
                           subject="AayDocCapio — Test Email",
                           body="This is a test email from AayDocCapio.\n\n"
                                "If you received this, your SMTP settings are working correctly.",
                           attachments=[])
                self._test_result.emit(True, f"Test email sent to {to}.\nCheck your inbox.")
            except Exception as e:
                from automation.emailer import friendly_smtp_error
                self._test_result.emit(False, friendly_smtp_error(e))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_test_result(self, success: bool, message: str):
        self._test_btn.setEnabled(True)
        self._test_btn.setText("Send Test Email")
        if success:
            QMessageBox.information(self, "Test Sent", message)
        else:
            QMessageBox.critical(self, "Send Failed", message)

    def _show_help(self):
        from ui.smtp_help import _write_smtp_help_html
        import webbrowser
        path = _write_smtp_help_html()
        webbrowser.open("file:///" + path.replace(os.sep, "/"))

    def _open_log(self):
        from automation.emailer import _email_log_path
        path = _email_log_path()
        if not os.path.exists(path):
            QMessageBox.information(self, "No Log Yet",
                "No email activity has been logged yet.\n"
                "Send a test email first, then check the log.")
            return
        EmailLogDialog(self, path).exec()



# ── Email Log Dialog ──────────────────────────────────────────────────────────

class EmailLogDialog(QDialog):
    """Shows email_log.txt with a refresh button and option to clear."""

    def __init__(self, parent, log_path: str):
        super().__init__(parent)
        self._path = log_path
        self.setWindowTitle("Email Activity Log")
        self.setModal(True)
        self.resize(700, 500)
        self._build_ui()

    def _build_ui(self):
        t = _t()
        self.setStyleSheet(f"QDialog{{background:{t.bg_window};}}")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Title bar
        title_bar = QWidget()
        title_bar.setStyleSheet(
            f"QWidget{{background:{t.bg_table_alt};border-bottom:1px solid {t.border};}}")
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(20, 10, 16, 10)
        tb.addWidget(_lbl("Email Activity Log", 13, bold=True))
        tb.addSpacing(8)
        tb.addWidget(_lbl(self._path, 10, color=t.text_muted))
        tb.addStretch()
        outer.addWidget(title_bar)

        # Log text area
        self._text = QTextEdit()
        self._text.setReadOnly(True)
        self._text.setFont(QFont(_MONO_FONT, 10))
        self._text.setStyleSheet(
            f"QTextEdit{{background:{t.bg_input};color:{t.text_primary};"
            f"border:none;padding:12px 16px;}}")
        outer.addWidget(self._text, stretch=1)
        self._reload()

        # Footer
        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        outer.addWidget(sep)

        footer = QWidget()
        footer.setStyleSheet(f"QWidget{{background:{t.bg_table_alt};}}")
        ft = QHBoxLayout(footer)
        ft.setContentsMargins(16, 10, 16, 10)
        ft.setSpacing(8)
        refresh_btn = _btn("Refresh", "outline", height=34, icon="btn_refresh.png")
        refresh_btn.clicked.connect(self._reload)
        ft.addWidget(refresh_btn)
        clear_btn = _btn("Clear Log", "danger", height=34, icon="btn_clear_log.png")
        clear_btn.clicked.connect(self._clear)
        ft.addWidget(clear_btn)
        ft.addStretch()
        close_btn = _btn("Close", "primary", height=34, icon="btn_close.png")
        close_btn.clicked.connect(self.accept)
        ft.addWidget(close_btn)
        outer.addWidget(footer)

    def _reload(self):
        try:
            with open(self._path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        except Exception as e:
            content = f"Could not read log: {e}"
        self._text.setPlainText(content)
        # Scroll to bottom so latest entry is visible
        self._text.moveCursor(self._text.textCursor().MoveOperation.End)

    def _clear(self):
        from PyQt6.QtWidgets import QMessageBox
        if QMessageBox.question(self, "Clear Log",
                "Delete all email log entries?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                ) == QMessageBox.StandardButton.Yes:
            try:
                open(self._path, "w").close()
            except Exception:
                pass
            self._text.clear()


# ── Mail Docs to Clients Dialog ───────────────────────────────────────────────

class MailDocsDialog(QDialog):
    """
    Tools → Mail Docs to Clients
    User picks a root folder, app scans for {PAN}-* sub-folders,
    matches against vault, filters by current AY, then lets user
    select clients and send emails with attachments.
    """

    _status_signal = pyqtSignal(str, str)   # pan, status text

    _COL_CHK   = 0
    _COL_NAME  = 1
    _COL_PAN   = 2
    _COL_EMAIL = 3
    _COL_CC    = 4
    _COL_FILES = 5

    def __init__(self, parent, vault, ay_label: str):
        super().__init__(parent)
        self._vault = vault
        self._ay_label = ay_label
        self._clients = []          # list of dicts from emailer.scan_for_clients
        self._sending = False
        self._client_status = {}    # pan → status string (✅ Sent / ❌ Failed / ⏳ …)
        self._scan_summary = ""     # base text from last scan

        self.setWindowTitle("Mail Docs to Clients")
        self.setMinimumSize(900, 560)
        self.resize(1020, 620)
        self.setSizeGripEnabled(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        self._build_ui()
        self._status_signal.connect(self._on_status)

    def _build_ui(self):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}"
            f"QLineEdit{{border:1px solid {t.border};border-radius:6px;padding:4px 8px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QLineEdit:focus{{border-color:{t.border_focus};}}"
            f"QTableWidget{{border:1.5px solid {t.border};border-radius:8px;"
            f"background:{t.bg_table};outline:0;gridline-color:{t.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {t.grid};padding:0 6px;}}"
        )

        main = QVBoxLayout(self)
        main.setContentsMargins(16, 14, 16, 12)
        main.setSpacing(8)

        # ── Title ─────────────────────────────────────────────────────────────
        title = QLabel("<b>Mail Docs to Clients</b>")
        title.setStyleSheet(f"font-size:14px;color:{t.text_primary};background:transparent;")
        main.addWidget(title)

        hint = QLabel(
            f"Select a folder, click <b>Scan</b> to find clients, then choose who to email."
            f"&nbsp;&nbsp;AY: <span style='color:{t.accent}'>{self._ay_label}</span>")
        hint.setStyleSheet(f"font-size:11px;color:{t.text_muted};background:transparent;")
        main.addWidget(hint)
        # ── Template / Folder / Filter — unified form rows ────────────────────
        from PyQt6.QtWidgets import QFormLayout
        _ROW_H = 30
        _lbl_ss  = f"font-size:11px;color:{t.text_muted};background:transparent;"
        _field_ss = (
            f"border:1px solid {t.border};border-radius:5px;padding:0 8px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};"
        )
        _combo_ss = (
            f"QComboBox{{border:1px solid {t.border};border-radius:5px;padding:0 8px;"
            f"font-size:11px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QComboBox::drop-down{{border:none;width:18px;}}"
            f"QComboBox QAbstractItemView{{background:{t.bg_input};color:{t.text_primary};"
            f"selection-background-color:{t.accent};}}"
        )
        _btn_ss_outline = (
            f"QPushButton{{background:transparent;color:{t.text_primary};"
            f"border:1px solid {t.border};border-radius:5px;padding:0 8px;"
            f"font-size:11px;font-weight:600;}}"
            f"QPushButton:hover{{background:{t.bg_table_alt};}}"
        )
        _btn_ss_primary = (
            f"QPushButton{{background:{t.accent};color:white;border:none;"
            f"border-radius:5px;padding:0 8px;font-size:11px;font-weight:600;}}"
            f"QPushButton:hover{{background:{t.accent_hover};}}"
            f"QPushButton:disabled{{background:{t.border};color:{t.text_muted};}}"
        )

        form = QFormLayout()
        form.setSpacing(8)
        form.setContentsMargins(0, 0, 0, 0)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        form.setFormAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)

        def _form_lbl(text):
            l = QLabel(text)
            l.setFixedHeight(_ROW_H)
            l.setStyleSheet(_lbl_ss)
            return l

        # Template row
        self._tpl_combo = QComboBox()
        self._tpl_combo.setFixedHeight(_ROW_H)
        self._tpl_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self._tpl_combo.setStyleSheet(_combo_ss)
        templates = self._vault.get_email_templates()
        active_name = self._vault.get_active_template_name()
        for tpl in templates:
            self._tpl_combo.addItem(tpl["name"])
        idx = next((i for i, t_ in enumerate(templates) if t_["name"] == active_name), 0)
        self._tpl_combo.setCurrentIndex(idx)
        form.addRow(_form_lbl("Template:"), self._tpl_combo)

        # Folder row
        folder_widget = QWidget()
        folder_widget.setStyleSheet("QWidget{background:transparent;}")
        folder_h = QHBoxLayout(folder_widget)
        folder_h.setContentsMargins(0, 0, 0, 0)
        folder_h.setSpacing(6)
        self._folder_edit = QLineEdit()
        self._folder_edit.setFixedHeight(_ROW_H)
        self._folder_edit.setPlaceholderText("Browse to the folder containing client sub-folders…")
        self._folder_edit.setStyleSheet(f"QLineEdit{{{_field_ss}}}QLineEdit:focus{{border-color:{t.border_focus};}}")
        default_dir = self._vault.get_setting("download_root_dir", "")
        if default_dir and os.path.isdir(default_dir):
            self._folder_edit.setText(default_dir)
        folder_h.addWidget(self._folder_edit, stretch=1)
        def _small_icon(name):
            p = _icon_path(name)
            if not p:
                return QIcon()
            return QIcon(QPixmap(p).scaled(16, 16, Qt.AspectRatioMode.KeepAspectRatio,
                                           Qt.TransformationMode.SmoothTransformation))

        browse_btn = QPushButton("Browse…")
        browse_btn.setIcon(_small_icon("btn_browse_folder.png"))
        browse_btn.setFixedSize(90, _ROW_H)
        browse_btn.setStyleSheet(_btn_ss_outline)
        browse_btn.clicked.connect(self._browse)
        folder_h.addWidget(browse_btn)
        self._scan_btn = QPushButton("Scan Folder")
        self._scan_btn.setIcon(_small_icon("btn_scan.png"))
        self._scan_btn.setFixedSize(110, _ROW_H)
        self._scan_btn.setStyleSheet(_btn_ss_primary)
        self._scan_btn.clicked.connect(self._scan)
        folder_h.addWidget(self._scan_btn)
        form.addRow(_form_lbl("Folder:"), folder_widget)

        # Filter row
        self._filter_edit = QLineEdit()
        self._filter_edit.setFixedHeight(_ROW_H)
        self._filter_edit.setPlaceholderText("Search by name, PAN or email…")
        self._filter_edit.setStyleSheet(f"QLineEdit{{{_field_ss}}}QLineEdit:focus{{border-color:{t.border_focus};}}")
        self._filter_edit.setClearButtonEnabled(True)
        self._filter_edit.textChanged.connect(self._apply_filter)
        form.addRow(_form_lbl("Filter:"), self._filter_edit)

        main.addLayout(form)

        # ── Table ─────────────────────────────────────────────────────────────
        self._table = QTableWidget(0, 6)
        self._table.setHorizontalHeaderLabels(
            ["", "Name  ⇅", "PAN  ⇅", "Email  ⇅", "CC  ⇅", "Files  ⇅"])

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_CHK,   QHeaderView.ResizeMode.Fixed)
        hdr.setSectionResizeMode(self._COL_NAME,  QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN,   QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_EMAIL, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(self._COL_CC,    QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_FILES, QHeaderView.ResizeMode.Fixed)
        hdr.setStretchLastSection(False)
        self._table.setColumnWidth(self._COL_CHK,   36)
        self._table.setColumnWidth(self._COL_NAME, 180)
        self._table.setColumnWidth(self._COL_PAN,  110)
        self._table.setColumnWidth(self._COL_CC,   180)
        self._table.setColumnWidth(self._COL_FILES, 80)

        t2 = _t()
        hdr.setStyleSheet(
            f"QHeaderView::section{{background-color:{t2.bg_header};border:none;"
            f"border-right:1px solid {t2.border};border-bottom:1px solid {t2.border};"
            f"font-weight:bold;color:{t2.text_muted};font-size:11px;height:32px;padding:0 6px;}}")
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setShowGrid(True)
        self._table.setAlternatingRowColors(False)
        self._table.setWordWrap(False)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self._sort_col = -1
        self._sort_order = Qt.SortOrder.AscendingOrder
        hdr.setSortIndicatorShown(False)
        hdr.sectionClicked.connect(self._on_header_clicked)

        main.addWidget(self._table, stretch=1)

        # ── Status bar ────────────────────────────────────────────────────────
        self._status_lbl = QLabel("")
        self._status_lbl.setStyleSheet(
            f"font-size:11px;color:{t.text_muted};background:transparent;")
        main.addWidget(self._status_lbl)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = QHBoxLayout(); footer.setSpacing(8)

        self._sel_all_btn = _btn("Select All", "outline", height=32, icon="btn_select_all.png")
        self._sel_all_btn.clicked.connect(self._select_all)
        footer.addWidget(self._sel_all_btn)

        self._sel_none_btn = _btn("Select None", "outline", height=32, icon="btn_select_none.png")
        self._sel_none_btn.clicked.connect(self._select_none)
        footer.addWidget(self._sel_none_btn)

        email_settings_btn = _btn("Email Settings", "outline", height=32, icon="icon_email.png")
        email_settings_btn.clicked.connect(self._open_email_settings)
        footer.addWidget(email_settings_btn)

        footer.addStretch()

        self._send_btn = _btn("Send to Selected", "success", height=34, icon="btn_send.png")
        self._send_btn.setMinimumWidth(150)
        self._send_btn.setEnabled(False)
        self._send_btn.clicked.connect(self._send)
        footer.addWidget(self._send_btn)

        self._close_btn = _btn("Close", "secondary", height=34, icon="btn_close.png")
        self._close_btn.clicked.connect(self.accept)
        footer.addWidget(self._close_btn)

        main.addLayout(footer)

    def _open_email_settings(self):
        SmtpSettingsDialog(self, self._vault).exec()
        self._refresh_tpl_combo()

    def _refresh_tpl_combo(self):
        templates = self._vault.get_email_templates()
        active_name = self._vault.get_active_template_name()
        current = self._tpl_combo.currentText()
        self._tpl_combo.blockSignals(True)
        self._tpl_combo.clear()
        for tpl in templates:
            self._tpl_combo.addItem(tpl["name"])
        # Restore previous selection if still exists, else use active
        names = [t["name"] for t in templates]
        if current in names:
            self._tpl_combo.setCurrentText(current)
        elif active_name in names:
            self._tpl_combo.setCurrentText(active_name)
        self._tpl_combo.blockSignals(False)

    # ── folder browse ─────────────────────────────────────────────────────────

    def _browse(self):
        start = self._folder_edit.text().strip() or os.path.expanduser("~")
        chosen = QFileDialog.getExistingDirectory(self, "Select Folder Containing Client Sub-Folders", start)
        if chosen:
            self._folder_edit.setText(chosen)

    # ── scan ──────────────────────────────────────────────────────────────────

    def _scan(self):
        from automation.emailer import scan_for_clients
        root = self._folder_edit.text().strip()
        if not root:
            QMessageBox.warning(self, "No Folder", "Please select a folder first.")
            return
        if not os.path.isdir(root):
            QMessageBox.warning(self, "Invalid Folder", f"Folder not found:\n{root}")
            return
        if not self._ay_label or self._ay_label == "Select AY/TY":
            QMessageBox.warning(self, "No AY Selected",
                                "Please select an Assessment Year in the main window first.")
            return

        assessees = self._vault.get_all_assessees()
        self._clients = scan_for_clients(root, self._ay_label, assessees)
        self._client_status = {}
        self._filter_edit.clear()
        self._populate_table()

        n = len(self._clients)
        if n == 0:
            self._status_lbl.setText(
                "No matching clients found. Check that the folder contains {PAN}-Name sub-folders.")
            self._send_btn.setEnabled(False)
        else:
            n_files = sum(1 for c in self._clients if c["attachments"])
            self._scan_summary = f"Found {n} client(s) — {n_files} with files for {self._ay_label}."
            self._update_status_label()
            self._send_btn.setEnabled(True)

    def _update_status_label(self):
        selected_n = sum(1 for chk in self._checkboxes.values() if chk.isChecked())
        base = getattr(self, "_scan_summary", "")
        if base:
            self._status_lbl.setText(f"{base}  ·  {selected_n} selected")
        else:
            self._status_lbl.setText(f"{selected_n} selected")

    def _apply_filter(self, text: str):
        q = text.strip().lower()
        for row in range(self._table.rowCount()):
            name  = (self._table.item(row, self._COL_NAME)  or QTableWidgetItem()).text().lower()
            pan   = (self._table.item(row, self._COL_PAN)   or QTableWidgetItem()).text().lower()
            email_w = self._table.cellWidget(row, self._COL_EMAIL)
            email = email_w.text().lower() if email_w else ""
            visible = not q or q in name or q in pan or q in email
            self._table.setRowHidden(row, not visible)

    def _populate_table(self, checked_pans: set = None, status_map: dict = None):
        t = _t()
        self._table.setRowCount(0)
        self._email_edits = {}
        self._cc_edits = {}
        self._checkboxes = {}

        for row, client in enumerate(self._clients):
            self._table.insertRow(row)
            self._table.setRowHeight(row, 38)

            pan = client["pan"]
            has_files = bool(client["attachments"])  # used for Files column display
            can_select = True  # always allow selection; doc filtering happens at send time

            # Checkbox cell
            row_bg = t.bg_table_alt if row % 2 else t.bg_table
            chk_widget = QWidget()
            chk_widget.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            chk_widget.setStyleSheet(f"QWidget{{background:{row_bg};}}")
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk = QCheckBox()
            chk.setChecked(checked_pans is not None and pan in checked_pans)
            chk.setEnabled(can_select)
            chk.setStyleSheet(
                f"QCheckBox{{background:transparent;}}"
                f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
                f"border-radius:3px;background:{t.bg_checkbox};}}"
                f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
                f"QCheckBox::indicator:disabled{{background:{t.border};}}")
            chk_layout.addWidget(chk)
            self._table.setCellWidget(row, self._COL_CHK, chk_widget)
            self._checkboxes[pan] = chk
            chk.stateChanged.connect(self._update_status_label)

            # Name
            name_item = QTableWidgetItem(client["name"])
            name_item.setForeground(QColor(t.text_primary if has_files else t.text_muted))
            self._table.setItem(row, self._COL_NAME, name_item)

            # PAN
            pan_item = QTableWidgetItem(pan)
            pan_item.setForeground(QColor(t.text_muted))
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_PAN, pan_item)

            # Email (editable inline + hidden sort item)
            email_val = client.get("email", "")
            email_edit = QLineEdit(email_val)
            email_edit.setPlaceholderText("type email…")
            email_edit.setStyleSheet(
                f"QLineEdit{{border:none;background:{row_bg};color:{t.text_primary};"
                f"font-size:11px;padding:0 4px;}}"
                f"QLineEdit:focus{{border-bottom:1px solid {t.accent};background:{t.bg_input_focus};}}")
            self._table.setCellWidget(row, self._COL_EMAIL, email_edit)
            email_sort = QTableWidgetItem(email_val)
            self._table.setItem(row, self._COL_EMAIL, email_sort)
            self._email_edits[pan] = email_edit

            # CC (editable inline + hidden sort item)
            cc_val = client.get("cc", "")
            cc_edit = QLineEdit(cc_val)
            cc_edit.setPlaceholderText("optional; separate with ;")
            cc_edit.setStyleSheet(
                f"QLineEdit{{border:none;background:{row_bg};color:{t.text_primary};"
                f"font-size:11px;padding:0 4px;}}"
                f"QLineEdit:focus{{border-bottom:1px solid {t.accent};background:{t.bg_input_focus};}}")
            self._table.setCellWidget(row, self._COL_CC, cc_edit)
            cc_sort = QTableWidgetItem(cc_val)
            self._table.setItem(row, self._COL_CC, cc_sort)
            self._cc_edits[pan] = cc_edit

            # Files — restore live status (e.g. ✅ Sent) if available, else show count
            n_files = len(client["attachments"])
            restored = status_map.get(pan) if status_map else None
            if restored and restored not in (f"{n_files} file{'s' if n_files != 1 else ''}", "⚠ No files"):
                files_item = QTableWidgetItem(restored)
                if "✅" in restored:
                    files_item.setForeground(QColor("#15803D"))
                elif "❌" in restored:
                    files_item.setForeground(QColor("#EF4444"))
                elif "⚠" in restored:
                    files_item.setForeground(QColor("#D97706"))
                else:
                    files_item.setForeground(QColor(t.text_primary))
            elif has_files:
                tip = "\n".join(os.path.basename(f) for f in client["attachments"])
                files_item = QTableWidgetItem(f"{n_files} file{'s' if n_files != 1 else ''}")
                files_item.setForeground(QColor(t.text_primary))
                files_item.setToolTip(tip)
            else:
                files_item = QTableWidgetItem("⚠ No files")
                files_item.setForeground(QColor("#EF4444"))
            files_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_FILES, files_item)

    # ── sort ──────────────────────────────────────────────────────────────────

    def _on_header_clicked(self, col: int):
        if col == self._COL_CHK:
            return
        hdr = self._table.horizontalHeader()
        if self._sort_col == col:
            self._sort_order = (
                Qt.SortOrder.DescendingOrder
                if self._sort_order == Qt.SortOrder.AscendingOrder
                else Qt.SortOrder.AscendingOrder
            )
        else:
            self._sort_col = col
            self._sort_order = Qt.SortOrder.AscendingOrder
        hdr.setSortIndicatorShown(True)
        hdr.setSortIndicator(self._sort_col, self._sort_order)
        self._sort_and_repopulate()

    def _sort_and_repopulate(self):
        # Capture current state keyed by PAN (stable, not row-index dependent)
        checked_pans = {pan for pan, chk in self._checkboxes.items() if chk.isChecked()}
        live_emails  = {pan: ed.text() for pan, ed in self._email_edits.items()}
        live_cc      = {pan: ed.text() for pan, ed in self._cc_edits.items()}
        live_status  = {pan: self._client_status.get(pan, "") for pan in self._checkboxes}

        # Merge typed email/cc back into clients list
        for c in self._clients:
            if c["pan"] in live_emails:
                c["email"] = live_emails[c["pan"]]
            if c["pan"] in live_cc:
                c["cc"] = live_cc[c["pan"]]

        # Sort clients list by chosen column
        rev = (self._sort_order == Qt.SortOrder.DescendingOrder)
        key_map = {
            self._COL_NAME:  lambda c: c["name"].lower(),
            self._COL_PAN:   lambda c: c["pan"].lower(),
            self._COL_EMAIL: lambda c: c.get("email", "").lower(),
            self._COL_CC:    lambda c: c.get("cc", "").lower(),
            self._COL_FILES: lambda c: len(c["attachments"]),
        }
        key_fn = key_map.get(self._sort_col)
        if key_fn:
            self._clients.sort(key=key_fn, reverse=rev)

        # Repopulate with sorted order, restoring checked state and status
        self._populate_table(checked_pans=checked_pans, status_map=live_status)
        # Re-apply active filter
        self._apply_filter(self._filter_edit.text())

    # ── select all / none ─────────────────────────────────────────────────────

    def _select_all(self):
        for row in range(self._table.rowCount()):
            if self._table.isRowHidden(row):
                continue
            item = self._table.item(row, self._COL_NAME)
            if not item:
                continue
            pan = self._table.item(row, self._COL_PAN)
            pan_text = pan.text() if pan else ""
            chk = self._checkboxes.get(pan_text)
            if chk and chk.isEnabled():
                chk.setChecked(True)

    def _select_none(self):
        for row in range(self._table.rowCount()):
            if self._table.isRowHidden(row):
                continue
            pan = self._table.item(row, self._COL_PAN)
            pan_text = pan.text() if pan else ""
            chk = self._checkboxes.get(pan_text)
            if chk:
                chk.setChecked(False)

    # ── send ──────────────────────────────────────────────────────────────────

    def _send(self):
        if self._sending:
            return

        cfg = self._vault.get_email_settings()
        if not cfg.get("smtp_host") or not cfg.get("smtp_user"):
            QMessageBox.warning(self, "SMTP Not Configured",
                                "Please configure email settings first.\n"
                                "Go to Settings → Email Settings.")
            return

        # Load selected template — use its subject, body and doc filter
        all_templates = self._vault.get_email_templates()
        tpl_name = self._tpl_combo.currentText()
        active_tpl = next((t for t in all_templates if t["name"] == tpl_name),
                          all_templates[0] if all_templates else None)
        if active_tpl:
            self._vault.set_active_template(active_tpl["name"])
            cfg["email_subject_tpl"] = active_tpl.get("subject", cfg.get("email_subject_tpl", ""))
            cfg["email_body_tpl"]    = active_tpl.get("body",    cfg.get("email_body_tpl", ""))
            doc_filter = active_tpl.get("docs", {})
        else:
            doc_filter = {}

        def _keep(path: str) -> bool:
            if not doc_filter:
                return True
            n = os.path.basename(path).upper()
            if "-26AS-" in n and n.endswith(".PDF"):  return doc_filter.get("26as_pdf",  True)
            if "-26AS-" in n and n.endswith(".XLSX"): return doc_filter.get("26as_xlsx", True)
            if "-168-"  in n and n.endswith(".PDF"):  return doc_filter.get("168_pdf",   True)
            if "-168-"  in n and n.endswith(".XLSX"): return doc_filter.get("168_xlsx",  True)
            if "-AIS-"  in n and n.endswith(".PDF"):  return doc_filter.get("ais_pdf",   True)
            if "-AIS-"  in n and n.endswith(".XLSX"): return doc_filter.get("ais_xlsx",  True)
            if "-TIS-"  in n and n.endswith(".PDF"):  return doc_filter.get("tis_pdf",   True)
            if "-ITR-"  in n and n.endswith("-FORM.PDF"):    return doc_filter.get("itr_form",    True)
            if "-ITR-"  in n and n.endswith("-RECEIPT.PDF"): return doc_filter.get("itr_receipt", True)
            if "-ITR-"  in n and n.endswith("-ITR-V.PDF"):   return doc_filter.get("itr_v",        True)
            if "-INTIMATION-" in n and n.endswith(".PDF"):   return doc_filter.get("intimation",  True)
            return True

        # Collect selected clients and validate emails
        selected = []
        missing_email = []
        skipped_no_docs = []   # (name, pan) of clients skipped due to no matching docs
        for i, client in enumerate(self._clients):
            pan = client["pan"]
            chk = self._checkboxes.get(pan)
            if not chk or not chk.isChecked():
                continue
            email = self._email_edits[pan].text().strip()
            if not email:
                missing_email.append(client["name"])
                self._email_edits[pan].setStyleSheet(
                    "QLineEdit{border:1.5px solid #EF4444;border-radius:4px;"
                    "background:#FEF2F2;color:#B91C1C;font-size:11px;padding:0 4px;}")
                continue
            cc = self._cc_edits[pan].text().strip()
            attachments = [a for a in client["attachments"] if _keep(a)]
            if not attachments:
                skipped_no_docs.append((client["name"], pan))
                continue
            selected.append({**client, "email": email, "cc": cc,
                             "ay_label": self._ay_label, "attachments": attachments})

        if missing_email:
            QMessageBox.warning(self, "Missing Email",
                                "Please enter email addresses for:\n" +
                                "\n".join(f"  • {n}" for n in missing_email))
            return

        if not selected and not skipped_no_docs:
            QMessageBox.information(self, "Nothing Selected", "No clients selected to email.")
            return

        if not selected:
            # All checked clients had no matching docs — mark rows and bail
            t = _t()
            for name, pan in skipped_no_docs:
                row = self._pan_to_row(pan)
                if row >= 0:
                    item = QTableWidgetItem("❌ Docs not found")
                    item.setForeground(QColor("#EF4444"))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                    self._table.setItem(row, self._COL_FILES, item)
            self._status_lbl.setText("No matching documents found for the selected template.")
            return

        # Save any inline-typed emails back to vault
        for client in selected:
            self._vault.update_assessee_email(client["pan"], client["email"], client.get("cc", ""))

        # Confirm
        reply = QMessageBox.question(self, "Confirm Send", "Send emails to selected clients?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        if reply != QMessageBox.StandardButton.Yes:
            return

        # Update UI to sending state
        self._sending = True
        self._send_btn.setEnabled(False)
        self._send_btn.setText("Sending…")
        self._scan_btn.setEnabled(False)

        t = _t()

        # Mark skipped clients with inline error status
        for name, pan in skipped_no_docs:
            row = self._pan_to_row(pan)
            if row >= 0:
                item = QTableWidgetItem("❌ Docs not found")
                item.setForeground(QColor("#EF4444"))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self._table.setItem(row, self._COL_FILES, item)

        # Mark rows as pending
        for client in selected:
            pan = client["pan"]
            row = self._pan_to_row(pan)
            if row >= 0:
                item = QTableWidgetItem("⏳ Sending…")
                item.setForeground(QColor("#92400E" if t.name != "light" else "#D97706"))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                self._table.setItem(row, self._COL_FILES, item)

        # Run send in background thread
        from ui.log_history import LogStore
        _log_store = LogStore()
        _pan_to_ay   = {c["pan"]: c.get("ay_label", self._ay_label) for c in selected}
        def _doc_label(path: str) -> str:
            n = os.path.basename(path).upper()
            if "-26AS-" in n and n.endswith(".PDF"):  return "26AS PDF"
            if "-26AS-" in n and n.endswith(".XLSX"): return "26AS Excel"
            if "-168-"  in n and n.endswith("-ITD.XLSX"): return "168 ITD Excel"
            if "-168-"  in n and n.endswith(".PDF"):  return "168 PDF"
            if "-168-"  in n and n.endswith(".XLSX"): return "168 Excel"
            if "-AIS-"  in n and n.endswith(".PDF"):  return "AIS PDF"
            if "-AIS-"  in n and n.endswith(".XLSX"): return "AIS Excel"
            if "-TIS-"  in n and n.endswith(".PDF"):  return "TIS"
            if "-ITR-"  in n and n.endswith("-FORM.PDF"):    return "ITR Form"
            if "-ITR-"  in n and n.endswith("-RECEIPT.PDF"): return "ITR Receipt"
            if "-ITR-"  in n and n.endswith("-ITR-V.PDF"):   return "ITR-V"
            if "-INTIMATION-" in n and n.endswith(".PDF"):   return "Intimation Order"
            return os.path.basename(path)

        _pan_to_docs = {
            c["pan"]: ", ".join(_doc_label(f) for f in c.get("attachments", []))
            for c in selected
        }

        def _on_progress(pan, status):
            self._status_signal.emit(pan, status)
            if status == "✅ Sent":
                try:
                    docs = _pan_to_docs.get(pan, "")
                    _log_store.record(pan, _pan_to_ay.get(pan, self._ay_label),
                                      f"[Email] {docs}" if docs else "[Email] Sent")
                except Exception:
                    pass

        def _worker():
            from automation.emailer import send_batch
            send_batch(
                cfg, selected,
                subject_tpl=cfg.get("email_subject_tpl", "Your Tax Documents — {ay}"),
                body_tpl=cfg.get("email_body_tpl", "Dear {client_name},\n\nPlease find attached your documents for {ay}.\n\nRegards,\n{firm_name}"),
                bcc_addresses=cfg.get("bcc_addresses", ""),
                progress_cb=_on_progress,
            )
            self._status_signal.emit("__done__", "")

        threading.Thread(target=_worker, daemon=True).start()

    def _pan_to_row(self, pan: str) -> int:
        for i, c in enumerate(self._clients):
            if c["pan"] == pan:
                return i
        return -1

    def _on_status(self, pan: str, status: str):
        if pan == "__done__":
            self._sending = False
            self._send_btn.setEnabled(True)
            self._send_btn.setText("Send to Selected")
            self._scan_btn.setEnabled(True)
            sent = sum(
                1 for c in self._clients
                if self._email_edits.get(c["pan"]) and
                   self._table.item(self._pan_to_row(c["pan"]), self._COL_FILES) and
                   (self._table.item(self._pan_to_row(c["pan"]), self._COL_FILES).text() or "").startswith("✅")
            )
            self._status_lbl.setText(f"Done — {sent} email(s) sent successfully.")
            return

        self._client_status[pan] = status
        row = self._pan_to_row(pan)
        if row < 0:
            return
        t = _t()
        item = QTableWidgetItem(status)
        if status.startswith("✅"):
            item.setForeground(QColor("#16A34A"))
        elif status.startswith("❌"):
            item.setForeground(QColor("#EF4444"))
        else:
            item.setForeground(QColor(t.text_muted))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        self._table.setItem(row, self._COL_FILES, item)
