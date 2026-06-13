"""
AayDocCapio — ITD Bulk Document Downloader (Form 26AS, AIS, TIS)
Run:  python3 app.py
"""
from version import __version__ as APP_VERSION

import sys, os, json, asyncio, threading, datetime, time, subprocess
from themes import THEMES, ThemeColors, build_stylesheet, get_theme, MONO_FONT_NAME as _MONO_FONT
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFrame, QLabel, QPushButton,
    QLineEdit, QCheckBox, QComboBox, QFileDialog, QScrollArea,
    QTabWidget, QHBoxLayout, QVBoxLayout, QGridLayout,
    QMessageBox, QTextEdit, QDialog, QRadioButton, QSplitter, QSizePolicy,
    QGraphicsDropShadowEffect, QListView, QStyledItemDelegate, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QToolButton, QMenu,
    QProgressBar, QCalendarWidget,
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer, QMetaObject, Q_ARG, QModelIndex
from PyQt6.QtGui import QFont, QTextCursor, QColor, QRegularExpressionValidator, QPalette, QAction, QIcon, QPixmap
from PyQt6.QtCore import QRegularExpression

def _app_dir() -> str:
    """
    Writable user-data directory for vault, settings, and outputs.
    - Windows compiled .exe : %LOCALAPPDATA%\\AayDocCapio
    - macOS compiled app    : ~/Library/Application Support/AayDocCapio
    - Linux/WSL compiled    : ~/.local/share/AayDocCapio
    - Running as script     : folder containing app.py
    """
    # sys.frozen = PyInstaller; __compiled__ = Nuitka
    if getattr(sys, "frozen", False) or globals().get("__compiled__"):
        if sys.platform == "win32":
            base = os.environ.get("LOCALAPPDATA", os.path.expanduser("~"))
        elif sys.platform == "darwin":
            base = os.path.join(os.path.expanduser("~"), "Library", "Application Support")
        else:
            base = os.path.join(os.path.expanduser("~"), ".local", "share")
        data_dir = os.path.join(base, "AayDocCapio")
        os.makedirs(data_dir, exist_ok=True)
        return data_dir
    return os.path.dirname(os.path.abspath(__file__))


def _default_download_dir() -> str:
    """
    Sensible default output directory per platform.
    - Windows (native or WSL): %USERPROFILE%\\Downloads → Documents → USERPROFILE
    - macOS  : ~/Downloads  (always exists on Mac)
    - Linux  : ~/Downloads if it exists, otherwise ~  (never create it)

    USERPROFILE takes priority over sys.platform so that WSL dev runs
    (sys.platform == 'linux' but USERPROFILE points to a real Windows path
    like C:\\Users\\deepak) pick the Windows Downloads folder correctly.
    Never creates the folder — just returns the path.
    """
    userprofile = os.environ.get("USERPROFILE", "")
    # USERPROFILE is only set on Windows (native) and WSL environments that
    # inherit the Windows environment. Treat its presence as "Windows paths apply."
    if userprofile and os.path.isdir(userprofile):
        downloads = os.path.join(userprofile, "Downloads")
        if os.path.isdir(downloads):
            return downloads
        documents = os.path.join(userprofile, "Documents")
        if os.path.isdir(documents):
            return documents
        return userprofile

    home = os.path.expanduser("~")
    if sys.platform == "darwin":
        return os.path.join(home, "Downloads")
    # Linux (no USERPROFILE) — use ~/Downloads only if it already exists
    downloads = os.path.join(home, "Downloads")
    return downloads if os.path.isdir(downloads) else home


def _open_path(path: str):
    """Open a file or folder in the OS file manager / default app."""
    if not path:
        return
    if not os.path.exists(path):
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.information(None, "Folder Not Found",
            f"The folder has not been created yet:\n{path}\n\n"
            "It will be created when the first file is downloaded.")
        return
    if sys.platform == "win32":
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        # WSL: prefer explorer.exe so the folder opens in Windows Explorer
        wsl_exe = "/mnt/c/Windows/explorer.exe"
        if os.path.exists(wsl_exe):
            wsl_path = subprocess.run(
                ["wslpath", "-w", path], capture_output=True, text=True).stdout.strip()
            subprocess.Popen([wsl_exe, wsl_path or path])
        else:
            subprocess.Popen(["xdg-open", path])


def _bundled_dir() -> str:
    """
    Directory for read-only assets bundled inside the .exe.
    PyInstaller uses _MEIPASS; Nuitka uses the exe directory.
    """
    if getattr(sys, "frozen", False):
        return sys._MEIPASS
    if globals().get("__compiled__"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


sys.path.insert(0, _bundled_dir())

try:
    from vault import VaultManager
    from automation.browser import browser_manager
    from automation.auth import login_itd, logout_itd
    from automation.downloader_26as import download_26as
    from automation.downloader_ais_tis import run_request_ais, run_download_ais_tis
except Exception as _import_err:
    import traceback
    _msg = (
        f"Failed to load required modules.\n\n"
        f"{traceback.format_exc()}\n\n"
        f"bundled_dir: {_bundled_dir()}\n"
        f"sys.path: {sys.path[:5]}"
    )
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, _msg, "AayDocCapio — Startup Error", 0x10)
    except Exception:
        pass
    sys.exit(1)


class _ComboDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        from PyQt6.QtWidgets import QStyle
        text = index.data() or ""
        painter.save()
        is_selected = bool(option.state & QStyle.StateFlag.State_Selected)
        is_hover    = bool(option.state & QStyle.StateFlag.State_MouseOver)
        t = _t()
        if is_selected:
            painter.fillRect(option.rect, QColor(t.accent))
            painter.setPen(QColor(t.accent_text))
        elif is_hover:
            painter.fillRect(option.rect, QColor(t.accent_light))
            painter.setPen(QColor(t.accent))
        else:
            painter.fillRect(option.rect, QColor(t.bg_input))
            painter.setPen(QColor(t.text_primary))
        painter.drawText(option.rect.adjusted(12, 0, -8, 0),
                         Qt.AlignmentFlag.AlignVCenter, text)
        painter.restore()

    def sizeHint(self, option, index):
        sh = super().sizeHint(option, index)
        return sh.__class__(sh.width(), max(sh.height(), 28))


class _ComboListView(QListView):
    """QListView that closes its parent ComboBox popup on mouse release."""
    def __init__(self, combo: 'StyledComboBox'):
        super().__init__()
        self._combo = combo
        self.setMouseTracking(True)
        self.setItemDelegate(_ComboDelegate(self))
        t = _t()
        self.setStyleSheet(
            f"QListView {{ border:1px solid {t.border}; background:{t.bg_input}; outline:none; color:{t.text_primary}; }}"
            f"QListView::item {{ padding:0px; color:{t.text_primary}; }}"
        )

    def mouseReleaseEvent(self, event):
        # macOS opens the popup on mouse-press, so the release of that same
        # click can land inside the popup and would instantly select + close
        # it ("vanishing" dropdown). Ignore releases right after opening.
        if time.monotonic() - getattr(self._combo, "_popup_opened_at", 0.0) < 0.30:
            super().mouseReleaseEvent(event)
            return
        index = self.indexAt(event.pos())
        if index.isValid():
            self._combo.setCurrentIndex(index.row())
            self._combo.hidePopup()
            event.accept()
            return
        super().mouseReleaseEvent(event)


class StyledComboBox(QComboBox):
    """QComboBox with reliable hover + click-closes behaviour."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._popup_was_open = False
        self.setView(_ComboListView(self))

    def showPopup(self):
        self._popup_was_open = True
        self._popup_opened_at = time.monotonic()
        super().showPopup()

    def hidePopup(self):
        super().hidePopup()
        # Keep flag set for one event cycle so row click guard can read it
        QTimer.singleShot(150, self._clear_popup_flag)

    def _clear_popup_flag(self):
        self._popup_was_open = False


def get_timestamp():
    import datetime
    try:
        from zoneinfo import ZoneInfo
        now = datetime.datetime.now(ZoneInfo("Asia/Kolkata"))
    except Exception:
        now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=5, minutes=30)))
    return now.strftime("%d-%m-%Y %H:%M:%S")


# ── Stylesheet ────────────────────────────────────────────────────────────────
# Font names for QFont() Python calls — full stacks live in themes.py
_UI_FONT   = "Segoe UI"      if sys.platform == "win32" else "Avenir Next"

# Active theme reference — updated by AayDocCapioApp._apply_theme()
# Used by module-level helpers (_btn, dialogs) so they stay theme-aware.
_active_theme: "ThemeColors" = get_theme("light")

def _t() -> "ThemeColors":
    """Return the currently active theme colours."""
    return _active_theme


def _shadow(blur=18, offset_y=3, alpha=22):
    e = QGraphicsDropShadowEffect()
    e.setBlurRadius(blur)
    e.setOffset(0, offset_y)
    e.setColor(QColor(0, 0, 0, alpha))
    return e


def _btn(text, style="secondary", height=36, min_width=None):
    b = QPushButton(text)
    b.setMinimumHeight(height)
    if min_width:
        b.setMinimumWidth(min_width)
    t = _t()
    COLORS = {
        "primary":   ("#2563EB",      "#1D4ED8",      "white"),
        "success":   ("#16A34A",      "#15803D",      "white"),
        "danger":    ("#EF4444",      "#DC2626",      "white"),
        "secondary": ("#64748B",      "#475569",      "white"),
        "warning":   ("#EAB308",      "#CA8A04",      "white"),
        "outline":   ("transparent",  t.bg_table_alt, t.text_primary),
        "edit":      ("#0284C7",      "#0369A1",      "white"),
        "delete":    ("#DC2626",      "#991B1B",      "white"),
    }
    bg, hov, fg = COLORS.get(style, COLORS["secondary"])
    border = f"1px solid {t.border}" if style == "outline" else "none"
    b.setStyleSheet(
        f"QPushButton {{ background:{bg}; color:{fg}; border:{border}; "
        f"border-radius:6px; padding:6px 14px; font-weight:bold; font-size:12px; }}"
        f"QPushButton:hover {{ background:{hov}; }}"
        f"QPushButton:disabled {{ background:{t.border}; color:{t.text_muted}; }}"
    )
    return b


def _lbl(text, size=12, bold=False, color=None):
    l = QLabel(text)
    c = color or _t().text_primary
    l.setStyleSheet(f"color:{c}; font-size:{size}px;" + (" font-weight:bold;" if bold else ""))
    return l


# ── Manage Years Dialog ───────────────────────────────────────────────────────
class ManageYearsDialog(QDialog):
    def __init__(self, parent, json_path: str, on_save):
        super().__init__(parent)
        self.setWindowTitle("Manage Assessment / Tax Years")
        self.setFixedSize(500, 560)
        self.setModal(True)
        self._json_path = json_path
        self._on_save = on_save
        self._checkboxes = []  # [(entry_dict, QCheckBox)]

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                entries = json.load(f)
        except Exception:
            entries = []

        self._build_ui(entries)

    def _build_ui(self, entries):
        t = _t()
        self.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{color:{t.text_primary};background:transparent;}}"
            f"QRadioButton{{color:{t.text_primary};background:transparent;spacing:6px;}}"
            f"QRadioButton::indicator{{width:14px;height:14px;border:1.5px solid {t.border};"
            f"border-radius:7px;background:{t.bg_checkbox};}}"
            f"QRadioButton::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
            f"QScrollArea{{background:{t.bg_input};border:1px solid {t.border};border-radius:6px;}}"
            f"QScrollArea > QWidget > QWidget{{background:{t.bg_input};}}"
        )
        main = QVBoxLayout(self)
        main.setContentsMargins(20, 16, 20, 16)
        main.setSpacing(8)

        main.addWidget(_lbl("Manage Assessment / Tax Years", 13, bold=True))
        main.addWidget(_lbl("Toggle enabled/disabled or add new years.", 10, color=t.text_muted))
        main.addWidget(_lbl("Existing Entries", 11, bold=True, color=t.text_muted))

        scroll = QScrollArea()
        scroll.setFixedHeight(180)
        scroll.setWidgetResizable(True)
        inner = QWidget()
        inner.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        inner.setStyleSheet(f"QWidget{{background:{t.bg_input};}}")
        self._list_layout = QVBoxLayout(inner)
        self._list_layout.setSpacing(2)
        self._list_layout.addStretch()
        scroll.setWidget(inner)
        main.addWidget(scroll)

        for e in entries:
            self._add_row(e)

        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep)

        main.addWidget(_lbl("Add New Year", 11, bold=True, color=t.text_muted))

        # Type radio
        type_row = QHBoxLayout()
        type_row.addWidget(_lbl("Type:", 11))
        self._type_ay = QRadioButton("AY (Assessment Year)"); self._type_ay.setChecked(True)
        self._type_ty = QRadioButton("TY (Tax Year)")
        self._type_ay.toggled.connect(self._auto_fy)
        type_row.addWidget(self._type_ay); type_row.addWidget(self._type_ty)
        type_row.addStretch(); main.addLayout(type_row)

        # Year + FY
        yr_row = QHBoxLayout()
        yr_row.addWidget(_lbl("Year:", 11))
        self._year_edit = QLineEdit(); self._year_edit.setPlaceholderText("e.g. 2027-28")
        self._year_edit.setFixedWidth(120); self._year_edit.textChanged.connect(self._auto_fy)
        yr_row.addWidget(self._year_edit)
        yr_row.addWidget(_lbl("FY:", 11))
        self._fy_edit = QLineEdit(); self._fy_edit.setPlaceholderText("auto-filled")
        self._fy_edit.setFixedWidth(120)
        yr_row.addWidget(self._fy_edit)
        yr_row.addWidget(_lbl("(editable)", 10, color=t.text_muted))
        yr_row.addStretch(); main.addLayout(yr_row)

        add_btn = _btn("＋ Add to List", "outline", height=32, min_width=130)
        add_btn.clicked.connect(self._add_entry)
        main.addWidget(add_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        sep2 = QFrame(); sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet(f"background:{t.border};border:none;max-height:1px;")
        main.addWidget(sep2)

        btns_row = QHBoxLayout()
        save_btn = _btn("💾 Save & Close", "primary", height=36)
        save_btn.clicked.connect(self._save)
        cancel_btn = _btn("Cancel", "secondary", height=36)
        cancel_btn.clicked.connect(self.reject)
        btns_row.addWidget(save_btn); btns_row.addWidget(cancel_btn)
        main.addLayout(btns_row)

    def _add_row(self, entry):
        t = _t()
        cb = QCheckBox(entry["label"])
        cb.setChecked(entry.get("enabled", True))
        cb.setStyleSheet(
            f"QCheckBox{{color:{t.text_primary};background:transparent;spacing:6px;}}"
            f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
            f"border-radius:3px;background:{t.bg_checkbox};}}"
            f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}"
        )
        self._checkboxes.append((entry, cb))
        self._list_layout.insertWidget(self._list_layout.count() - 1, cb)

    def _auto_fy(self):
        import re
        year = self._year_edit.text().strip()
        m = re.match(r"(\d{4})-(\d{2}|\d{4})$", year)
        if not m:
            return
        y1 = int(m.group(1)); suffix = m.group(2)
        y2 = int(str(y1)[:2] + suffix) if len(suffix) == 2 else int(suffix)
        fy = f"{y1}-{str(y2)[-2:]}" if self._type_ty.isChecked() else f"{y1-1}-{str(y2-1)[-2:]}"
        self._fy_edit.setText(fy)

    def _add_entry(self):
        year_type = "TY" if self._type_ty.isChecked() else "AY"
        year = self._year_edit.text().strip()
        fy = self._fy_edit.text().strip()
        if not year or not fy:
            QMessageBox.warning(self, "Missing Fields", "Please fill in Year and FY.")
            return
        label = f"{year_type} {year} (FY {fy})"
        if any(e["label"] == label for e, _ in self._checkboxes):
            QMessageBox.warning(self, "Duplicate", f'"{label}" already exists.')
            return
        year_obj = {"TY": year, "FY": fy} if year_type == "TY" else {"AY": year, "FY": fy}
        self._add_row({"label": label, "enabled": True, "year": year_obj})
        self._year_edit.clear(); self._fy_edit.clear()

    def _save(self):
        final = [{**e, "enabled": cb.isChecked()} for e, cb in self._checkboxes]
        def _sort_key(e):
            y = e.get("year", {})
            label_year = y.get("AY") or y.get("TY") or y.get("FY") or "0000-00"
            try:
                return (0 if not e.get("enabled", True) else 1, -int(label_year[:4]))
            except ValueError:
                return (1, 0)
        final.sort(key=_sort_key)
        try:
            with open(self._json_path, "w", encoding="utf-8") as f:
                json.dump(final, f, indent=2, ensure_ascii=False)
            self._on_save()
            self.accept()
        except Exception as ex:
            QMessageBox.critical(self, "Save Error", str(ex))


# ── Batch Progress Dialog ─────────────────────────────────────────────────────

def _friendly_error(raw: str) -> str:
    """
    Translate technical Playwright / network exception messages into plain
    English suitable for display in the progress table status column.
    """
    r = raw.lower()
    # Wrong password / auth failures (already human-readable from auth.py)
    if "authentication failed" in r:
        # Strip the "AUTHENTICATION FAILED: " prefix, keep the reason
        for prefix in ("authentication failed: ", "authentication failed"):
            if raw.lower().startswith(prefix):
                return raw[len(prefix):].strip() or "Incorrect credentials"
        return raw
    # Network / portal unreachable
    if "err_empty_response" in r or "empty response" in r:
        return "ITD portal is not responding — portal may be down for maintenance. Open incometax.gov.in in a browser to check, then retry."
    if "err_connection_refused" in r or "connection refused" in r:
        return "ITD portal refused the connection — it may be down for maintenance. Try again in a few minutes."
    if "err_name_not_resolved" in r or "name not resolved" in r:
        return "Cannot reach ITD portal — check your internet connection and try again."
    if "err_timed_out" in r or "timed out" in r or "timeout" in r:
        return "ITD portal took too long to respond — portal may be busy. Try again in a few minutes."
    if "err_connection_reset" in r or "connection reset" in r:
        return "Connection was reset by the portal — try again in a few minutes."
    if "net::" in r:
        # Generic net:: error — strip the technical prefix
        import re as _re
        m = _re.search(r"net::(\w+)", raw)
        code = m.group(1) if m else "NETWORK_ERROR"
        return f"Network error — ITD portal may be temporarily unavailable. ({code})"
    # Playwright target/browser closed
    if "target closed" in r or "browser has been closed" in r:
        return "Browser closed unexpectedly — try again"
    # Aborted by user
    if "aborted by user" in r or "cancelled" in r:
        return "Stopped by user"
    # 2FA
    if "2fa" in r or "otp" in r:
        return raw  # already human-readable
    # Generic fallback — truncate but don't show raw technical noise
    clean = raw.split("\n")[0].strip()   # first line only
    if len(clean) > 72:
        clean = clean[:69] + "..."
    return clean


# Status → (icon prefix, background colour, text colour)
_STATUS_STYLE = {
    "waiting":  ("⬜", "#F8FAFC", "#64748B"),
    "running":  ("⏳", "#FFF7ED", "#92400E"),
    "success":  ("✅", "#F0FDF4", "#15803D"),
    "queued":   ("🕐", "#FFFBEB", "#92400E"),
    "failed":   ("❌", "#FEF2F2", "#B91C1C"),
}

def _status_style(text: str):
    t = text.lower()
    if any(x in t for x in ("waiting",)):          return _STATUS_STYLE["waiting"]
    if any(x in t for x in ("✅", "downloaded")):  return _STATUS_STYLE["success"]
    if any(x in t for x in ("🕐", "queued", "pls use")):  return _STATUS_STYLE["queued"]
    if any(x in t for x in ("❌", "failed", "error")): return _STATUS_STYLE["failed"]
    return _STATUS_STYLE["running"]  # ⏳ anything in-progress


class BatchProgressDialog(QDialog):
    """
    Live progress popup shown during any batch run.
    Columns: Name | PAN | Status | Save Path (clickable link)
    Status and path updates arrive from the worker thread via Qt signals.
    """
    # (pan, status_text), (pan, folder_path), and () for resume
    _update_signal    = pyqtSignal(str, str)
    _path_signal      = pyqtSignal(str, str)
    _resume_signal    = pyqtSignal(list)   # emitted with remaining targets

    # column indices
    _COL_NAME   = 0
    _COL_PAN    = 1
    _COL_STATUS = 2
    _COL_PATH   = 3

    def __init__(self, targets: list, mode: str, ay: str = "",
                 stop_callback=None, resume_callback=None,
                 output_dir: str = "", parent=None):
        super().__init__(parent)
        self._stop_callback   = stop_callback
        self._resume_callback = resume_callback
        self._output_dir      = output_dir
        self._mode          = mode
        self._ay            = ay
        self._targets       = targets          # kept for Excel report
        self._pan_to_path   = {}               # pan → folder path (filled at runtime)

        # Human-readable label per mode
        mode_label = {
            "26as":        "Downloading 26AS",
            "request_ais": "Requesting AIS Generation",
            "ais_tis":     "Downloading AIS / TIS",
        }.get(mode, "Batch Run")
        self._mode_label = mode_label

        self.setWindowTitle(f"{mode_label} — Batch Progress")
        self.setMinimumSize(900, 500)
        self.resize(1060, min(160 + len(targets) * 42, 720))
        self.setSizeGripEnabled(True)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint)
        _bt = _t()
        self.setStyleSheet(f"QDialog{{background:{_bt.bg_window};}}")

        self._pan_to_row = {}

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 12)
        layout.setSpacing(8)

        # ── Title bar ────────────────────────────────────────────────────────
        ay_tag = (f" &nbsp;·&nbsp; <span style='color:{_bt.accent}'>{ay}</span>") if ay else ""
        title = QLabel(f"<b>{mode_label}</b> — {len(targets)} client(s){ay_tag}")
        title.setStyleSheet(f"font-size:14px; color:{_bt.text_primary}; background:transparent;")
        layout.addWidget(title)

        # ── Table ────────────────────────────────────────────────────────────
        self._table = QTableWidget(len(targets), 4)
        self._table.setHorizontalHeaderLabels(["Name", "PAN", "Status", "Save Path"])

        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(self._COL_NAME,   QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PAN,    QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_STATUS, QHeaderView.ResizeMode.Interactive)
        hdr.setSectionResizeMode(self._COL_PATH,   QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(self._COL_NAME,   180)
        self._table.setColumnWidth(self._COL_PAN,    120)
        self._table.setColumnWidth(self._COL_STATUS, 260)

        hdr.setStyleSheet(
            f"QHeaderView::section{{"
            f"background-color:{_bt.bg_header};"
            f"border:none;"
            f"border-right:1px solid {_bt.border};"
            f"border-bottom:1px solid {_bt.border};"
            f"font-weight:bold;color:{_bt.text_muted};"
            f"font-size:11px;height:34px;"
            f"padding:0 8px;}}")
        self._table.verticalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._table.setShowGrid(True)
        self._table.setAlternatingRowColors(False)
        self._table.setWordWrap(False)
        self._table.setStyleSheet(
            f"QTableWidget{{border:1.5px solid {_bt.border};border-radius:8px;"
            f"background:{_bt.bg_table};outline:0;gridline-color:{_bt.grid};}}"
            f"QTableWidget::item{{border-bottom:1px solid {_bt.grid};padding:0 8px;}}"
            f"QPushButton{{border:none;background:transparent;font-size:14px;}}"
            f"QPushButton:hover{{background:{_bt.bg_table_alt};border-radius:4px;}}")
        self._table.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        for row, t in enumerate(targets):
            pan  = t.get("pan", "")
            name = t.get("name", "—")
            self._pan_to_row[pan] = row
            self._table.setRowHeight(row, 40)

            name_item = QTableWidgetItem(name)
            name_item.setForeground(QColor("#1E293B"))
            self._table.setItem(row, self._COL_NAME, name_item)

            pan_item = QTableWidgetItem(pan)
            pan_item.setForeground(QColor("#475569"))
            pan_item.setFont(QFont(_MONO_FONT, 10))
            pan_item.setTextAlignment(
                Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            self._table.setItem(row, self._COL_PAN, pan_item)

            self._set_status_item(row, "⬜ Waiting")

            # Save Path — shown as a clickable link once the path is known
            path_lbl = QLabel("—")
            path_lbl.setStyleSheet(
                f"color:{_bt.text_muted};font-size:11px;padding:0 8px;background:transparent;")
            path_lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            path_lbl.setWordWrap(False)
            path_lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
            path_lbl.setOpenExternalLinks(False)
            path_lbl.linkActivated.connect(lambda url: _open_path(url))
            self._table.setCellWidget(row, self._COL_PATH, path_lbl)

        layout.addWidget(self._table, stretch=1)

        # ── Progress bar row ──────────────────────────────────────────────────
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, len(targets))
        self._progress_bar.setValue(0)
        self._progress_bar.setFixedHeight(18)
        self._progress_bar.setTextVisible(True)
        self._progress_bar.setFormat(f"0 / {len(targets)} done")
        self._progress_bar.setStyleSheet(
            f"QProgressBar{{border:1px solid {_bt.border};border-radius:9px;"
            f"background:{_bt.scrollbar_handle};text-align:center;font-size:11px;"
            f"font-weight:600;color:{_bt.accent_text};}}"
            f"QProgressBar::chunk{{background:#16A34A;border-radius:9px;}}")
        layout.addWidget(self._progress_bar)

        # ── Footer: saving-to + buttons ───────────────────────────────────────
        footer = QHBoxLayout()
        footer.setSpacing(8)
        footer.setContentsMargins(0, 0, 0, 0)

        # Saving-to (left side, stretches)
        loc_cap = QLabel("📁")
        loc_cap.setStyleSheet("font-size:13px;background:transparent;")
        loc_cap.setFixedWidth(18)
        footer.addWidget(loc_cap)

        self._loc_val = QLabel(output_dir or "—")
        self._loc_val.setStyleSheet(f"color:{_bt.text_muted};font-size:11px;background:transparent;")
        self._loc_val.setWordWrap(False)
        self._loc_val.setMinimumWidth(0)
        self._loc_val.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self._loc_val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        footer.addWidget(self._loc_val, stretch=1)

        # Open Download Folder (right next to the path label)
        self._open_folder_btn = QPushButton("📂  Open Folder")
        self._open_folder_btn.setFixedHeight(32)
        self._open_folder_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._open_folder_btn.clicked.connect(lambda: _open_path(self._output_dir))
        footer.addWidget(self._open_folder_btn)

        # Download Report
        self._report_btn = QPushButton("⬇  Download Report")
        self._report_btn.setFixedHeight(32)
        self._report_btn.setEnabled(False)
        self._report_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.bg_table_alt};color:{_bt.text_primary};border:1px solid {_bt.border};"
            f"border-radius:6px;font-size:12px;padding:0 12px;}}"
            f"QPushButton:enabled:hover{{background:{_bt.bg_input};}}"
            f"QPushButton:disabled{{color:{_bt.text_muted};border-color:{_bt.border};}}")
        self._report_btn.clicked.connect(self._export_report)
        footer.addWidget(self._report_btn)

        # Stop
        self._stop_btn = QPushButton("⏹  Stop")
        self._stop_btn.setFixedHeight(32)
        self._stop_btn.setMinimumWidth(90)
        self._stop_btn.setStyleSheet(
            "QPushButton{background:#EF4444;color:#FFFFFF;border:none;"
            "border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:#DC2626;}"
            "QPushButton:disabled{background:#E2E8F0;color:#94A3B8;}")
        self._stop_btn.clicked.connect(self._on_stop_clicked)
        footer.addWidget(self._stop_btn)

        # Resume (hidden until aborted)
        self._resume_btn = QPushButton("▶  Resume")
        self._resume_btn.setFixedHeight(32)
        self._resume_btn.setMinimumWidth(100)
        self._resume_btn.setVisible(False)
        self._resume_btn.setStyleSheet(
            "QPushButton{background:#16A34A;color:#FFFFFF;border:none;"
            "border-radius:6px;font-size:12px;font-weight:600;padding:0 12px;}"
            "QPushButton:hover{background:#15803D;}")
        self._resume_btn.clicked.connect(self._on_resume_clicked)
        footer.addWidget(self._resume_btn)

        # Close
        self._close_btn = QPushButton("Close")
        self._close_btn.setFixedSize(80, 32)
        self._close_btn.setEnabled(False)
        self._close_btn.setStyleSheet(
            f"QPushButton{{background:{_bt.border};color:{_bt.text_muted};border:none;"
            f"border-radius:6px;font-size:12px;}}"
            f"QPushButton:enabled{{background:{_bt.accent};color:{_bt.accent_text};}}"
            f"QPushButton:enabled:hover{{background:{_bt.accent_hover};}}")
        self._close_btn.clicked.connect(self.accept)
        footer.addWidget(self._close_btn)

        layout.addLayout(footer)

        self._done_count = 0
        self._total      = len(targets)
        self._rows_data  = {}   # pan → {"name", "path", "status"} for report

        for t in targets:
            self._rows_data[t.get("pan", "")] = {
                "name": t.get("name", ""), "path": "", "status": "Waiting", "ts": ""}

        self._update_signal.connect(self._on_update)
        self._path_signal.connect(self._on_path_update)

    # ── internal helpers ──────────────────────────────────────────────────────

    def _set_status_item(self, row: int, text: str):
        _, bg, fg = _status_style(text)
        item = QTableWidgetItem(text)
        item.setForeground(QColor(fg))
        item.setBackground(QColor(bg))
        item.setFont(QFont(_UI_FONT, 10))
        item.setToolTip(text)
        self._table.setItem(row, self._COL_STATUS, item)

    def _on_update(self, pan: str, status: str):
        row = self._pan_to_row.get(pan)
        if row is None:
            return
        self._set_status_item(row, status)
        if pan in self._rows_data:
            self._rows_data[pan]["status"] = status
            self._rows_data[pan]["ts"] = datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S")
        terminal = ("✅", "❌", "🕐", "⬜", "⏹")
        if any(status.startswith(p) for p in terminal):
            self._done_count += 1
            self._progress_bar.setValue(self._done_count)
            self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")
        if self._done_count >= self._total:
            self._close_btn.setEnabled(True)
            self._report_btn.setEnabled(True)
            self._progress_bar.setFormat(f"All {self._total} done")

    def _on_path_update(self, pan: str, folder: str):
        row = self._pan_to_row.get(pan)
        if row is None:
            return
        self._pan_to_path[pan] = folder
        if pan in self._rows_data:
            self._rows_data[pan]["path"] = folder

        # Update path cell widget (QLabel link)
        lbl = self._table.cellWidget(row, self._COL_PATH)
        if isinstance(lbl, QLabel):
            lbl.setText(
                f'<a href="{folder}" style="color:#2563EB;text-decoration:underline;">'
                f'{folder}</a>')
            lbl.setToolTip(folder)
            lbl.setStyleSheet(
                "font-size:11px;padding:0 8px;background:transparent;")

    def _on_stop_clicked(self):
        if self._stop_callback:
            self._stop_callback()
        self._stop_btn.setEnabled(False)
        self._stop_btn.setText("⏹  Stopping...")

    # ── Excel report ──────────────────────────────────────────────────────────

    def _export_report(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"BatchReport_{self._ay.replace(' ','_')}_{timestamp}.xlsx"
        default_path = os.path.join(self._output_dir or os.path.expanduser("~"), default_name)

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Download Report", default_path,
            "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Download Report"

        # Header styling
        hdr_fill  = PatternFill("solid", fgColor="0F172A")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        link_font = Font(color="2563EB", underline="single", size=10)
        body_font = Font(size=10)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        thin      = Side(style="thin", color="CBD5E1")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["#", "Client Name", "Save Folder", "Status", "Timestamp"]
        col_widths = [5, 30, 60, 40, 22]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[1].height = 22

        for seq, t in enumerate(self._targets, start=1):
            pan  = t.get("pan", "")
            data = self._rows_data.get(pan, {})
            row_num = seq + 1

            folder = data.get("path", "")
            status = data.get("status", "—")
            name   = data.get("name", t.get("name", ""))
            row_ts = data.get("ts", "")

            ws.cell(row=row_num, column=1, value=seq).alignment = center

            ws.cell(row=row_num, column=2, value=name).alignment = left

            # Folder as clickable hyperlink if it exists
            if folder and os.path.exists(folder):
                cell = ws.cell(row=row_num, column=3, value=folder)
                cell.hyperlink = folder
                cell.font      = link_font
                cell.alignment = left
            else:
                ws.cell(row=row_num, column=3, value=folder or "—").alignment = left

            # Strip emoji from status for cleaner Excel display
            import re as _re
            status_clean = _re.sub(r'[^\x00-\x7F✅❌🕐⬜⏹⏳]+', '', status).strip()
            ws.cell(row=row_num, column=4, value=status_clean).alignment = left

            ws.cell(row=row_num, column=5, value=row_ts or "—").alignment = center

            for col_idx in range(1, 6):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border
                if not cell.font or cell.font == Font():
                    cell.font = body_font
            ws.row_dimensions[row_num].height = 18

        # Freeze header row
        ws.freeze_panes = "A2"

        try:
            wb.save(path)
            _open_path(path)
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", str(e))

    # ── public API (called from worker thread / main thread) ──────────────────

    def set_status(self, pan: str, status: str):
        """Thread-safe status update."""
        self._update_signal.emit(pan, status)

    def set_client_path(self, pan: str, folder: str):
        """Thread-safe path update — call once the client folder is known."""
        self._path_signal.emit(pan, folder)

    def batch_finished(self, aborted: bool = False):
        """Called when batch ends to enable Close/Report and hide Stop.
        If aborted, sweeps any non-terminal rows to ⏹ Stopped and shows Resume."""
        terminal = ("✅", "❌", "🕐", "⏹", "⬜ Skipped")
        if aborted:
            for pan, row in self._pan_to_row.items():
                item = self._table.item(row, self._COL_STATUS)
                current = item.text() if item else ""
                if not any(current.startswith(t) for t in terminal):
                    self._set_status_item(row, "⏹ Stopped")
                    if pan in self._rows_data:
                        self._rows_data[pan]["status"] = "⏹ Stopped"
        self._stop_btn.setVisible(False)
        self._resume_btn.setVisible(aborted)
        self._close_btn.setEnabled(True)
        self._report_btn.setEnabled(True)
        n = self._done_count
        self._progress_bar.setValue(n)
        label = "Stopped" if aborted else "All done"
        self._progress_bar.setFormat(f"{label} — {n} / {self._total} processed")

    def batch_resumed(self):
        """Called when the user clicks Resume — resets UI back to running state."""
        self._resume_btn.setVisible(False)
        self._stop_btn.setText("⏹  Stop")
        self._stop_btn.setEnabled(True)
        self._stop_btn.setVisible(True)
        self._close_btn.setEnabled(False)
        self._report_btn.setEnabled(False)
        self._progress_bar.setFormat(f"{self._done_count} / {self._total} done")

    def _on_resume_clicked(self):
        # Collect targets that are ⏹ Stopped (retry those only)
        remaining = []
        for t in self._targets:
            pan = t.get("pan", "")
            status = (self._rows_data.get(pan) or {}).get("status", "")
            if status.startswith("⏹"):
                remaining.append(t)
                # Reset row to Waiting
                row = self._pan_to_row.get(pan)
                if row is not None:
                    self._set_status_item(row, "⬜ Waiting")
                    self._rows_data[pan]["status"] = "⬜ Waiting"
        if remaining and self._resume_callback:
            self._resume_callback(remaining)


# ── Main Window ───────────────────────────────────────────────────────────────
class AayDocCapioApp(QMainWindow):
    _log_signal = pyqtSignal(str)
    _batch_done_signal = pyqtSignal()
    _show_progress_signal = pyqtSignal(list, str, str, str)   # (targets, mode, ay, output_dir)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("AayDocCapio — Tax Documents. Delivered to You.")
        self.setMinimumSize(1100, 720)
        self.resize(1200, 780)

        self.vault = VaultManager(
            vault_path=os.path.join(_app_dir(), "tax_vault.json"))
        self.selected_ids = set()
        self.editing_id = None
        self.is_running = False
        self._checkbox_map = {}
        
        # Generate checkmark image for custom check box styling
        self.checkmark_path = os.path.join(_app_dir(), "checkmark.png")
        if not os.path.exists(self.checkmark_path):
            try:
                from PIL import Image, ImageDraw
                img = Image.new("RGBA", (16, 16), (0, 0, 0, 0))
                draw = ImageDraw.Draw(img)
                draw.line([(3, 8), (7, 12), (13, 4)], fill=(255, 255, 255, 255), width=2)
                img.save(self.checkmark_path, "PNG")
            except Exception as e:
                print(f"Error generating checkmark: {e}")

        self._ais_requested_time = None   # datetime when Request AIS last completed
        self._last_mode = None            # mode of last completed batch
        self._ais_results = {}            # pan → "instant" | "queued" | "failed" | "skipped"
        self._last_errors = {}            # pan → error message string
        self._batch_loop = None           # asyncio event loop for the running batch
        self._batch_task = None           # asyncio Task for the running batch
        self._batch_aborted = False       # True if user clicked Stop
        self._last_batch_params = None    # (ay, fy, root_dir, mode) for resume

        self._log_signal.connect(self._append_log)
        self._batch_done_signal.connect(self._on_batch_done)
        self._show_progress_signal.connect(self._show_progress_dialog)
        self._progress_dialog = None   # BatchProgressDialog instance

        try:
            log_path = os.path.join(_app_dir(), "app.log")
            _MAX_LOG_BYTES = 1 * 1024 * 1024  # 1 MB
            if os.path.exists(log_path) and os.path.getsize(log_path) > _MAX_LOG_BYTES:
                with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write(content[len(content) // 2:])
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n=== Session Started {get_timestamp()} ===\n")
        except Exception:
            pass

        self._current_theme = self.vault.get_setting("theme", "light")
        self._build_ui()
        self.refresh_grid()
        # Apply saved theme after UI is built
        QTimer.singleShot(0, lambda: self._apply_theme(self._current_theme))

        # Check Chromium on startup in background — installs silently if missing
        QTimer.singleShot(1500, self._check_browser)

    # ── Build UI ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Menu bar ──────────────────────────────────────────────────────────
        menubar = self.menuBar()

        # Client Master menu
        cm_menu = menubar.addMenu("Client Master")
        act_add  = QAction("➕  Add New Client",          self); act_add.triggered.connect(self._open_add_client)
        act_imp  = QAction("📥  Import from CSV / Excel", self); act_imp.triggered.connect(self.bulk_import)
        act_exp  = QAction("📤  Export Client Data",      self); act_exp.triggered.connect(self.export_data)
        act_tpl  = QAction("📄  Download Import Template",self); act_tpl.triggered.connect(self.generate_template)
        cm_menu.addAction(act_add)
        cm_menu.addSeparator()
        cm_menu.addAction(act_imp)
        cm_menu.addAction(act_exp)
        cm_menu.addSeparator()
        cm_menu.addAction(act_tpl)

        # Settings menu
        st_menu = menubar.addMenu("Settings")
        act_yr   = QAction("📅  Manage Assessment Years", self); act_yr.triggered.connect(self.open_manage_years)
        act_dir  = QAction("📂  Change Output Folder",    self); act_dir.triggered.connect(self.browse_output_dir)
        st_menu.addAction(act_yr)
        st_menu.addAction(act_dir)
        st_menu.addSeparator()

        # Appearance submenu — built dynamically from THEMES registry
        appear_menu = st_menu.addMenu("🎨  Appearance")
        _icons = {"light": "☀", "dark": "🌙"}
        for theme_key, theme_colors in THEMES.items():
            icon = _icons.get(theme_key, "●")
            act = QAction(f"{icon}  {theme_colors.name}", self, checkable=True)
            act.triggered.connect(lambda _, k=theme_key: self._apply_theme(k))
            setattr(self, f"_theme_{theme_key}_act", act)
            appear_menu.addAction(act)

        # Tools menu
        tools_menu = menubar.addMenu("Tools")
        act_convert = QAction("📊  Convert 26AS TXT → Excel + HTML…", self)
        act_convert.triggered.connect(self._convert_26as_manual)
        tools_menu.addAction(act_convert)
        self._tools_menu = tools_menu

        # Help menu
        help_menu = menubar.addMenu("Help")
        about_action = QAction("About AayDocCapio", self)
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)

        root_widget = QWidget()
        self.setCentralWidget(root_widget)
        root = QVBoxLayout(root_widget)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)

        root.addWidget(self._mk_header())
        root.addWidget(self._mk_main_panel(), 1)
        root.addWidget(self._mk_footer())

    def _apply_theme(self, theme: str):
        """Switch theme by name and persist the choice."""
        global _active_theme
        self._current_theme = theme
        self.vault.update_setting("theme", theme)
        t = get_theme(theme)
        _active_theme = t
        app = QApplication.instance()
        if app:
            app.setStyleSheet(build_stylesheet(t))
        if hasattr(self, "_theme_light_act"):
            for name, action in THEMES.items():
                act = getattr(self, f"_theme_{name}_act", None)
                if act:
                    act.setChecked(name == theme)
        self._repaint_theme(t)

    def _repaint_theme(self, t: ThemeColors = None):
        """Repaint widgets whose colours are set imperatively (not via QSS)."""
        if t is None:
            t = get_theme(self._current_theme)

        # ── Header ────────────────────────────────────────────────────────────
        if hasattr(self, "_hdr_frame"):
            self._hdr_frame.setStyleSheet(
                f"QFrame#header {{ background: {t.bg_window}; border: none; }}"
                f" QLabel {{ border: none; text-decoration: none; }}"
            )
        if hasattr(self, "_hdr_aay"):
            self._hdr_aay.setStyleSheet(
                f"color:{t.text_primary}; font-family:'Avenir Next'; font-size:36px;"
                f" font-weight:700; background:transparent; text-decoration:none; border:none;")
        if hasattr(self, "_hdr_capio"):
            self._hdr_capio.setStyleSheet(
                f"color:{t.accent}; font-family:'Avenir Next'; font-size:36px;"
                f" font-weight:700; background:transparent; text-decoration:none; border:none;")
        if hasattr(self, "_hdr_tm"):
            self._hdr_tm.setStyleSheet(
                f"color:{t.accent}; font-family:'Avenir Next'; font-size:14px;"
                f" font-weight:700; background:transparent; padding-bottom:18px;"
                f" text-decoration:none; border:none;")
        if hasattr(self, "_hdr_sep"):
            self._hdr_sep.setStyleSheet(
                f"color:{t.border}; font-size:22px; background:transparent; border:none;")
        if hasattr(self, "_hdr_tagline"):
            self._hdr_tagline.setStyleSheet(
                f"color:{t.text_muted}; font-family:'Arial'; font-size:13px;"
                f" font-weight:400; background:transparent; border:none;")
        for lbl in (getattr(self, "_hdr_version", None), getattr(self, "_hdr_copy", None)):
            if lbl:
                lbl.setStyleSheet(
                    f"color:{t.text_muted}; font-family:'Arial'; font-size:11px;"
                    f" background:transparent; border:none;")

        # ── Main panel + settings/control bar backgrounds ─────────────────────
        if hasattr(self, "_main_panel"):
            self._main_panel.setStyleSheet(f"background:{t.bg_window};")
        for bar in (getattr(self, "_settings_bar", None), getattr(self, "_control_bar", None)):
            if bar:
                bar.setStyleSheet(f"QFrame{{background:{t.bg_panel};border-radius:10px;}}")

        # ── Settings bar caption labels ───────────────────────────────────────
        for lbl in getattr(self, "_settings_caps", []):
            lbl.setStyleSheet(
                f"color:{t.text_muted};font-size:10px;font-weight:700;"
                f"letter-spacing:0.8px;background:transparent;")

        # ── Search box + status filter ────────────────────────────────────────
        if hasattr(self, "search_box"):
            self.search_box.setStyleSheet(
                f"QLineEdit{{background:{t.bg_input};border:1.5px solid {t.border};"
                f"border-radius:6px;padding:4px 10px;font-size:13px;"
                f"color:{t.text_primary};min-height:28px;}}"
                f"QLineEdit:focus{{border-color:{t.border_focus};background:{t.bg_input_focus};}}"
                f"QLineEdit::placeholder{{color:{t.text_muted};}}"
            )

        # ── dir_lbl + browse button ───────────────────────────────────────────
        if hasattr(self, "dir_lbl"):
            self.dir_lbl.setStyleSheet(
                f"color:{t.text_primary};font-size:12px;background:transparent;")
        if hasattr(self, "_browse_btn"):
            self._browse_btn.setStyleSheet(
                f"QPushButton{{background:transparent;color:{t.text_primary};"
                f"border:1px solid {t.border};border-radius:6px;"
                f"padding:4px 12px;font-weight:bold;font-size:12px;}}"
                f"QPushButton:hover{{background:{t.bg_table_alt};}}"
                f"QPushButton:disabled{{color:{t.text_muted};border-color:{t.border};}}"
            )

        # ── chk_headless ──────────────────────────────────────────────────────
        if hasattr(self, "chk_headless"):
            self.chk_headless.setStyleSheet(
                f"QCheckBox{{font-size:12px;color:{t.text_muted};background:transparent;spacing:6px;}}"
                f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
                f"border-radius:3px;background:{t.bg_checkbox};}}"
                f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};}}")

        # ── Run button ────────────────────────────────────────────────────────
        if hasattr(self, "btn_run"):
            self.btn_run.setStyleSheet(
                f"QToolButton{{background:{t.accent};color:{t.accent_text};border:none;"
                f"border-radius:6px;font-size:13px;font-weight:600;padding:0 14px;}}"
                f"QToolButton:hover{{background:{t.accent_hover};}}"
                f"QToolButton::menu-button{{border:none;width:20px;}}"
                f"QToolButton:disabled{{background:{t.border};color:{t.text_muted};}}")
            if hasattr(self.btn_run, "menu") and self.btn_run.menu():
                self.btn_run.menu().setStyleSheet(
                    f"QMenu{{background:{t.bg_menu};border:1px solid {t.border};"
                    f"border-radius:8px;padding:4px 0;}}"
                    f"QMenu::item{{padding:8px 18px;font-size:13px;color:{t.text_primary};}}"
                    f"QMenu::item:selected{{background:{t.accent};color:{t.accent_text};}}"
                    f"QMenu::separator{{height:1px;background:{t.border};margin:4px 0;}}")

        # ── CLIENTS caption + selected count label ────────────────────────────
        if hasattr(self, "_lbl_clients_cap"):
            self._lbl_clients_cap.setStyleSheet(
                f"color:{t.text_muted};font-size:10px;font-weight:700;"
                f"letter-spacing:0.8px;background:transparent;")
        if hasattr(self, "lbl_selected"):
            cur = self.lbl_selected.text()
            self.lbl_selected.setStyleSheet(
                f"color:{t.accent};font-size:11px;font-weight:bold;background:transparent;")

        # ── Header checkbox (select-all) ──────────────────────────────────────
        if hasattr(self, "header_cb"):
            _chk_path2 = self.checkmark_path.replace("\\", "/")
            self.header_cb.setStyleSheet(
                f"QCheckBox{{background:transparent;}}"
                f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {t.border};"
                f"border-radius:3px;background:{t.bg_checkbox};}}"
                f"QCheckBox::indicator:hover{{border-color:{t.border_focus};}}"
                f"QCheckBox::indicator:checked{{background:{t.accent};border-color:{t.accent};"
                f"image:url('{_chk_path2}');}}"
            )

        # ── Table widget stylesheet + header ──────────────────────────────────
        if hasattr(self, "client_table"):
            chk_path = self.checkmark_path.replace("\\", "/")
            chk_ss = (
                "QCheckBox { background: transparent; }"
                f"QCheckBox::indicator {{ width: 15px; height: 15px; border: 1.5px solid {t.border};"
                f" border-radius: 3px; background: {t.bg_checkbox}; }}"
                f"QCheckBox::indicator:hover {{ border-color: {t.border_focus}; }}"
                f"QCheckBox::indicator:checked {{ background-color: {t.accent}; border-color: {t.accent};"
                f" image: url('{chk_path}'); }}"
            )
            self.client_table.setStyleSheet(
                f"QTableWidget {{ border: 1.5px solid {t.border}; border-radius: 8px;"
                f" background: {t.bg_table}; outline: 0; gridline-color: {t.grid}; }}"
                f"QTableWidget::item {{ border-bottom: 1px solid {t.grid}; padding: 5px; color: {t.text_primary}; }}"
                + chk_ss
            )
            self.client_table.horizontalHeader().setStyleSheet(
                f"QHeaderView::section {{ background-color: {t.bg_header}; border: none;"
                f" border-right: 1px solid {t.border}; border-bottom: 1px solid {t.border};"
                f" font-weight: bold; color: {t.text_muted}; font-size: 11px; height: 34px; }}"
            )
            self.refresh_grid()

        # ── Log box ───────────────────────────────────────────────────────────
        if hasattr(self, "log_box"):
            self.log_box.setStyleSheet(
                f"QTextEdit {{ background: {t.bg_log}; border: none;"
                f" font-family: '{_MONO_FONT}', monospace;"
                f" font-size: 11px; color: {t.text_log}; padding: 8px 16px; }}"
            )

    def _show_about(self):
        import webbrowser
        dlg = QDialog(self)
        dlg.setWindowTitle("About AayDocCapio")
        dlg.setFixedSize(500, 520)
        _ab = _t()
        dlg.setStyleSheet(
            f"QDialog {{ background:{_ab.bg_window}; }}"
            f"QLabel {{ border:none; background:transparent; color:{_ab.text_primary}; }}"
            f"QLabel[link=true] {{ color:{_ab.accent}; }}"
        )

        vl = QVBoxLayout(dlg)
        vl.setContentsMargins(36, 28, 36, 28)
        vl.setSpacing(0)

        # ── App logo + name ───────────────────────────────────────────────────
        logo_row = QHBoxLayout(); logo_row.setSpacing(14)
        icon_lbl = QLabel()
        icon_path = os.path.join(_bundled_dir(), "resources", "app_icon.png")
        if os.path.exists(icon_path):
            icon_lbl.setPixmap(QPixmap(icon_path).scaled(52, 52, Qt.AspectRatioMode.KeepAspectRatio,
                                                          Qt.TransformationMode.SmoothTransformation))
        logo_row.addWidget(icon_lbl)
        name_col = QVBoxLayout(); name_col.setSpacing(2)
        name_lbl = QLabel(
            f'<span style="color:{_ab.text_primary};font-family:\'Avenir Next\';font-size:22px;font-weight:700;">AayDoc </span>'
            f'<span style="color:{_ab.accent};font-family:\'Avenir Next\';font-size:22px;font-weight:700;">Capio™</span>'
        )
        ver_lbl = QLabel(f"Version {APP_VERSION}")
        ver_lbl.setStyleSheet(f"color:{_ab.text_muted}; font-size:12px;")
        name_col.addWidget(name_lbl); name_col.addWidget(ver_lbl)
        logo_row.addLayout(name_col); logo_row.addStretch()
        vl.addLayout(logo_row)
        vl.addSpacing(14)

        desc = QLabel("Automates the secure bulk retrieval of Form 26AS, AIS and TIS directly from the Income Tax e-Filing Portal.")
        desc.setStyleSheet(f"color:{_ab.text_primary}; font-size:13px;")
        desc.setWordWrap(True)
        vl.addWidget(desc)
        vl.addSpacing(12)

        # Name explanation
        name_box = QFrame()
        name_box.setStyleSheet(f"QFrame {{ background:{_ab.bg_panel}; border-radius:6px; border:1px solid {_ab.accent}; }}")
        nb_l = QVBoxLayout(name_box)
        nb_l.setContentsMargins(14, 10, 14, 10)
        nb_l.setSpacing(6)
        name_head = QLabel("Why AayDoc Capio?")
        name_head.setStyleSheet(f"color:{_ab.accent}; font-size:11px; font-weight:700; letter-spacing:0.5px; background:transparent; border:none;")
        name_exp = QLabel(
            f'<b style="color:{_ab.text_primary};">Aay</b> (Income) · '
            f'<b style="color:{_ab.text_primary};">Doc</b> (Documents) · '
            f'<b style="color:{_ab.accent};">Capio</b> <span style="color:{_ab.text_muted};">(Latin: To Obtain)</span>'
        )
        name_exp.setStyleSheet("font-size:13px; background:transparent; border:none;")
        name_sub = QLabel("AayDoc Capio is designed to securely retrieve and deliver income tax documents, "
                          "eliminating repetitive manual downloads and improving efficiency for tax professionals.")
        name_sub.setStyleSheet(f"color:{_ab.text_primary}; font-size:12px; background:transparent; border:none;")
        name_sub.setWordWrap(True)
        nb_l.addWidget(name_head)
        nb_l.addWidget(name_exp)
        nb_l.addWidget(name_sub)
        vl.addWidget(name_box)
        vl.addSpacing(16)

        # ── Divider ───────────────────────────────────────────────────────────
        div1 = QFrame(); div1.setFrameShape(QFrame.Shape.HLine)
        div1.setStyleSheet(f"background:{_ab.border}; border:none; max-height:1px;")
        vl.addWidget(div1)
        vl.addSpacing(16)

        # ── Developer info ────────────────────────────────────────────────────
        dev_title = QLabel("Contact Us")
        dev_title.setStyleSheet(f"color:{_ab.text_muted}; font-size:10px; font-weight:700; letter-spacing:1px;")
        vl.addWidget(dev_title)
        vl.addSpacing(8)

        def _link_row(icon_file, display_text, url=None):
            row = QHBoxLayout(); row.setSpacing(10); row.setContentsMargins(0,0,0,0)
            icon_l = QLabel()
            icon_l.setFixedSize(22, 22)
            icon_path = os.path.join(_bundled_dir(), "resources", icon_file)
            if os.path.exists(icon_path):
                icon_l.setPixmap(QPixmap(icon_path).scaled(22, 22, Qt.AspectRatioMode.KeepAspectRatio,
                                                            Qt.TransformationMode.SmoothTransformation))
            icon_l.setStyleSheet("background:transparent; border:none;")
            row.addWidget(icon_l)
            if url:
                lbl = QLabel(f'<a href="{url}" style="color:{_ab.accent}; text-decoration:none;">{display_text}</a>')
                lbl.setOpenExternalLinks(True)
                lbl.setStyleSheet("background:transparent; border:none; font-size:13px;")
            else:
                lbl = QLabel(display_text)
                lbl.setStyleSheet(f"color:{_ab.text_primary}; font-size:13px; font-weight:600; background:transparent; border:none;")
            row.addWidget(lbl)
            row.addStretch()
            return row

        vl.addLayout(_link_row("icon_person.png",   "CA. Deepak Bhholusaria"))
        vl.addSpacing(6)
        vl.addLayout(_link_row("icon_email.png",    "deepak@ailearrning.guru",          "mailto:deepak@ailearrning.guru"))
        vl.addSpacing(6)
        vl.addLayout(_link_row("icon_linkedin.png", "linkedin.com/in/bhholusaria",      "https://www.linkedin.com/in/bhholusaria/"))
        vl.addSpacing(6)
        vl.addLayout(_link_row("icon_vcard.png",    "www.ailearrning.guru",             "https://www.ailearrning.guru"))
        vl.addSpacing(6)
        vl.addLayout(_link_row("icon_vcard.png",    "Virtual Card",                     "https://www.qrcodechimp.com/page/deepakb?chk1668183417"))
        vl.addSpacing(16)

        # ── Divider ───────────────────────────────────────────────────────────
        div2 = QFrame(); div2.setFrameShape(QFrame.Shape.HLine)
        div2.setStyleSheet(f"background:{_ab.border}; border:none; max-height:1px;")
        vl.addWidget(div2)
        vl.addSpacing(12)

        copy = QLabel("© 2026 Deepak Bhholusaria. All rights reserved.")
        copy.setStyleSheet(f"color:{_ab.text_muted}; font-size:11px;")
        vl.addWidget(copy)
        vl.addStretch()

        # ── Close button ──────────────────────────────────────────────────────
        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(100)
        close_btn.setStyleSheet(
            f"QPushButton {{ background:{_ab.accent}; color:{_ab.accent_text}; border:none; border-radius:6px; padding:8px 16px; font-size:13px; }}"
            f"QPushButton:hover {{ background:{_ab.accent_hover}; }}")
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout(); btn_row.addStretch(); btn_row.addWidget(close_btn)
        vl.addLayout(btn_row)

        dlg.exec()

    def _mk_header(self):
        hdr = QFrame()
        hdr.setFixedHeight(80)
        hdr.setObjectName("header")
        self._hdr_frame = hdr
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(32, 0, 32, 0)
        hl.setSpacing(0)

        # App icon — large enough to anchor the header
        icon_label = QLabel()
        icon_path = os.path.join(_bundled_dir(), "resources", "app_icon.png")
        if os.path.exists(icon_path):
            icon_label.setPixmap(
                QPixmap(icon_path).scaled(62, 62, Qt.AspectRatioMode.KeepAspectRatio,
                                          Qt.TransformationMode.SmoothTransformation)
            )
        hl.addWidget(icon_label)
        hl.addSpacing(18)

        # Name + tagline stacked
        name_block = QWidget()
        name_block.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(name_block)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(3)

        title_row = QWidget()
        title_row.setStyleSheet("background:transparent;")
        tl = QHBoxLayout(title_row)
        tl.setContentsMargins(0, 0, 0, 0)
        tl.setSpacing(0)
        tl.setAlignment(Qt.AlignmentFlag.AlignVCenter)

        aay = QLabel("AayDoc ")
        self._hdr_aay = aay
        tl.addWidget(aay)

        capio = QLabel("Capio")
        self._hdr_capio = capio
        tl.addWidget(capio)

        tm = QLabel("™")
        self._hdr_tm = tm
        tl.addWidget(tm)

        # Separator + tagline inline with title
        sep = QLabel("  |  ")
        self._hdr_sep = sep
        sep.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        tl.addWidget(sep)

        tagline = QLabel("Tax Documents. Delivered to You.")
        self._hdr_tagline = tagline
        tagline.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        tl.addWidget(tagline)
        tl.addStretch()

        title = title_row

        vl.addStretch()
        vl.addWidget(title)
        vl.addStretch()

        hl.addWidget(name_block)
        hl.addStretch()

        # Copyright + version on the right
        meta_block = QWidget()
        meta_block.setStyleSheet("background:transparent;")
        ml = QVBoxLayout(meta_block)
        ml.setContentsMargins(0, 0, 0, 0)
        ml.setSpacing(2)
        ml.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        version_lbl = QLabel(f"v{APP_VERSION}")
        self._hdr_version = version_lbl
        version_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)

        copy_lbl = QLabel("© 2026 Deepak Bhholusaria")
        self._hdr_copy = copy_lbl
        copy_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)

        ml.addStretch()
        ml.addWidget(version_lbl)
        ml.addWidget(copy_lbl)
        ml.addStretch()

        hl.addWidget(meta_block)
        hl.addSpacing(12)

        # ⓘ About button
        about_btn = QPushButton("ⓘ")
        about_btn.setFixedSize(32, 32)
        about_btn.setToolTip("About AayDocCapio")
        about_btn.setStyleSheet(
            "QPushButton { background:transparent; border:none; font-size:20px; color:#DC2626; }"
            "QPushButton:hover { color:#B91C1C; }")
        about_btn.clicked.connect(self._show_about)
        hl.addWidget(about_btn, 0, Qt.AlignmentFlag.AlignVCenter)

        return hdr

    def _mk_main_panel(self):
        panel = QWidget()
        self._main_panel = panel
        panel.setStyleSheet(f"background:{_t().bg_window};")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(14, 10, 14, 8)
        layout.setSpacing(6)

        settings = self._mk_settings_bar()
        settings.setGraphicsEffect(_shadow(18, 3, 18))
        layout.addWidget(settings)

        # Grid label row
        grid_hdr_row = QHBoxLayout()
        self._lbl_clients_cap = _lbl("CLIENTS", 10, bold=True, color=_t().text_muted)
        grid_hdr_row.addWidget(self._lbl_clients_cap)
        grid_hdr_row.addStretch()
        self.lbl_selected = _lbl("0 selected", 11, bold=True, color=_t().accent)
        grid_hdr_row.addWidget(self.lbl_selected)
        layout.addLayout(grid_hdr_row)

        # Search / filter bar
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)
        filter_row.setContentsMargins(0, 0, 0, 0)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍  Search by name or PAN...")
        self.search_box.setFixedHeight(28)
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setAttribute(Qt.WidgetAttribute.WA_MacShowFocusRect, False)
        self.search_box.textChanged.connect(self._apply_filter)
        filter_row.addWidget(self.search_box, 1)

        self.status_filter = StyledComboBox()
        self.status_filter.addItems([
            "All statuses",
            "✅  Downloaded",
            "⚠  Partially Completed",
            "❌  Failed",
            "🕐  Queued / Pending",
            "—  Not run yet",
        ])
        self.status_filter.setFixedHeight(28)
        self.status_filter.setFixedWidth(185)
        self.status_filter.currentIndexChanged.connect(lambda _: self._apply_filter(self.search_box.text()))
        filter_row.addWidget(self.status_filter)

        layout.addLayout(filter_row)

        layout.addWidget(self._mk_client_table(), 1)

        ctrl = self._mk_control_bar()
        ctrl.setGraphicsEffect(_shadow(18, 3, 18))
        layout.addWidget(ctrl)
        return panel

    def _mk_settings_bar(self):
        bar = QFrame()
        self._settings_bar = bar
        bar.setFixedHeight(68)
        bar.setStyleSheet(f"QFrame{{background:{_t().bg_panel};border-radius:10px;}}")
        hl = QHBoxLayout(bar)
        hl.setContentsMargins(20, 0, 20, 0)
        hl.setSpacing(0)
        hl.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        self._settings_caps = []

        def _cap(text):
            l = QLabel(text)
            self._settings_caps.append(l)
            return l

        def _divider():
            d = QFrame(); d.setFrameShape(QFrame.Shape.VLine)
            d.setFixedSize(1, 32)
            d.setStyleSheet(f"background:{_t().border};border:none;")
            return d

        # ── Assessment Year ───────────────────────────────────────────────────
        self._ay_entries = self._load_ay_list()
        ay_labels = [e["label"] for e in self._ay_entries if e.get("enabled", True)]
        saved_ay = self.vault.get_setting("assessment_year", "Select AY/TY")
        self.ay_combo = StyledComboBox()
        self.ay_combo.addItem("Select AY/TY")
        self.ay_combo.addItems(ay_labels)
        self.ay_combo.setCurrentText(saved_ay if saved_ay in ay_labels else "Select AY/TY")
        self.ay_combo.setFixedWidth(220)
        self.ay_combo.currentTextChanged.connect(self.save_ay_setting)
        self.ay_combo.currentTextChanged.connect(lambda _: self.refresh_grid())

        manage_btn = QPushButton("⚙")
        manage_btn.setFixedSize(24, 24)
        manage_btn.setToolTip("Manage Years")
        manage_btn.setStyleSheet(
            f"QPushButton{{background:transparent;border:none;font-size:14px;color:{_t().text_muted};}}"
            f"QPushButton:hover{{color:{_t().accent};}}")
        manage_btn.clicked.connect(self.open_manage_years)

        ay_col = QWidget(); ay_col.setStyleSheet("background:transparent;")
        ay_vl = QVBoxLayout(ay_col); ay_vl.setContentsMargins(0,0,0,0); ay_vl.setSpacing(2)
        ay_vl.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        ay_vl.addWidget(_cap("ASSESSMENT YEAR"))
        ay_hl = QHBoxLayout(); ay_hl.setContentsMargins(0,0,0,0); ay_hl.setSpacing(5)
        ay_hl.addWidget(self.ay_combo); ay_hl.addWidget(manage_btn)
        ay_vl.addLayout(ay_hl)
        hl.addWidget(ay_col)

        hl.addSpacing(28); hl.addWidget(_divider()); hl.addSpacing(28)

        # ── Output Directory ──────────────────────────────────────────────────
        # Prefer Windows path whenever USERPROFILE is set (native Windows or WSL).
        # A saved /home/... path from a WSL dev session must not be shown when
        # the Windows Downloads folder is reachable.
        _saved_dir = self.vault.get_setting("download_root_dir", "")
        _win_path = _default_download_dir() if os.environ.get("USERPROFILE") else None
        if _win_path and _win_path != _saved_dir:
            # USERPROFILE is available — always use the Windows-derived path
            default_dir = _win_path
            self.vault.update_setting("download_root_dir", default_dir)
        elif _saved_dir and os.path.isdir(_saved_dir):
            default_dir = _saved_dir
        else:
            default_dir = _default_download_dir()
            self.vault.update_setting("download_root_dir", default_dir)
        self.dir_lbl = QLabel(default_dir)
        self.dir_lbl.setStyleSheet(f"color:{_t().text_primary};font-size:12px;background:transparent;")
        self.dir_lbl.setWordWrap(False)
        self.dir_lbl.setMaximumWidth(320)
        self.dir_lbl.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        browse_btn = _btn("📂  Browse", "outline", height=26)
        self._browse_btn = browse_btn
        browse_btn.clicked.connect(self.browse_output_dir)

        dir_row = QHBoxLayout(); dir_row.setContentsMargins(0,0,0,0); dir_row.setSpacing(8)
        dir_row.addWidget(self.dir_lbl); dir_row.addWidget(browse_btn)

        dir_col = QWidget(); dir_col.setStyleSheet("background:transparent;")
        dir_col.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
        dir_vl = QVBoxLayout(dir_col); dir_vl.setContentsMargins(0,0,0,0); dir_vl.setSpacing(2)
        dir_vl.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        dir_vl.addWidget(_cap("OUTPUT DIRECTORY"))
        dir_vl.addLayout(dir_row)
        hl.addWidget(dir_col)
        hl.addStretch()

        return bar

    # Column indices for client table
    _TC_CHK    = 0
    _TC_NAME   = 1
    _TC_PAN    = 2
    _TC_DOB    = 3
    _TC_STATUS = 4
    _TC_PATH   = 5
    _TC_ACTS   = 6

    def _mk_client_table(self):
        self.client_table = QTableWidget(0, 7)
        self.client_table.setHorizontalHeaderLabels([
            "", "Name  ⇅", "PAN  ⇅", "Date of Birth",
            "Last Download Status", "Last Saved Location", ""
        ])

        _tbl = _t()
        self.client_table.horizontalHeader().setStyleSheet(
            f"QHeaderView::section {{ background-color: {_tbl.bg_header}; border: none; "
            f"border-right: 1px solid {_tbl.border}; border-bottom: 1px solid {_tbl.border}; "
            f"font-weight: bold; color: {_tbl.text_muted}; font-size: 11px; height: 34px; }}"
        )
        self.client_table.verticalHeader().setVisible(False)
        self.client_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.client_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.client_table.setShowGrid(True)
        self.client_table.setAlternatingRowColors(False)

        chk_path = self.checkmark_path.replace("\\", "/")
        checkbox_style = (
            "QCheckBox { background: transparent; }"
            f"QCheckBox::indicator {{ width: 15px; height: 15px; border: 1.5px solid {_tbl.border}; border-radius: 3px; background: {_tbl.bg_checkbox}; }}"
            f"QCheckBox::indicator:hover {{ border-color: {_tbl.border_focus}; }}"
            f"QCheckBox::indicator:checked {{ background-color: {_tbl.accent}; border-color: {_tbl.accent}; image: url('{chk_path}'); }}"
        )
        self.client_table.setStyleSheet(
            f"QTableWidget {{ border: 1.5px solid {_tbl.border}; border-radius: 8px; background: {_tbl.bg_table}; outline: 0; gridline-color: {_tbl.grid}; }}"
            f"QTableWidget::item {{ border-bottom: 1px solid {_tbl.grid}; padding: 5px; }}"
            + checkbox_style
        )

        for col, align in [
            (self._TC_NAME,   Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            (self._TC_PAN,    Qt.AlignmentFlag.AlignCenter),
            (self._TC_DOB,    Qt.AlignmentFlag.AlignCenter),
            (self._TC_STATUS, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            (self._TC_PATH,   Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            (self._TC_ACTS,   Qt.AlignmentFlag.AlignCenter),
        ]:
            item = self.client_table.horizontalHeaderItem(col)
            if item:
                item.setTextAlignment(align)

        self.client_table.setColumnWidth(self._TC_CHK,    45)
        self.client_table.setColumnWidth(self._TC_PAN,   130)
        self.client_table.setColumnWidth(self._TC_DOB,   120)
        self.client_table.setColumnWidth(self._TC_STATUS, 170)
        self.client_table.setColumnWidth(self._TC_ACTS,   52)

        header = self.client_table.horizontalHeader()
        header.setSectionResizeMode(self._TC_CHK,    QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(self._TC_NAME,   QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(self._TC_PAN,    QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(self._TC_DOB,    QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(self._TC_STATUS, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(self._TC_PATH,   QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(self._TC_ACTS,   QHeaderView.ResizeMode.Interactive)

        header.setSortIndicatorShown(False)
        header.sectionClicked.connect(self._on_header_clicked)
        self._current_sort_col = -1
        self._current_sort_order = Qt.SortOrder.AscendingOrder

        self.client_table.cellClicked.connect(self._on_cell_clicked)
        self.client_table.setMouseTracking(True)
        self.client_table.viewport().setMouseTracking(True)
        self.client_table.cellEntered.connect(self._on_cell_entered)
        self._hovered_acts_row = -1

        self.header_cb = QCheckBox(header)
        self.header_cb.setFixedSize(18, 18)
        self.header_cb.setStyleSheet(checkbox_style)
        self.header_cb.toggled.connect(self.toggle_select_all)
        header.geometriesChanged.connect(self._position_header_checkbox)

        self.client_table.setMinimumHeight(200)
        return self.client_table

    def _position_header_checkbox(self):
        if not hasattr(self, "client_table") or not hasattr(self, "header_cb"):
            return
        header = self.client_table.horizontalHeader()
        x = header.sectionPosition(0)
        w = header.sectionSize(0)
        h = header.height()
        cb_x = x + (w - 18) // 2
        cb_y = (h - 18) // 2
        self.header_cb.move(cb_x, cb_y)

    def _on_cell_clicked(self, row, col):
        if hasattr(self, "ay_combo") and self.ay_combo._popup_was_open:
            return

        if col == self._TC_ACTS:
            item = self.client_table.item(row, self._TC_ACTS)
            if not item:
                return
            a = item.data(Qt.ItemDataRole.UserRole + 1)
            if not a:
                return
            a_id = a.get("id")
            _mt = _t()
            menu = QMenu(self)
            menu.setStyleSheet(
                f"QMenu {{ background:{_mt.bg_menu}; border:1px solid {_mt.border}; border-radius:6px; padding:4px 0; }}"
                f"QMenu::item {{ padding:8px 20px; font-size:12px; color:{_mt.text_primary}; }}"
                f"QMenu::item:selected {{ background:{_mt.accent}; color:{_mt.accent_text}; }}"
                f"QMenu::separator {{ height:1px; background:{_mt.border}; margin:3px 0; }}"
            )
            menu.addAction("✏  Edit Client",   lambda av=a:   self._open_edit_client(av))
            menu.addSeparator()
            menu.addAction("🗑  Delete Client", lambda id_=a_id: self.delete_assessee(id_))
            # Show menu at the cell's bottom-left corner
            rect = self.client_table.visualItemRect(item)
            pos  = self.client_table.viewport().mapToGlobal(rect.bottomLeft())
            menu.exec(pos)
            return

        if col in (self._TC_CHK, self._TC_PATH):
            return
        # Toggle checkbox for any other column click
        cb_container = self.client_table.cellWidget(row, self._TC_CHK)
        if cb_container:
            cb = cb_container.findChild(QCheckBox)
            if cb:
                cb.setChecked(not cb.isChecked())

    def _on_cell_entered(self, row, col):
        pass

    def _on_header_clicked(self, logical_index):
        if logical_index not in (self._TC_NAME, self._TC_PAN):
            return
            
        header = self.client_table.horizontalHeader()
        
        # Toggle or initialize sort
        if self._current_sort_col == logical_index:
            self._current_sort_order = (
                Qt.SortOrder.DescendingOrder 
                if self._current_sort_order == Qt.SortOrder.AscendingOrder 
                else Qt.SortOrder.AscendingOrder
            )
        else:
            self._current_sort_col = logical_index
            self._current_sort_order = Qt.SortOrder.AscendingOrder
            
        # Visually show active sort indicator
        header.setSortIndicatorShown(True)
        header.setSortIndicator(self._current_sort_col, self._current_sort_order)
        
        # Perform sort
        self.client_table.setSortingEnabled(True)
        self.client_table.sortByColumn(self._current_sort_col, self._current_sort_order)
        self.client_table.setSortingEnabled(False)
        
        for row_idx in range(self.client_table.rowCount()):
            item = self.client_table.item(row_idx, self._TC_NAME)
            if item:
                row_id = item.data(Qt.ItemDataRole.UserRole)
                row_selected = row_id in self.selected_ids
                self._apply_row_style(row_idx, row_selected, row_idx)

    def _mk_control_bar(self):
        bar = QFrame()
        self._control_bar = bar
        bar.setFixedHeight(54)
        bar.setStyleSheet(f"QFrame{{background:{_t().bg_panel};border-radius:10px;}}")
        hl = QHBoxLayout(bar)
        hl.setContentsMargins(22, 0, 16, 0)

        # Headless toggle — when checked, the automation browser runs hidden.
        # Default ON; uncheck to watch progress or handle a CAPTCHA.
        self.chk_headless = QCheckBox("Run in background (hide browser)")
        self.chk_headless.setChecked(True)
        self.chk_headless.setToolTip(
            "When ON, the automation Chrome window is hidden (headless).\n"
            "Keep OFF to watch progress or handle any CAPTCHA.")
        _ck = _t()
        self.chk_headless.setStyleSheet(
            f"QCheckBox{{font-size:12px;color:{_ck.text_muted};background:transparent;spacing:6px;}}"
            f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {_ck.border};"
            f"border-radius:3px;background:{_ck.bg_checkbox};}}"
            f"QCheckBox::indicator:checked{{background:{_ck.accent};border-color:{_ck.accent};}}")
        hl.addWidget(self.chk_headless)

        hl.addStretch()

        self.btn_delete_sel = _btn("🗑  Delete Selected", "danger", height=34, min_width=130)
        self.btn_delete_sel.setEnabled(False)
        self.btn_delete_sel.clicked.connect(self.delete_selected)
        hl.addWidget(self.btn_delete_sel)
        hl.addSpacing(8)

        # ── Run dropdown (split-style: label + arrow) ─────────────────────────
        self.btn_run = QToolButton()
        self.btn_run.setText("▶  Run")
        self.btn_run.setPopupMode(QToolButton.ToolButtonPopupMode.MenuButtonPopup)
        self.btn_run.setFixedHeight(34)
        self.btn_run.setMinimumWidth(110)
        self.btn_run.setStyleSheet(
            "QToolButton{"
            "  background:#16A34A; color:#FFFFFF; border:none;"
            "  border-radius:8px; font-size:13px; font-weight:600; padding:0 14px;"
            "}"
            "QToolButton:hover{ background:#15803D; }"
            "QToolButton:disabled{ background:#D1FAE5; color:#6EE7B7; }"
            "QToolButton::menu-button{"
            "  border-left:1px solid rgba(255,255,255,0.35);"
            "  border-radius:0 8px 8px 0; width:20px;"
            "}"
            "QToolButton::menu-arrow{ image:none; }"
        )

        run_menu = QMenu(self.btn_run)
        _rm = _t()
        run_menu.setStyleSheet(
            f"QMenu{{ background:{_rm.bg_menu}; border:1px solid {_rm.border}; border-radius:8px; padding:4px 0; }}"
            f"QMenu::item{{ padding:8px 18px; font-size:13px; color:{_rm.text_primary}; }}"
            f"QMenu::item:selected{{ background:{_rm.accent}; color:{_rm.accent_text}; }}"
            f"QMenu::separator{{ height:1px; background:{_rm.border}; margin:4px 0; }}"
        )

        act_26as = QAction("▶  Download 26AS", self)
        act_26as.setToolTip("Downloads Form 26AS PDF + TXT for selected clients.")
        act_26as.triggered.connect(lambda: self.start_automation("26as"))

        act_request_ais = QAction("📋  Download / Request TIS & AIS", self)
        act_request_ais.setToolTip(
            "Opens AIS portal for each client.\n"
            "• If AIS is ready — downloads instantly.\n"
            "• If not ready — places generation request (~5 min on ITD servers).")
        act_request_ais.triggered.connect(lambda: self.start_automation("request_ais"))

        act_dl_ais = QAction("⬇  Download Previously Requested AIS", self)
        act_dl_ais.setToolTip(
            "Fetches AIS PDF from Activity History for clients\n"
            "whose AIS was requested earlier and is now ready.")
        act_dl_ais.triggered.connect(lambda: self.start_automation("ais_tis"))

        run_menu.addAction(act_26as)
        run_menu.addSeparator()
        run_menu.addAction(act_request_ais)
        run_menu.addAction(act_dl_ais)

        self.btn_run.setMenu(run_menu)
        self.btn_run.clicked.connect(lambda: self.btn_run.showMenu())
        hl.addWidget(self.btn_run)

        # ── AIS status line (hidden until Request AIS runs) ───────────────────
        self.ais_status_bar = QFrame()
        self.ais_status_bar.setFixedHeight(28)
        self.ais_status_bar.setStyleSheet(
            "QFrame{background:#FFF7ED;border-top:1px solid #FED7AA;}")
        self.ais_status_bar.setVisible(False)
        asl = QHBoxLayout(self.ais_status_bar)
        asl.setContentsMargins(22, 0, 16, 0)
        self.ais_status_lbl = QLabel()
        self.ais_status_lbl.setStyleSheet(
            "color:#92400E; font-size:11px; background:transparent;")
        asl.addWidget(self.ais_status_lbl)
        asl.addStretch()
        ais_dismiss = QPushButton("✕")
        ais_dismiss.setFixedSize(18, 18)
        ais_dismiss.setStyleSheet(
            "QPushButton{background:transparent;border:none;"
            "color:#92400E;font-size:10px;}"
            "QPushButton:hover{color:#78350F;}")
        ais_dismiss.clicked.connect(lambda: self.ais_status_bar.setVisible(False))
        asl.addWidget(ais_dismiss)

        # Wrap bar + status into a column so status sits just below the bar
        container = QWidget()
        container.setStyleSheet("background:transparent;")
        col = QVBoxLayout(container)
        col.setContentsMargins(0, 0, 0, 0)
        col.setSpacing(0)
        col.addWidget(bar)
        col.addWidget(self.ais_status_bar)
        return container

    def _mk_footer(self):
        footer = QFrame()
        footer.setFixedHeight(190)
        footer.setStyleSheet("QFrame{background:#0F172A;}")
        fl = QVBoxLayout(footer)
        fl.setContentsMargins(0, 0, 0, 0)
        fl.setSpacing(0)

        log_hdr = QFrame()
        log_hdr.setFixedHeight(32)
        log_hdr.setStyleSheet("QFrame{background:#1E293B;}")
        hhl = QHBoxLayout(log_hdr); hhl.setContentsMargins(16, 0, 12, 0)
        dot = QLabel("●")
        dot.setStyleSheet("color:#22C55E; font-size:9px; margin-right:4px;")
        hhl.addWidget(dot)
        hhl.addWidget(_lbl("LIVE LOGS", 10, bold=True, color="#64748B"))
        hhl.addStretch()
        copy_btn = QPushButton("Copy")
        copy_btn.setFixedHeight(22)
        copy_btn.setStyleSheet(
            f"QPushButton{{background:transparent;color:{_t().text_muted};border:1px solid {_t().border};"
            f"border-radius:4px;padding:0 10px;font-size:10px;}}"
            f"QPushButton:hover{{color:{_t().text_primary};border-color:{_t().text_muted};}}")
        copy_btn.clicked.connect(self.copy_logs_to_clipboard)
        hhl.addWidget(copy_btn)
        fl.addWidget(log_hdr)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(
            "QTextEdit{background:#0F172A;border:none;"
            f"font-family:'{_MONO_FONT}',monospace;"
            "font-size:11px;color:#7DD3FC;padding:8px 16px;}")
        fl.addWidget(self.log_box)
        return footer

    # ── Grid ──────────────────────────────────────────────────────────────────

    def _apply_row_style(self, row_idx, selected, index=0):
        t  = get_theme(getattr(self, "_current_theme", "light"))
        bg = t.bg_table if index % 2 == 0 else t.bg_table_alt
        fg = t.row_selected_fg if selected else t.text_primary
        
        # Apply style to all items in the row
        for col in range(self.client_table.columnCount()):
            item = self.client_table.item(row_idx, col)
            if item:
                item.setBackground(QColor(bg))
                item.setForeground(QColor(fg))
                # Set font
                font = item.font()
                # PAN (col 2) is bold by default, or bold if selected
                font.setBold(col == 2 or selected)
                item.setFont(font)
        
        cb_container = self.client_table.cellWidget(row_idx, self._TC_CHK)
        if cb_container:
            cb_container.setStyleSheet(f"background:{bg}; border:none;")
            cb = cb_container.findChild(QCheckBox)
            if cb:
                cb.setStyleSheet("background:transparent;")

        # _TC_ACTS is a plain QTableWidgetItem — styled via item background/foreground above

    def _apply_filter(self, text=""):
        if not hasattr(self, "client_table"):
            return
        q = text.strip().lower()
        sf_idx = self.status_filter.currentIndex() if hasattr(self, "status_filter") else 0
        # Map dropdown index → status prefix(es) to match
        _sf_map = {
            0: None,              # All statuses
            1: ("✅",),           # Downloaded
            2: ("⚠",),           # Partially Completed
            3: ("❌",),           # Failed
            4: ("🕐", "⏹"),      # Queued / Pending
            5: ("—",),            # Not run yet
        }
        status_prefixes = _sf_map.get(sf_idx)
        for row_idx in range(self.client_table.rowCount()):
            name_item   = self.client_table.item(row_idx, self._TC_NAME)
            pan_item    = self.client_table.item(row_idx, self._TC_PAN)
            status_item = self.client_table.item(row_idx, self._TC_STATUS)
            if not name_item or not pan_item:
                continue
            text_match = (not q
                          or q in name_item.text().lower()
                          or q in pan_item.text().lower())
            if status_prefixes is None:
                status_match = True
            else:
                st = (status_item.text() if status_item else "—")
                status_match = any(st.startswith(p) for p in status_prefixes)
            self.client_table.setRowHidden(row_idx, not (text_match and status_match))

    def refresh_grid(self):
        self._checkbox_map.clear()
        self.assessee_list = self.vault.get_all_assessees()
        
        if not hasattr(self, "client_table"):
            return
            
        # Block signals on header_cb during refresh
        if hasattr(self, "header_cb"):
            self.header_cb.blockSignals(True)
            self.header_cb.setEnabled(False)
            
        self.client_table.setRowCount(0)
        
        if not self.assessee_list:
            if hasattr(self, "header_cb"):
                self.header_cb.setChecked(False)
                self.header_cb.blockSignals(False)
            self._update_count()
            return

        if hasattr(self, "header_cb"):
            self.header_cb.setEnabled(True)
            
        # Load download history for currently selected AY
        current_ay = self.ay_combo.currentText() if hasattr(self, "ay_combo") else ""
        dl_history = {}
        if current_ay and current_ay != "Select AY/TY":
            try:
                dl_history = self.vault.get_download_history(current_ay)
            except Exception:
                pass

        for i, a in enumerate(self.assessee_list):
            a_id = a.get("id")
            pan  = a.get("pan", "")
            is_selected = a_id in self.selected_ids

            self.client_table.insertRow(i)

            # Col 0: Checkbox
            _ct = _t()
            _chk_path = self.checkmark_path.replace("\\", "/")
            cb_container = QWidget()
            cb_container.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            cb_container.setStyleSheet(f"QWidget{{background:{_ct.bg_table};}}")
            cb_layout = QHBoxLayout(cb_container)
            cb_layout.setContentsMargins(0, 0, 0, 0)
            cb_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cb = QCheckBox()
            cb.setStyleSheet(
                f"QCheckBox{{background:transparent;}}"
                f"QCheckBox::indicator{{width:15px;height:15px;border:1.5px solid {_ct.border};"
                f"border-radius:3px;background:{_ct.bg_checkbox};}}"
                f"QCheckBox::indicator:hover{{border-color:{_ct.border_focus};}}"
                f"QCheckBox::indicator:checked{{background:{_ct.accent};border-color:{_ct.accent};"
                f"image:url('{_chk_path}');}}"
            )
            cb.setChecked(is_selected)
            cb.toggled.connect(lambda checked, id_=a_id: self._on_check(id_, checked))
            self._checkbox_map[a_id] = cb
            cb_layout.addWidget(cb)
            self.client_table.setCellWidget(i, self._TC_CHK, cb_container)

            # Col 1: Name
            name_item = QTableWidgetItem(a.get("name", ""))
            name_item.setData(Qt.ItemDataRole.UserRole, a_id)
            name_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            self.client_table.setItem(i, self._TC_NAME, name_item)

            # Col 2: PAN (monospace)
            pan_item = QTableWidgetItem(pan)
            pan_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            f = pan_item.font(); f.setFamily(_MONO_FONT); pan_item.setFont(f)
            self.client_table.setItem(i, self._TC_PAN, pan_item)

            # Col 3: Date of Birth
            dob_item = QTableWidgetItem(a.get("dob", ""))
            dob_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.client_table.setItem(i, self._TC_DOB, dob_item)

            # Col 4: Last Download Status (from history)
            hist = dl_history.get(pan, {})
            status_text = hist.get("status", "—")
            status_item = QTableWidgetItem(status_text)
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            status_item.setToolTip(status_text)
            if status_text.startswith("✅"):
                status_item.setForeground(QColor("#15803D"))
            elif status_text.startswith("❌"):
                status_item.setForeground(QColor("#DC2626"))
            elif status_text.startswith("⚠"):
                status_item.setForeground(QColor("#D97706"))
            else:
                status_item.setForeground(QColor("#64748B"))
            self.client_table.setItem(i, self._TC_STATUS, status_item)

            # Col 5: Last Saved Location (hyperlink QLabel)
            saved_path = hist.get("path", "")
            path_lbl = QLabel()
            path_lbl.setContentsMargins(6, 0, 6, 0)
            path_lbl.setStyleSheet("background:transparent; border:none; font-size:11px;")
            if saved_path and os.path.isdir(saved_path):
                path_lbl.setText(
                    f'<a href="{saved_path}" style="color:#1D4ED8;text-decoration:underline;">'
                    f'{saved_path}</a>'
                )
                path_lbl.setToolTip(saved_path)
                path_lbl.linkActivated.connect(lambda p=saved_path: _open_path(p))
            else:
                path_lbl.setText('<span style="color:#94A3B8;">—</span>')
            self.client_table.setCellWidget(i, self._TC_PATH, path_lbl)

            # Col 6: Actions
            dots_item = QTableWidgetItem("• • •")
            dots_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            dots_item.setToolTip("Click for actions")
            dots_item.setForeground(QColor("#64748B"))
            f = dots_item.font()
            f.setPointSize(9)
            f.setBold(True)
            dots_item.setFont(f)
            dots_item.setData(Qt.ItemDataRole.UserRole + 1, a)
            self.client_table.setItem(i, self._TC_ACTS, dots_item)

            self._apply_row_style(i, is_selected, i)

        # Re-apply active sort if any
        if self._current_sort_col in (self._TC_NAME, self._TC_PAN):
            self.client_table.setSortingEnabled(True)
            self.client_table.sortByColumn(self._current_sort_col, self._current_sort_order)
            self.client_table.setSortingEnabled(False)

            for row_idx in range(self.client_table.rowCount()):
                item = self.client_table.item(row_idx, self._TC_NAME)
                if item:
                    row_id = item.data(Qt.ItemDataRole.UserRole)
                    row_selected = row_id in self.selected_ids
                    self._apply_row_style(row_idx, row_selected, row_idx)
                    
        if hasattr(self, "header_cb"):
            self.header_cb.blockSignals(False)
            
        self._update_count()
        if hasattr(self, "search_box"):
            self._apply_filter(self.search_box.text())

    def _on_check(self, id_, checked):
        if checked:
            self.selected_ids.add(id_)
        else:
            self.selected_ids.discard(id_)
            
        # Find the row in table widget and apply selection style
        if hasattr(self, "client_table"):
            for row_idx in range(self.client_table.rowCount()):
                item = self.client_table.item(row_idx, self._TC_NAME)
                if item and item.data(Qt.ItemDataRole.UserRole) == id_:
                    self._apply_row_style(row_idx, checked, row_idx)
                    break
                    
        self._update_count()

    def toggle_select_all(self, checked):
        for a_id, cb in self._checkbox_map.items():
            cb.blockSignals(True)
            cb.setChecked(checked)
            cb.blockSignals(False)
            if checked:
                self.selected_ids.add(a_id)
            else:
                self.selected_ids.discard(a_id)
                
        if hasattr(self, "client_table"):
            for row_idx in range(self.client_table.rowCount()):
                item = self.client_table.item(row_idx, self._TC_NAME)
                if item:
                    self._apply_row_style(row_idx, checked, row_idx)
                    
        self._update_count()

    def _update_count(self):
        n = len(self.selected_ids)
        self.lbl_selected.setText(f"{n} selected" if n else "")
        total = len(self._checkbox_map)
        if hasattr(self, "header_cb"):
            self.header_cb.blockSignals(True)
            self.header_cb.setChecked(total > 0 and n == total)
            self.header_cb.blockSignals(False)
        if hasattr(self, "btn_delete_sel"):
            self.btn_delete_sel.setEnabled(n > 0)

    # ── Form Operations ───────────────────────────────────────────────────────

    # ── Client Dialog (Add / Edit popup) ─────────────────────────────────────

    def _open_add_client(self):
        self._client_dialog(None)

    def _open_edit_client(self, a):
        self._client_dialog(a)

    def _client_dialog(self, a=None):
        editing = a is not None
        dlg = QDialog(self)
        dlg.setWindowTitle("Edit Client" if editing else "Add New Client")
        dlg.setFixedWidth(400)
        t = _t()
        dlg.setStyleSheet(
            f"QDialog{{background:{t.bg_window};}}"
            f"QLabel{{background:transparent;border:none;color:{t.text_primary};font-size:12px;}}"
            f"QLineEdit{{border:1px solid {t.border};border-radius:6px;padding:6px 10px;"
            f"font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
            f"QLineEdit:focus{{border-color:{t.border_focus};background:{t.bg_input_focus};}}"
        )
        vl = QVBoxLayout(dlg)
        vl.setContentsMargins(28, 24, 28, 24)
        vl.setSpacing(0)

        title_lbl = QLabel("Edit Client" if editing else "Add New Client")
        title_lbl.setStyleSheet(f"font-size:16px;font-weight:700;color:{t.text_primary};")
        vl.addWidget(title_lbl)
        vl.addSpacing(18)

        fields = {}

        def _field_label(text):
            l = QLabel(text)
            l.setStyleSheet(f"font-size:11px;font-weight:600;color:{t.text_muted};margin-bottom:3px;")
            return l

        # ── Name ──────────────────────────────────────────────────────────────
        vl.addWidget(_field_label("Full Name"))
        fields["name"] = QLineEdit()
        fields["name"].setPlaceholderText("e.g. Deepak Bholusaria")
        fields["name"].setFixedHeight(34)
        fields["name"].setAttribute(Qt.WidgetAttribute.WA_MacShowFocusRect, False)
        vl.addWidget(fields["name"])
        vl.addSpacing(10)

        # ── PAN ───────────────────────────────────────────────────────────────
        vl.addWidget(_field_label("PAN Number"))
        fields["pan"] = QLineEdit()
        fields["pan"].setPlaceholderText("e.g. AEKPB0205L")
        fields["pan"].setFixedHeight(34)
        fields["pan"].setMaxLength(10)
        fields["pan"].setValidator(
            QRegularExpressionValidator(QRegularExpression("[A-Za-z0-9]{0,10}")))
        fields["pan"].setAttribute(Qt.WidgetAttribute.WA_MacShowFocusRect, False)
        fields["pan"].textChanged.connect(
            lambda txt, e=fields["pan"]: e.setText(txt.upper()) if txt != txt.upper() else None)
        vl.addWidget(fields["pan"])
        vl.addSpacing(10)

        # ── Date of Birth — text input + calendar picker + live hint ─────────
        vl.addWidget(_field_label("Date of Birth"))
        dob_row = QHBoxLayout()
        dob_row.setSpacing(6)
        dob_row.setContentsMargins(0, 0, 0, 0)

        dob_edit = QLineEdit()
        dob_edit.setPlaceholderText("DD-MM-YYYY  or type digits freely")
        dob_edit.setFixedHeight(34)
        dob_edit.setAttribute(Qt.WidgetAttribute.WA_MacShowFocusRect, False)
        fields["dob"] = dob_edit

        cal_btn = QPushButton("📅")
        cal_btn.setFixedSize(34, 34)
        cal_btn.setToolTip("Pick date from calendar")
        cal_btn.setStyleSheet(
            f"QPushButton{{background:{t.bg_input};border:1px solid {t.border};"
            f"border-radius:6px;font-size:15px;}}"
            f"QPushButton:hover{{border-color:{t.border_focus};}}")

        dob_row.addWidget(dob_edit, 1)
        dob_row.addWidget(cal_btn)
        vl.addLayout(dob_row)

        # Hint label: shows normalised date or error inline
        dob_hint = QLabel("")
        dob_hint.setStyleSheet("font-size:10px;background:transparent;border:none;")
        dob_hint.setFixedHeight(16)
        vl.addWidget(dob_hint)
        vl.addSpacing(6)

        def _smart_normalise(text: str) -> str:
            """Pre-process raw text before vault normalisation.
            Handles pure digit strings: 8 digits → DD-MM-YYYY, 6 digits → DD-MM-YY.
            """
            import re as _re
            from vault import _normalise_dob
            s = text.strip()
            # Pure digits only — insert dashes
            if _re.fullmatch(r"\d{8}", s):          # 12121975 → 12-12-1975
                s = f"{s[:2]}-{s[2:4]}-{s[4:]}"
            elif _re.fullmatch(r"\d{6}", s):         # 121275  → 12-12-75
                s = f"{s[:2]}-{s[2:4]}-{s[4:]}"
            return _normalise_dob(s)

        def _normalise_and_hint(text):
            """Show a live ✓/✗ hint as the user types."""
            from vault import _DOB_RE
            raw = text.strip()
            if not raw:
                dob_hint.setText("")
                dob_edit.setStyleSheet(
                    f"QLineEdit{{border:1px solid {t.border};border-radius:6px;"
                    f"padding:6px 10px;font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
                    f"QLineEdit:focus{{border-color:{t.border_focus};}}")
                return
            # Only validate when enough chars typed (avoid red flash on first keystrokes)
            if len(raw) < 6:
                dob_hint.setText("")
                dob_edit.setStyleSheet(
                    f"QLineEdit{{border:1px solid {t.border};border-radius:6px;"
                    f"padding:6px 10px;font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
                    f"QLineEdit:focus{{border-color:{t.border_focus};}}")
                return
            normed = _smart_normalise(raw)
            if _DOB_RE.match(normed):
                dob_hint.setText(f"✓  Will save as: {normed}")
                dob_hint.setStyleSheet("font-size:10px;color:#16A34A;background:transparent;border:none;")
                dob_edit.setStyleSheet(
                    f"QLineEdit{{border:1.5px solid #16A34A;border-radius:6px;"
                    f"padding:6px 10px;font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
                    f"QLineEdit:focus{{border-color:#16A34A;}}")
            else:
                dob_hint.setText("✗  Use DDMMYYYY or DD-MM-YYYY  e.g. 06071974")
                dob_hint.setStyleSheet("font-size:10px;color:#EF4444;background:transparent;border:none;")
                dob_edit.setStyleSheet(
                    f"QLineEdit{{border:1.5px solid #EF4444;border-radius:6px;"
                    f"padding:6px 10px;font-size:12px;background:{t.bg_input};color:{t.text_primary};}}"
                    f"QLineEdit:focus{{border-color:#EF4444;}}")

        def _on_dob_commit():
            """On Tab/Enter/focus-out, rewrite the field to canonical DD-MM-YYYY."""
            from vault import _DOB_RE
            raw = dob_edit.text().strip()
            if not raw:
                return
            normed = _smart_normalise(raw)
            if _DOB_RE.match(normed):
                # Block signal to avoid recursive hint update while rewriting
                dob_edit.blockSignals(True)
                dob_edit.setText(normed)
                dob_edit.blockSignals(False)
                _normalise_and_hint(normed)

        dob_edit.textChanged.connect(_normalise_and_hint)
        dob_edit.editingFinished.connect(_on_dob_commit)

        # Calendar popup
        def _show_calendar():
            from PyQt6.QtCore import QDate
            from PyQt6.QtGui import QPalette, QColor
            from vault import _normalise_dob, _DOB_RE

            popup = QDialog(dlg)
            popup.setWindowTitle("Pick Date of Birth")
            popup.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
            popup.setStyleSheet(
                f"QDialog{{background:{t.bg_panel};border:2px solid {t.border};border-radius:8px;}}"
                f"QToolButton{{color:{t.text_primary};background:transparent;border:none;"
                f"border-radius:4px;padding:3px 8px;font-size:12px;font-weight:600;}}"
                f"QToolButton:hover{{background:{t.accent};color:{t.accent_text};}}"
                f"QToolButton::menu-indicator{{image:none;}}"
                f"QSpinBox{{color:{t.text_primary};background:{t.bg_input};"
                f"border:1px solid {t.border};border-radius:4px;padding:2px 6px;}}"
                f"QSpinBox::up-button,QSpinBox::down-button{{background:transparent;border:none;}}"
                f"QMenu{{background:{t.bg_menu};color:{t.text_primary};border:1px solid {t.border};}}"
                f"QMenu::item:selected{{background:{t.accent};color:{t.accent_text};}}"
                f"QHeaderView::section{{background:{t.bg_header};color:{t.text_muted};"
                f"border:none;font-size:11px;font-weight:600;padding:4px 0;}}"
                f"#qt_calendar_navigationbar{{background:{t.bg_header};padding:6px 8px;}}"
                f"QAbstractItemView{{background:{t.bg_table};color:{t.text_primary};"
                f"selection-background-color:{t.accent};selection-color:{t.accent_text};"
                f"outline:none;border:none;}}"
            )

            pl = QVBoxLayout(popup)
            pl.setContentsMargins(0, 0, 0, 0)
            pl.setSpacing(0)

            cal = QCalendarWidget(popup)
            cal.setGridVisible(False)
            cal.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)

            # Set QPalette on the calendar so Qt's internal painter uses theme colours
            pal = QPalette()
            pal.setColor(QPalette.ColorRole.Window,      QColor(t.bg_table))
            pal.setColor(QPalette.ColorRole.Base,        QColor(t.bg_table))
            pal.setColor(QPalette.ColorRole.AlternateBase, QColor(t.bg_table_alt))
            pal.setColor(QPalette.ColorRole.Text,        QColor(t.text_primary))
            pal.setColor(QPalette.ColorRole.BrightText,  QColor(t.accent_text))
            pal.setColor(QPalette.ColorRole.Highlight,   QColor(t.accent))
            pal.setColor(QPalette.ColorRole.HighlightedText, QColor(t.accent_text))
            pal.setColor(QPalette.ColorRole.ButtonText,  QColor(t.text_primary))
            pal.setColor(QPalette.ColorRole.Button,      QColor(t.bg_input))
            pal.setColor(QPalette.ColorRole.WindowText,  QColor(t.text_primary))
            pal.setColor(QPalette.ColorRole.ToolTipBase, QColor(t.bg_menu))
            pal.setColor(QPalette.ColorRole.ToolTipText, QColor(t.text_primary))

            # Disabled role — other-month days
            pal.setColor(QPalette.ColorGroup.Disabled, QPalette.ColorRole.Text,
                         QColor(t.text_muted))
            cal.setPalette(pal)

            # Apply palette to all child widgets (nav bar, table view, header)
            for child in cal.findChildren(QWidget):
                child.setPalette(pal)
                child.setAutoFillBackground(True)

            # Pre-select current value if valid
            raw = dob_edit.text().strip()
            normed = _normalise_dob(raw) if raw else ""
            if _DOB_RE.match(normed):
                try:
                    d, m, y = normed.split("-")
                    cal.setSelectedDate(QDate(int(y), int(m), int(d)))
                except Exception:
                    pass

            cal.setMaximumDate(QDate.currentDate())
            pl.addWidget(cal)

            def _picked(date):
                dob_edit.setText(f"{date.day():02d}-{date.month():02d}-{date.year()}")
                popup.accept()

            cal.clicked.connect(_picked)
            pos = cal_btn.mapToGlobal(cal_btn.rect().bottomLeft())
            popup.move(pos)
            popup.exec()

        cal_btn.clicked.connect(_show_calendar)

        # ── Portal Password ───────────────────────────────────────────────────
        vl.addWidget(_field_label("Portal Password"))
        fields["pwd"] = QLineEdit()
        fields["pwd"].setPlaceholderText("Enter password")
        fields["pwd"].setFixedHeight(34)
        fields["pwd"].setEchoMode(QLineEdit.EchoMode.Password)
        fields["pwd"].setAttribute(Qt.WidgetAttribute.WA_MacShowFocusRect, False)
        vl.addWidget(fields["pwd"])
        vl.addSpacing(10)

        # Show password toggle
        show_cb = QCheckBox("Show password")
        show_cb.setStyleSheet(f"color:{t.text_muted};font-size:11px;background:transparent;")
        show_cb.toggled.connect(
            lambda v: fields["pwd"].setEchoMode(
                QLineEdit.EchoMode.Normal if v else QLineEdit.EchoMode.Password))
        vl.addWidget(show_cb)
        vl.addSpacing(18)

        # Pre-fill if editing
        if editing:
            fields["name"].setText(a.get("name", ""))
            fields["pan"].setText(a.get("pan", ""))
            fields["dob"].setText(a.get("dob", ""))
            fields["pwd"].setText(a.get("password", ""))

        # Buttons
        btn_row = QHBoxLayout()
        btn_cancel = _btn("Cancel", "secondary", height=34)
        btn_save   = _btn("Update Client" if editing else "Add Client", "primary", height=34)
        btn_row.addWidget(btn_cancel)
        btn_row.addStretch()
        btn_row.addWidget(btn_save)
        vl.addLayout(btn_row)

        btn_cancel.clicked.connect(dlg.reject)

        def _save():
            try:
                edit_id = a.get("id") if editing else None
                self.vault.add_update_assessee(
                    fields["name"].text(), fields["pan"].text(),
                    fields["dob"].text(), fields["pwd"].text(), edit_id)
                action = "updated" if editing else "added"
                self.log(f"[Vault] {fields['pan'].text()} {action}.")
                dlg.accept()
                self.refresh_grid()
            except ValueError as ve:
                QMessageBox.critical(dlg, "Validation Error", str(ve))
            except Exception as ex:
                QMessageBox.critical(dlg, "Error", str(ex))

        btn_save.clicked.connect(_save)
        dlg.exec()

    def save_assessee(self):
        self._open_add_client()

    def delete_assessee(self, assessee_id):
        if QMessageBox.question(self, "Confirm Delete",
            "Delete this assessee from the vault?") == QMessageBox.StandardButton.Yes:
            try:
                self.vault.delete_assessee(assessee_id)
                self.selected_ids.discard(assessee_id)
                self.log("[Vault] Assessee deleted.")
                self.refresh_grid()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete: {e}")

    def delete_selected(self):
        n = len(self.selected_ids)
        if not n:
            return
        if QMessageBox.question(self, "Confirm Bulk Delete",
                f"Delete {n} selected assessee{'s' if n > 1 else ''} from the vault?\n\nThis cannot be undone."
                ) != QMessageBox.StandardButton.Yes:
            return
        errors = []
        for a_id in list(self.selected_ids):
            try:
                self.vault.delete_assessee(a_id)
            except Exception as e:
                errors.append(str(e))
        self.selected_ids.clear()
        self.log(f"[Vault] {n} assessee{'s' if n > 1 else ''} deleted.")
        if errors:
            QMessageBox.warning(self, "Partial Delete", f"{len(errors)} deletion(s) failed:\n" + "\n".join(errors))
        self.refresh_grid()

    # ── Settings ──────────────────────────────────────────────────────────────

    def browse_output_dir(self):
        chosen = QFileDialog.getExistingDirectory(self, "Select Output Directory",
            self.dir_lbl.text())
        if chosen:
            self.dir_lbl.setText(chosen)
            self.vault.update_setting("download_root_dir", chosen)
            self.log(f"[Settings] Output folder: {chosen}")

    def _ay_json_path(self) -> str:
        """
        Writable path for assessment_years.json next to the exe / script.
        On first run when frozen, seeds the file from the bundled read-only copy.
        """
        writable = os.path.join(_app_dir(), "assessment_years.json")
        if not os.path.exists(writable):
            bundled = os.path.join(_bundled_dir(), "assessment_years.json")
            if os.path.exists(bundled):
                import shutil
                shutil.copy2(bundled, writable)
        return writable

    def _load_ay_list(self):
        path = self._ay_json_path()
        try:
            with open(path, "r", encoding="utf-8") as f:
                entries = json.load(f)
            # Sort: disabled (future TY) first, then descending by year
            def _sort_key(e):
                y = e.get("year", {})
                label_year = (y.get("AY") or y.get("TY") or y.get("FY") or "0000-00")
                try:
                    return (0 if not e.get("enabled", True) else 1, -int(label_year[:4]))
                except ValueError:
                    return (1, 0)
            return sorted(entries, key=_sort_key)
        except Exception:
            return [
                {"label": "AY 2025-26 (FY 2024-25)", "enabled": True, "year": {"AY": "2025-26", "FY": "2024-25"}},
                {"label": "AY 2024-25 (FY 2023-24)", "enabled": True, "year": {"AY": "2024-25", "FY": "2023-24"}},
            ]

    def _resolve_ay_fy(self, label: str):
        """Returns (ay_or_ty_value, fy_value, year_type) where year_type is 'AY' or 'TY'."""
        for e in self._ay_entries:
            if e["label"] == label:
                y = e["year"]
                if y.get("AY"):
                    return y["AY"], y.get("FY"), "AY"
                if y.get("TY"):
                    return y["TY"], y.get("FY"), "TY"
        return None, None, "AY"

    def open_manage_years(self):
        ManageYearsDialog(self, self._ay_json_path(), on_save=self.refresh_ay_combo).exec()

    def refresh_ay_combo(self):
        self._ay_entries = self._load_ay_list()
        ay_labels = [e["label"] for e in self._ay_entries if e.get("enabled", True)]
        current = self.ay_combo.currentText()
        self.ay_combo.blockSignals(True)
        self.ay_combo.clear()
        self.ay_combo.addItem("Select AY/TY")
        self.ay_combo.addItems(ay_labels)
        self.ay_combo.setCurrentText(current if current in ay_labels else "Select AY/TY")
        self.ay_combo.blockSignals(False)
        self.log("[Settings] Assessment Year list refreshed.")

    def save_ay_setting(self, val):
        if val and val != "Select AY/TY":
            self.vault.update_setting("assessment_year", val)
            self.log(f"[Settings] Assessment Year → {val}")

    # ── Logging ───────────────────────────────────────────────────────────────

    def log(self, message):
        text = f"[{get_timestamp()}] {message}"
        self._log_signal.emit(text)
        try:
            with open(os.path.join(_app_dir(), "app.log"), "a", encoding="utf-8") as f:
                f.write(text + "\n")
        except Exception:
            pass

    def _append_log(self, text):
        self.log_box.append(text)
        self.log_box.moveCursor(QTextCursor.MoveOperation.End)

    def copy_logs_to_clipboard(self):
        QApplication.clipboard().setText(self.log_box.toPlainText())
        self.log("[System] Logs copied to clipboard.")

    # ── Bulk Import ───────────────────────────────────────────────────────────

    def bulk_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import File", "",
            "Excel / CSV files (*.xlsx *.csv)")
        if not path:
            return
        self.log(f"[Vault] Importing {os.path.basename(path)}...")
        added, updated, errors = self.vault.import_bulk(path)
        total = added + updated
        parts = []
        if added:
            parts.append(f"{added} new record{'s' if added != 1 else ''} added")
        if updated:
            parts.append(f"{updated} existing record{'s' if updated != 1 else ''} updated")
        summary = ", ".join(parts) if parts else "No records imported"
        self.log(f"[Vault] Import complete — {summary}.")
        if errors:
            for err in errors:
                self.log(f"  - {err}")
            QMessageBox.warning(self, "Import Complete",
                f"{summary}.\n\n{len(errors)} row(s) had errors — see logs for details.")
        else:
            QMessageBox.information(self, "Import Complete", f"{summary}.")
        self.refresh_grid()

    def generate_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Template",
            "Assessee_Import_Template", "Excel Workbook (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        try:
            self.vault.generate_template(path)
            self.log(f"[Vault] Template generated: {path}")
            QMessageBox.information(self, "Success", f"Template generated at:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed: {e}")

    def export_data(self):
        assessees = self.vault.get_all_assessees()
        if not assessees:
            QMessageBox.information(self, "No Data", "No assessees saved yet.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export Saved Data",
            "Assessee_Export", "Excel Workbook (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        if not (path.endswith('.xlsx') or path.endswith('.csv')):
            path += ".xlsx"
        try:
            self.vault.export_data(path)
            self.log(f"[Vault] Data exported: {path}")
            QMessageBox.information(self, "Export Complete",
                f"{len(assessees)} record(s) exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed: {e}")

    # ── Browser health check ──────────────────────────────────────────────────

    def _check_browser(self):
        """Run silently on startup; installs Chromium if missing."""
        def _run():
            import asyncio
            from automation.browser import _playwright_browsers_dir, _install_chromium
            import os
            # Check if chromium executable already exists in our browsers dir
            browsers_dir = _playwright_browsers_dir()
            chromium_exists = any(
                f.startswith("chromium") for f in os.listdir(browsers_dir)
            ) if os.path.exists(browsers_dir) else False
            if not chromium_exists:
                self.log("[Browser] Chromium not found — installing in background...")
                loop = asyncio.new_event_loop()
                try:
                    loop.run_until_complete(_install_chromium(self.log))
                except Exception as e:
                    self.log(f"[Browser] Auto-install failed: {e}")
                    self.log("[Browser] Run 'playwright install chromium' manually if downloads fail.")
                finally:
                    loop.close()
            else:
                self.log("[Browser] Chromium ready.")
        threading.Thread(target=_run, daemon=True).start()

    # ── Automation ────────────────────────────────────────────────────────────

    def _lock_ui(self, lock: bool):
        widgets = [self.ay_combo, self.btn_delete_sel, self.btn_run, self.chk_headless]
        if hasattr(self, "header_cb"):
            widgets.append(self.header_cb)
        for w in widgets:
            w.setEnabled(not lock)
        # Disable Client Master / Settings menus during batch
        menubar = self.menuBar()
        for action in menubar.actions():
            if action.text() in ("Client Master", "Settings"):
                if action.menu():
                    action.menu().setEnabled(not lock)

    def start_automation(self, mode: str):
        """
        mode: "26as"        — download 26AS only
              "request_ais" — fire AIS generation request (Phase 1)
              "ais_tis"     — download AIS (from Activity History) + TIS (Phase 2)
        """
        if self.is_running:
            return
        if not self.selected_ids:
            QMessageBox.warning(self, "Selection Required",
                "Please select at least one client.")
            return
        ay_label = self.ay_combo.currentText()
        if not ay_label or ay_label == "Select AY/TY":
            QMessageBox.warning(self, "Selection Required",
                "Please select an Assessment / Tax Year.")
            return
        ay, fy, year_type = self._resolve_ay_fy(ay_label)
        if not ay:
            QMessageBox.warning(self, "Invalid", f"Cannot resolve year: {ay_label}")
            return

        self.is_running = True
        self._batch_aborted = False
        self._lock_ui(True)
        self.log_box.clear()

        targets = [a for a in self.assessee_list if a.get("id") in self.selected_ids]
        output_dir = self.dir_lbl.text()
        self._last_batch_params = (ay, fy, output_dir, mode, ay_label)

        self.btn_run.setText("⏳ Running...")

        mode_log = {"26as": "26AS download",
                    "request_ais": "AIS generation requests",
                    "ais_tis": "AIS/TIS download"}
        self.log(f"[System] Starting {mode_log[mode]}...")

        # Year tag shown in progress dialog:
        #   26AS  → "AY 2025-26" or "TY 2025-26" (whatever the user configured)
        #   AIS/TIS → "FY 2024-25"
        if mode == "26as":
            year_tag = f"{year_type} {ay}"
        else:
            year_tag = f"FY {fy}" if fy else ay

        # Show progress dialog (on main thread via signal)
        self._show_progress_signal.emit(targets, mode, year_tag, output_dir)

        threading.Thread(
            target=self._run_wrapper,
            args=(targets, ay, fy, output_dir, mode, ay_label),
            daemon=True).start()

    def _show_progress_dialog(self, targets: list, mode: str, ay: str, output_dir: str = ""):
        """Called on main thread to create and show the progress dialog."""
        self._progress_dialog = BatchProgressDialog(
            targets, mode, ay=ay, stop_callback=self.stop_automation,
            resume_callback=self.resume_batch,
            output_dir=output_dir, parent=self)
        # Window-modal: blocks the parent window (so it can't be clicked behind
        # the dialog) but still allows the worker thread's Qt-signal updates.
        # We use show() rather than exec() so the event loop keeps processing
        # the live status signals.
        self._progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        self._progress_dialog.show()
        self._progress_dialog.raise_()
        self._progress_dialog.activateWindow()

    def stop_automation(self):
        if not self.is_running:
            return
        _sm = _t()
        _mb = QMessageBox(self)
        _mb.setWindowTitle("Stop")
        _mb.setText("Abort the active batch?")
        _mb.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        _mb.setDefaultButton(QMessageBox.StandardButton.No)
        _mb.setStyleSheet(
            f"QMessageBox{{background:{_sm.bg_window};color:{_sm.text_primary};}}"
            f"QLabel{{color:{_sm.text_primary};background:transparent;}}"
            f"QPushButton{{background:{_sm.accent};color:{_sm.accent_text};border:none;"
            f"border-radius:5px;padding:6px 18px;font-size:13px;}}"
            f"QPushButton:hover{{background:{_sm.accent_hover};}}"
            f"QPushButton[text='No']{{background:{_sm.border};color:{_sm.text_primary};}}"
        )
        if _mb.exec() == QMessageBox.StandardButton.Yes:
            self.log("[System] Abort requested...")
            self.is_running = False
            self._batch_aborted = True
            # Cancel the asyncio task immediately — raises CancelledError into
            # whatever await is currently blocking (goto, wait_for_selector, sleep…)
            if self._batch_task and self._batch_loop:
                self._batch_loop.call_soon_threadsafe(self._batch_task.cancel)

    def resume_batch(self, remaining_targets: list):
        """Called from the dialog Resume button — restart batch with unfinished clients."""
        if not self._last_batch_params or not remaining_targets:
            return
        ay, fy, root_dir, mode, ay_label = self._last_batch_params
        self.is_running = True
        self._batch_aborted = False
        self._lock_ui(True)
        self.btn_run.setText("⏳ Running...")
        self.log(f"[System] Resuming — {len(remaining_targets)} client(s) remaining...")
        if self._progress_dialog:
            self._progress_dialog.batch_resumed()
        threading.Thread(
            target=self._run_wrapper,
            args=(remaining_targets, ay, fy, root_dir, mode, ay_label),
            daemon=True).start()

    def _run_wrapper(self, targets, ay, fy, root_dir, mode, ay_label=""):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        self._batch_loop = loop

        # Suppress "Future exception was never retrieved" noise from Playwright
        # futures that are orphaned when the browser closes mid-await after Stop.
        def _exc_handler(loop, ctx):
            exc = ctx.get("exception")
            if exc is not None:
                name = type(exc).__name__
                msg  = str(exc).lower()
                # Suppress orphaned futures from browser close / abort
                if name in ("TargetClosedError", "ConnectionClosedError",
                            "CancelledError"):
                    return
                # Playwright base Error with net:: abort codes that appear when
                # the browser is closed mid-navigation after Stop
                if name == "Error" and any(code in msg for code in (
                        "err_aborted", "err_empty_response",
                        "err_connection_reset", "err_connection_refused",
                        "target page, context or browser has been closed")):
                    return
            loop.default_exception_handler(ctx)

        loop.set_exception_handler(_exc_handler)

        try:
            self._batch_task = loop.create_task(
                self._execute_batch(targets, ay, fy, root_dir, mode, ay_label))
            loop.run_until_complete(self._batch_task)
        except asyncio.CancelledError:
            self.log("[System] Batch cancelled.")
        except Exception as e:
            self.log(f"[System Error] Batch crashed: {e}")
        finally:
            self._batch_task = None
            self._batch_loop = None
            loop.close()
            self.is_running = False
            self._last_mode = mode
            if self._progress_dialog:
                self._progress_dialog.batch_finished(aborted=self._batch_aborted)
            self._batch_done_signal.emit()

    def _on_batch_done(self):
        import datetime
        self.btn_run.setText("▶  Run")
        self._lock_ui(False)
        self.log("[System] Engine Idle.")
        # Refresh grid so Last Download Status / Last Saved Location columns update
        QTimer.singleShot(200, self.refresh_grid)

        mode = self._last_mode

        if self._batch_aborted:
            self._ais_results = {}
            return

        if mode == "request_ais":
            results = self._ais_results
            n_instant = sum(1 for v in results.values() if v == "instant")
            n_queued  = sum(1 for v in results.values() if v == "queued")
            n_failed  = sum(1 for v in results.values() if v == "failed")

            self._ais_requested_time = datetime.datetime.now()
            t = self._ais_requested_time.strftime("%I:%M %p")

            # Build status line text — only mention what actually happened
            parts = []
            if n_instant:
                parts.append(f"{n_instant} downloaded instantly")
            if n_queued:
                parts.append(f"{n_queued} queued on ITD servers")
            if n_failed:
                parts.append(f"{n_failed} failed")
            status_summary = " · ".join(parts) if parts else "no results"

            if n_queued:
                self.ais_status_lbl.setText(
                    f"⏳  AIS at {t}: {status_summary} — "
                    f"wait ~5 min then click ▶ Run → Download Previously Requested AIS for the {n_queued} queued client(s).")
                self.ais_status_bar.setVisible(True)
            else:
                # All instant or failed — no need to show the waiting reminder
                self.ais_status_bar.setVisible(False)

            # Build dialog text based on actual breakdown
            lines = ["<b>AIS Request Results:</b><br>"]
            if n_instant:
                lines.append(
                    f"✅ <b>{n_instant} client(s)</b> — AIS file was small, "
                    f"downloaded instantly. No further action needed for these.")
            if n_queued:
                lines.append(
                    f"⏳ <b>{n_queued} client(s)</b> — AIS file is large, "
                    f"generation request queued on ITD servers.<br>"
                    f"&nbsp;&nbsp;Wait <b>~5 minutes</b>, then select these clients "
                    f"and click <b>▶ Run → ⬇ Download Previously Requested AIS</b>.")
            if n_failed:
                # Show the distinct error reasons so the user knows WHY.
                reasons = sorted(set(self._last_errors.values()))
                reason_html = ""
                if reasons:
                    reason_html = "<br>" + "<br>".join(
                        f"&nbsp;&nbsp;• {r}" for r in reasons[:5])
                lines.append(
                    f"❌ <b>{n_failed} client(s)</b> — Request failed.{reason_html}")

            msg = QMessageBox(self)
            msg.setWindowTitle("AIS Request Complete")
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setText("<br><br>".join(lines))
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg.exec()

        elif mode == "ais_tis":
            # Hide the status line once download is done
            self.ais_status_bar.setVisible(False)
            self._ais_requested_time = None

        elif mode == "26as" and not self._batch_aborted:
            def _set_status(pan, text):
                if self._progress_dialog:
                    self._progress_dialog.set_status(pan, text)
            self._auto_convert_26as(_set_status)

    def _convert_26as_manual(self):
        from as26_converter import convert_26as_txt
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Select 26AS TXT file(s)",
            self.vault.get_setting("download_root_dir", ""),
            "Text Files (*.txt);;All Files (*)")
        if not paths:
            return
        ok, fail = 0, 0
        locked_warnings = []  # collect "saved as alternate" warnings for the dialog

        def _log_capturing(msg):
            self.log(msg)
            if msg.startswith("[Warning]") and "was open in Excel" in msg:
                locked_warnings.append(msg[len("[Warning] "):])

        for p in paths:
            try:
                xlsx, html = convert_26as_txt(p, log_callback=_log_capturing)
                self.log(f"[Victory] Converted: {os.path.basename(xlsx)} + {os.path.basename(html)}")
                ok += 1
            except Exception as e:
                self.log(f"[Error] Convert failed for {os.path.basename(p)}: {e}")
                fail += 1
        msg = f"Converted {ok} file(s) — Excel + HTML saved alongside each TXT."
        if locked_warnings:
            msg += "\n\nNote — the following file(s) were open in Excel and saved under an alternate name:\n"
            msg += "\n".join(f"  • {w}" for w in locked_warnings)
        if fail:
            msg += f"\n\n{fail} file(s) failed — check the log for details."
        QMessageBox.information(self, "Convert Complete", msg)

    def _auto_convert_26as(self, set_status=None):
        from as26_converter import convert_26as_txt
        items = getattr(self, "_batch_26as_txts", [])
        if not items:
            return
        converted = 0
        for pan, txt in items:
            if not os.path.exists(txt):
                continue
            if set_status:
                set_status(pan, "⏳ Converting to Excel...")
            try:
                convert_26as_txt(txt, log_callback=self.log)
                converted += 1
                if set_status:
                    set_status(pan, "✅ 26AS + Excel + HTML")
            except Exception as e:
                self.log(f"[Warning] Auto-convert failed for {os.path.basename(txt)}: {e}")
                if set_status:
                    set_status(pan, "⚠ Excel convert failed")
        if converted:
            self.log(f"[Victory] Auto-converted {converted} 26AS TXT file(s) to Excel + HTML.")

    async def _execute_batch(self, targets, ay, fy, root_dir, mode, ay_label=""):
        self.log(f"[System] Batch: {len(targets)} client(s) | AY: {ay} | Mode: {mode}")
        self._ais_results = {}
        self._last_errors = {}
        self._batch_26as_txts = []

        _client_out = {}   # pan → out path, populated below

        def set_status(pan, text):
            """Update progress dialog and persist terminal status to vault."""
            if self._progress_dialog:
                self._progress_dialog.set_status(pan, text)
            terminal = ("✅", "❌", "🕐", "⏹", "⬜", "⚠")
            if ay_label and any(text.startswith(p) for p in terminal):
                try:
                    self.vault.record_download(
                        pan, ay_label, text, _client_out.get(pan, ""))
                except Exception:
                    pass

        try:
            interactive = not self.chk_headless.isChecked()
            context = await browser_manager.get_context(
                log_callback=self.log, interactive=interactive)
        except Exception as e:
            self.log(f"[System Error] Browser init failed: {e}"); return

        try:
            for i, target in enumerate(targets):
                if not self.is_running:
                    self.log("[System] Aborted."); break

                pan  = target.get("pan", "")
                name = target.get("name", "")
                dob  = target.get("dob", "")
                self.log("──────────────────────────────────────────────────")
                self.log(f"[{i+1}/{len(targets)}] {name}")

                name_safe = "".join(c if c.isalnum() or c in " _-" else "" for c in name)
                out = os.path.join(root_dir, f"{pan}-{name_safe}", f"AY_{ay.replace('-','_')}")

                _client_out[pan] = out
                # Tell the progress dialog the client's save folder immediately
                if self._progress_dialog:
                    self._progress_dialog.set_client_path(pan, out)

                page = None
                try:
                    # ── Login ────────────────────────────────────────────────────
                    set_status(pan, "⏳ Logging in to ITD...")
                    page = await login_itd(pan, target.get("password"), self.log, context,
                                           is_running=lambda: self.is_running)
                    set_status(pan, "⏳ Logged in to ITD")

                    # ── 26AS ─────────────────────────────────────────────────────
                    if mode == "26as" and self.is_running:
                        set_status(pan, "⏳ Downloading 26AS...")
                        ok, err_msg, txt_path = await download_26as(page, ay, out, self.log, pan=pan, dob=dob)
                        if ok:
                            if err_msg:
                                # PDF saved but TXT failed (bad password / extraction error)
                                set_status(pan, f"⚠ Partially Completed — {err_msg}")
                            else:
                                set_status(pan, "✅ 26AS Downloaded")
                            if txt_path:
                                self._batch_26as_txts.append((pan, txt_path))
                        else:
                            set_status(pan, f"❌ 26AS Failed — {err_msg}")

                    # ── Request AIS ───────────────────────────────────────────────
                    elif mode == "request_ais" and self.is_running:
                        await self._ensure_dashboard(page)
                        set_status(pan, "⏳ Opening AIS portal...")
                        result = await run_request_ais(
                            page, fy, out, self.log, pan=pan, dob=dob,
                            status_callback=lambda t, _p=pan: set_status(_p, t))
                        ais_status = result.get("status")

                        # Track AIS queue state for the "Download AIS" button
                        if ais_status == "downloaded":
                            self._ais_results[pan] = "instant"
                        elif ais_status == "requested":
                            self._ais_results[pan] = "queued"
                            ref = result.get("ref_id", "")
                            if ref:
                                self.log(f"[AIS] Generation queued — Ref ID: {ref}")
                        elif ais_status == "skipped":
                            self._ais_results[pan] = "skipped"
                        else:
                            self._ais_results[pan] = "failed"
                        # combined_status_label already set via status_callback at end of run_request_ais

                    # ── Download AIS/TIS ──────────────────────────────────────────
                    elif mode == "ais_tis" and self.is_running:
                        # "Download Previously Requested AIS" — fetch ONLY the AIS PDF
                        # from Activity History. TIS is not re-downloaded here (it was
                        # already grabbed during the Request step).
                        await self._ensure_dashboard(page)
                        set_status(pan, "⏳ Downloading AIS from Activity History...")

                        dl_result = await run_download_ais_tis(
                            page, fy, out, self.log, pan=pan, dob=dob,
                            dl_ais=True, dl_tis=False,
                            should_continue=lambda: self.is_running,
                            status_callback=lambda t, _p=pan: set_status(_p, t))
                        # combined_status_label already set via status_callback at end of run_download_ais_tis

                    if self.is_running:
                        await logout_itd(page, self.log)
                        page = None

                    pan_masked = pan[:3] + "XXXXXXX" if pan and len(pan) >= 3 else "UNKNOWN"
                    self.log(f"[Victory] {pan_masked} done.")

                except Exception as e:
                    pan_masked = pan[:3] + "XXXXXXX" if pan and len(pan) >= 3 else "UNKNOWN"
                    self.log(f"[Error] {pan_masked}: {e}")
                    # Record the failure so the summary dialog/counts reflect it.
                    self._ais_results[pan] = "failed"
                    self._last_errors[pan] = str(e)
                    set_status(pan, f"❌ Failed — {_friendly_error(str(e))}")
                    if page:
                        try: await logout_itd(page, self.log)
                        except Exception: pass
                await asyncio.sleep(3)

        finally:
            await browser_manager.close()
        self.log("[System] Batch finished.")

    async def _ensure_dashboard(self, page):
        """Navigate back to ITD dashboard if not already there."""
        try:
            if "dashboard" not in page.url.lower():
                await page.goto(
                    "https://eportal.incometax.gov.in/iec/fo/dashboard",
                    wait_until="domcontentloaded", timeout=30000)
                await asyncio.sleep(2)
        except Exception:
            pass


def _fatal(msg: str):
    """Show a visible error dialog even before QApplication exists, then exit."""
    try:
        # Try Qt dialog first (works if Qt DLLs loaded successfully)
        _a = QApplication.instance() or QApplication(sys.argv)
        from PyQt6.QtWidgets import QMessageBox
        box = QMessageBox()
        box.setWindowTitle("AayDocCapio — Startup Error")
        box.setIcon(QMessageBox.Icon.Critical)
        box.setText("AayDocCapio could not start.")
        box.setDetailedText(msg)
        box.exec()
    except Exception:
        # Qt itself failed — fall back to Windows MessageBox via ctypes
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(
                0,
                f"AayDocCapio could not start.\n\n{msg}\n\n"
                "If this persists, install the Microsoft Visual C++ Redistributable:\n"
                "https://aka.ms/vs/17/release/vc_redist.x64.exe",
                "AayDocCapio — Startup Error",
                0x10  # MB_ICONERROR
            )
        except Exception:
            pass
    sys.exit(1)


if __name__ == "__main__":
    # Called by Inno Setup [Run] step to pre-install Chromium silently
    if "--install-browsers" in sys.argv:
        from automation.browser import _install_chromium
        loop = asyncio.new_event_loop()
        try:
            loop.run_until_complete(_install_chromium())
        except Exception as e:
            print(f"Browser install failed: {e}", file=sys.stderr)
            sys.exit(1)
        finally:
            loop.close()
        sys.exit(0)

    # Write a startup trace log using only builtins — before any Qt call
    _diag_path = os.path.join(_app_dir(), "startup_diag.log")
    def _diag(msg):
        try:
            os.makedirs(os.path.dirname(_diag_path), exist_ok=True)
            with open(_diag_path, "a", encoding="utf-8") as _f:
                _f.write(msg + "\n")
        except Exception:
            pass

    _diag(f"\n=== Startup {datetime.datetime.now()} ===")
    _diag(f"bundled_dir : {_bundled_dir()}")
    _diag(f"app_dir     : {_app_dir()}")
    _diag(f"sys.argv    : {sys.argv}")
    _diag(f"platform    : {sys.platform}")

    try:
        _diag("Step 1: QApplication()")
        app = QApplication(sys.argv)
        _diag("Step 2: setApplicationName")
        app.setApplicationName("AayDocCapio")
        app.setDesktopFileName("aay-doc-capio")
        app.setStyle("Fusion")
        # Force light palette so macOS dark mode doesn't corrupt unstyled
        # native widgets (dialogs, menus, headers). Will be overridden once
        # the theme system initialises inside AayDocCapioApp.
        try:
            app.styleHints().setColorScheme(Qt.ColorScheme.Light)
        except AttributeError:
            pass  # Qt < 6.8
        _diag("Step 3: font loading")
        from PyQt6.QtGui import QFontDatabase
        _fonts_dir = os.path.join(_bundled_dir(), "resources", "fonts")
        if os.path.isdir(_fonts_dir):
            for _ttf in os.listdir(_fonts_dir):
                if _ttf.endswith(".ttf"):
                    QFontDatabase.addApplicationFont(os.path.join(_fonts_dir, _ttf))
        _diag("Step 4: setStyleSheet")
        app.setStyleSheet(build_stylesheet(get_theme("light")))
        _diag("Step 5: AayDocCapioApp()")
        window = AayDocCapioApp()
        _diag("Step 6: setWindowIcon")
        _app_icon_path = os.path.join(_bundled_dir(), "resources", "app_icon.png")
        if os.path.exists(_app_icon_path):
            _icon = QIcon(_app_icon_path)
            app.setWindowIcon(_icon)
            window.setWindowIcon(_icon)
        _diag("Step 7: window.show()")
        window.show()
        _diag("Step 8: app.exec() — entering event loop")
        sys.exit(app.exec())
    except Exception as _startup_err:
        import traceback
        _tb = traceback.format_exc()
        _diag(f"CRASH:\n{_tb}")
        _fatal(_tb)
