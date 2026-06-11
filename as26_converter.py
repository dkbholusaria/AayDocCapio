"""
26AS TXT → Excel + HTML converter.

Public API:
    convert_26as_txt(txt_path, log_callback=None) -> (xlsx_path, html_path)
"""

import os
import re
from datetime import datetime

# ── Part metadata ──────────────────────────────────────────────────────────────

PART_META = {
    "I":    {"title": "TDS — Salary / Professional / Interest / VDA",             "credit": True},
    "II":   {"title": "TDS on Interest (15G/15H — No TDS Deducted)",              "credit": True},
    "III":  {"title": "TDS on Winnings / Benefits / VDA (Tax Paid Before Release)","credit": True},
    "IV":   {"title": "TDS u/s 194IA/IB/M/S (Seller of Property / VDA)",          "credit": True},
    "V":    {"title": "26QE — Seller of Virtual Digital Asset",                    "credit": True},
    "VI":   {"title": "TCS — Tax Collected at Source",                             "credit": True},
    "VII":  {"title": "Refunds Paid",                                              "credit": False},
    "VIII": {"title": "TDS u/s 194IA/IB/M/S (Buyer of Property / VDA)",           "credit": False},
    "IX":   {"title": "26QE — Buyer of Virtual Digital Asset",                     "credit": False},
}

NON_FINAL_STATUSES = {"U", "M", "O", "P", "Z"}

STATUS_FULL = {
    "F": "Final",
    "U": "Unmatched",
    "M": "Matched",
    "O": "Overbooked",
    "P": "Provisional",
    "Z": "Mismatch",
}

# ── Parser ─────────────────────────────────────────────────────────────────────

def _parse(txt_path: str) -> dict:
    """Return a dict with keys: header, parts.
    header = {field: value, ...}
    parts  = { "I": {"rows": [...], "empty": bool}, ... }
    Each row is a dict of field->value.
    """
    with open(txt_path, encoding="utf-8", errors="replace") as f:
        raw = f.read()

    lines = [ln.rstrip("\n\r") for ln in raw.splitlines()]

    # ── File header (lines 0-2) ────────────────────────────────────────────
    header_keys   = [c.strip() for c in lines[2].split("^") if c.strip()]
    header_values = [c.strip() for c in lines[3].split("^") if c.strip()] if len(lines) > 3 else []
    header = dict(zip(header_keys, header_values))

    # ── Split into Part blocks ─────────────────────────────────────────────
    part_re = re.compile(r"^\^PART-([IVX]+)\s*[-–]", re.IGNORECASE)
    no_txn_re = re.compile(r"No Transactions Present", re.IGNORECASE)

    parts = {}
    current_part = None
    current_lines = []

    for ln in lines[4:]:
        m = part_re.match(ln)
        if m:
            if current_part:
                parts[current_part] = current_lines
            current_part = m.group(1).upper()
            current_lines = []
        elif current_part:
            current_lines.append(ln)

    if current_part and current_part != "X":
        parts[current_part] = current_lines

    # ── Parse each Part ────────────────────────────────────────────────────
    parsed_parts = {}
    for roman, plines in parts.items():
        if roman == "X":
            continue
        # Check empty
        empty = any(no_txn_re.search(ln) for ln in plines)
        if empty:
            parsed_parts[roman] = {"empty": True, "rows": [], "col_headers": []}
            continue

        # Find column header lines (first non-blank line = summary headers,
        # second non-blank = detail headers)
        non_blank = [ln for ln in plines if ln.strip()]
        if not non_blank:
            parsed_parts[roman] = {"empty": True, "rows": [], "col_headers": []}
            continue

        summary_headers = [c.strip() for c in non_blank[0].split("^")]
        detail_headers  = []
        # detail header = first line that starts with ^ (blank first field)
        for ln in non_blank[1:]:
            if ln.startswith("^"):
                detail_headers = [c.strip() for c in ln.split("^")]
                break

        rows = []
        i = 0
        data_lines = non_blank[1:]  # skip the summary-col header line

        # skip detail header line itself
        skip_next_detail_hdr = True

        while i < len(data_lines):
            ln = data_lines[i]
            if not ln.strip():
                i += 1
                continue

            fields = [c.strip() for c in ln.split("^")]

            # Deductor/summary row: first field is a digit (Sr. No.)
            if fields[0].isdigit():
                # Skip detail-header lines that appear after each deductor block
                # (they start with ^ and contain "Sr. No." etc.)
                deductor_row = dict(zip(summary_headers, fields))
                deductor_row["_type"] = "deductor"
                deductor_row["_details"] = []
                i += 1
                # Collect detail rows (start with ^)
                while i < len(data_lines):
                    dl = data_lines[i]
                    if not dl.strip():
                        i += 1
                        break
                    dfields = [c.strip() for c in dl.split("^")]
                    # skip if it looks like a header (contains "Sr. No." or "Section")
                    if dfields[0] == "" and dfields[1:2] == ["Sr. No."]:
                        i += 1
                        continue
                    if dfields[0] == "" and dfields[1].isdigit():
                        detail_row = dict(zip(detail_headers, dfields))
                        detail_row["_type"] = "detail"
                        deductor_row["_details"].append(detail_row)
                        i += 1
                    else:
                        break
                rows.append(deductor_row)
            else:
                i += 1

        parsed_parts[roman] = {
            "empty": False,
            "rows": rows,
            "col_headers": summary_headers,
            "detail_headers": detail_headers,
        }

    return {"header": header, "parts": parsed_parts}


def _fmt_num(val: str) -> float | None:
    """Convert string number to float; None if blank/dash."""
    v = val.strip().replace(",", "")
    if not v or v == "-":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _ind(n: float | None) -> str:
    """Format float as Indian number string for HTML."""
    if n is None:
        return ""
    if n < 0:
        return f"({abs(n):,.2f})"
    return f"{n:,.2f}"


# ── HTML generator ────────────────────────────────────────────────────────────

_CSS = """
* { box-sizing:border-box; margin:0; padding:0; }
html, body { height:100%; }
body { font-family:"Segoe UI",Calibri,system-ui,-apple-system,Arial,sans-serif; font-size:12px; background:#e8e8e8;
  display:flex; flex-direction:column; }
.tab-bar { display:flex; align-items:flex-end; flex-wrap:wrap; background:#d6d6d6;
  border-bottom:2px solid #1a5c32; padding:4px 8px 0; gap:2px; flex-shrink:0; }
.tab { padding:5px 12px; background:#c0c0c0; border:1px solid #a0a0a0;
  border-bottom:none; border-radius:3px 3px 0 0; cursor:pointer; font-size:11px;
  color:#333; white-space:nowrap; transition:background 0.12s; }
.tab:hover { background:#d8d8d8; }
.tab.active { background:#fff; color:#1a5c32; font-weight:700; border-color:#1a5c32; }
.sheet { display:none; background:#fff; flex-direction:column; flex:1; min-height:0; }
.sheet.active { display:flex; }
.brand { background:linear-gradient(135deg,#0A1628 0%,#1e3a5f 100%);
  color:#fff; padding:10px 20px; display:flex; align-items:center; gap:14px; flex-shrink:0; }
.brand-title { font-size:16px; font-weight:700; }
.brand-sub { font-size:11px; color:#94A3B8; margin-top:2px; }
.brand-badge { margin-left:auto; background:#217346; color:#fff; padding:3px 10px;
  border-radius:10px; font-size:10px; font-weight:600; }
.freeze-bar { background:#fffbe6; border-bottom:2px solid #e6b800; padding:3px 10px;
  font-size:10px; color:#7a5c00; flex-shrink:0; }
.note-bar { background:#e8f0fe; border-bottom:1px solid #bad0f8; padding:4px 10px;
  font-size:10px; color:#1a3a8f; flex-shrink:0; }
.tbl-wrap { overflow-x:auto; overflow-y:auto; flex:1; min-height:0; }
.footer { background:#0A1628; color:#94A3B8; font-size:10px; padding:4px 14px;
  display:flex; align-items:center; gap:16px; flex-shrink:0; }
.footer a { color:#7da9d8; text-decoration:none; }
.footer a:hover { color:#fff; text-decoration:underline; }
.footer-copy { margin-right:auto; }
table { border-collapse:collapse; width:100%; }
td,th { border:1px solid #d0d0d0; padding:3px 7px; white-space:nowrap; font-size:11px; }
thead th { background:#1a5c32; color:#fff; font-weight:700; text-align:center;
  position:sticky; top:0; z-index:2; }
tr.alt td { background:#f5fbf5; }
tr:hover td { background:#e8f5e9; }
tr.subtotal td { background:#d0e8c8; font-weight:700; border-top:2px solid #1a5c32; color:#1a3a22; }
tr.subtotal:hover td { background:#c4e0bc; }
tr.grandtotal td { background:#0A1628; color:#fff; font-weight:700; border-top:2px solid #fff; }
tr.grandtotal td.num { color:#7fff8a; }
tr.empty-note td { background:#fff8f0; color:#a05000; font-style:italic; text-align:center; padding:10px; }
td.ded-name { font-weight:700; color:#1a3a22; }
td.num { text-align:right; font-family:"Cascadia Code","Cascadia Mono",Consolas,"Courier New",monospace; }
td.label { background:#f0f4f0; font-weight:600; color:#444; min-width:160px; }
td.link a { color:#1155cc; text-decoration:underline; cursor:pointer; font-weight:600; }
td.link a:hover { color:#0b3d91; }
td.sec { color:#555; font-style:italic; }
.st { display:inline-block; padding:1px 6px; border-radius:3px; font-size:10px; font-weight:700; }
.st-F { background:#d4edda; color:#155724; }
.st-M { background:#d1ecf1; color:#0c5460; }
.st-U { background:#fff3cd; color:#856404; }
.st-O { background:#ffe0b2; color:#7a3000; }
.st-P { background:#e2d9f3; color:#4a235a; }
.st-Z { background:#f8d7da; color:#721c24; }
tr.non-final td { background:#fff0f0 !important; color:#800 !important; }
tr.non-final td .st { border:1px solid #f5c2c7; }
tr.non-final:hover td { background:#ffe0e0 !important; }
.rmk { display:inline-block; padding:1px 5px; border-radius:3px; font-size:10px;
  background:#f0f0f0; color:#333; border:1px solid #ccc; }
.rmk-G { background:#fff3cd; color:#856404; border-color:#ffc107; }
.rmk-B { background:#d1ecf1; color:#0c5460; border-color:#bee5eb; }
.sec-hdr td { background:#1a5c32 !important; color:#fff !important;
  border-color:#1a5c32; font-weight:700; padding:5px 10px; font-size:11px; }
.ass-tbl td { padding:5px 10px; }
.leg-title { font-size:13px; font-weight:700; padding:8px 10px;
  background:#1a5c32; color:#fff; }
td.pbadge { background:#0A1628 !important; color:#fff !important; font-weight:700; text-align:center;
  font-size:10px; }
"""

_JS = """
function show(id) {
  document.querySelectorAll('.sheet').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById('sheet-' + id).classList.add('active');
  const idx = TAB_IDS.indexOf(id);
  if (idx >= 0) document.querySelectorAll('.tab')[idx].classList.add('active');
}
function jumpTo(rowId) {
  setTimeout(() => {
    const el = document.getElementById(rowId);
    if (!el) return;
    const orig = el.style.outline;
    el.scrollIntoView({behavior:'smooth', block:'center'});
    el.style.outline = '3px solid #e6b800';
    setTimeout(() => { el.style.outline = orig; }, 2200);
  }, 60);
}
"""

def _st_badge(status: str) -> str:
    s = (status or "").strip()
    css = f"st-{s}" if s else ""
    return f'<span class="st {css}">{s}</span>' if s else ""


def _rmk_badge(rmk: str) -> str:
    r = (rmk or "").strip()
    if not r or r == "-":
        return ""
    css = f"rmk-{r}" if r in ("G", "B") else "rmk"
    return f'<span class="{css}">{r}</span>'


def _non_final_class(rows_for_deductor) -> bool:
    """True if any detail row has non-F status."""
    for r in rows_for_deductor:
        st = (r.get("Status of Booking") or r.get("Status") or "").strip()
        if st and st not in ("F", "", "-"):
            return True
    return False


def _td_num(val) -> str:
    if isinstance(val, (int, float)):
        v = val
    else:
        v = _fmt_num(str(val))
    if v is None:
        return '<td class="num"></td>'
    return f'<td class="num">{_ind(v)}</td>'


# ── Per-Part HTML builders ────────────────────────────────────────────────────

def _html_part_I_VI(roman: str, pdata: dict, row_ids: dict) -> str:
    """Parts I and VI: deductor + detail rows, flat layout."""
    is_VI = roman == "VI"
    if is_VI:
        hdrs = ["Collector Name", "TAN", "Section", "Transaction Date",
                "Booking Status", "Remarks",
                "Total Amount Paid/Debited (₹)", "Tax Collected (₹)", "TCS Deposited (₹)"]
    else:
        hdrs = ["Deductor Name", "TAN", "Section", "Transaction Date",
                "Booking Status", "Remarks",
                "Amount Paid/Credited (₹)", "Tax Deducted (₹)", "TDS Deposited (₹)"]

    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    sr = 0
    alt = False

    for ded in pdata["rows"]:
        sr += 1
        name = ded.get("Name of Deductor") or ded.get("Name of Collector") or ""
        tan  = ded.get("TAN of Deductor")  or ded.get("TAN of Collector")  or ""
        details = ded.get("_details", [])

        has_nf = _non_final_class(details)
        first_row_id = f"p{roman}r{sr}"
        row_ids[roman][sr] = {"name": name, "tan": tan}

        for di, d in enumerate(details):
            sec    = d.get("Section", "")
            txdate = d.get("Transaction Date", "")
            status = d.get("Status of Booking", "")
            rmk    = d.get("Remarks", "")
            amt    = _fmt_num(d.get("Amount Paid / Credited(Rs.)") or d.get("Amount Paid / Debited(Rs.)") or d.get("Total Amount Paid / Debited(Rs.)") or "")
            tds    = _fmt_num(d.get("Tax Deducted(Rs.)") or d.get("Tax Collected(Rs.)") or "")
            dep    = _fmt_num(d.get("TDS Deposited(Rs.)") or d.get("TCS Deposited(Rs.)") or "")

            nf_cls  = " non-final" if status.strip() in NON_FINAL_STATUSES else ""
            alt_cls = " alt" if alt else ""
            row_id  = f' id="{first_row_id}"' if di == 0 else ""

            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}"{row_id}>'
                f'<td class="ded-name">{name}</td>'
                f'<td>{tan}</td>'
                f'<td class="sec">{sec}</td>'
                f'<td>{txdate}</td>'
                f'<td>{_st_badge(status)}</td>'
                f'<td>{_rmk_badge(rmk)}</td>'
                f'{_td_num(amt)}{_td_num(tds)}{_td_num(dep)}'
                f'</tr>'
            )
            alt = not alt

        # Subtotal
        s_amt = _fmt_num(ded.get("Total Amount Paid / Credited(Rs.)") or ded.get("Total Amount Paid / Debited(Rs.)") or "")
        s_tds = _fmt_num(ded.get("Total Tax Deducted(Rs.)")  or ded.get("Total Tax Collected(Rs.)")  or "")
        s_dep = _fmt_num(ded.get("Total TDS Deposited(Rs.)") or ded.get("Total TCS Deposited(Rs.)") or "")
        nf_warn = " ⚠" if has_nf else ""
        rows_html.append(
            f'<tr class="subtotal">'
            f'<td colspan="5" style="text-align:right;padding-right:8px">'
            f'Sr. {sr}&nbsp;·&nbsp;Subtotal — {name}{nf_warn}</td>'
            f'<td></td>'
            f'{_td_num(s_amt)}{_td_num(s_tds)}{_td_num(s_dep)}'
            f'</tr>'
        )

    # Grand total
    g_amt = sum(_fmt_num(d.get("Total Amount Paid / Credited(Rs.)") or d.get("Total Amount Paid / Debited(Rs.)") or "") or 0 for d in pdata["rows"])
    g_tds = sum(_fmt_num(d.get("Total Tax Deducted(Rs.)") or d.get("Total Tax Collected(Rs.)") or "") or 0 for d in pdata["rows"])
    g_dep = sum(_fmt_num(d.get("Total TDS Deposited(Rs.)") or d.get("Total TCS Deposited(Rs.)") or "") or 0 for d in pdata["rows"])
    rows_html.append(
        f'<tr class="grandtotal">'
        f'<td colspan="5" style="text-align:right;padding-right:8px">GRAND TOTAL — Part-{roman}</td>'
        f'<td></td>'
        f'{_td_num(g_amt)}{_td_num(g_tds)}{_td_num(g_dep)}'
        f'</tr>'
    )

    return (
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_part_II(pdata: dict, row_ids: dict) -> str:
    hdrs = ["Deductor Name", "TAN", "Section", "Transaction Date",
            "Date of Booking", "Remarks",
            "Amount Paid/Credited (₹)", "Tax Deducted (₹)", "TDS Deposited (₹)"]
    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    sr = 0
    alt = False

    for ded in pdata["rows"]:
        sr += 1
        name = ded.get("Name of Deductor", "")
        tan  = ded.get("TAN of Deductor", "")
        details = ded.get("_details", [])
        first_row_id = f"pIIr{sr}"
        row_ids["II"][sr] = {"name": name, "tan": tan}

        for di, d in enumerate(details):
            sec    = d.get("Section", "")
            txdate = d.get("Transaction Date", "")
            bkdate = d.get("Date of Booking", "")
            rmk    = d.get("Remarks", "")
            amt    = _fmt_num(d.get("Amount Paid / Credited(Rs.)", ""))
            tds    = _fmt_num(d.get("Tax Deducted(Rs.)", ""))
            dep    = _fmt_num(d.get("TDS Deposited(Rs.)", ""))
            status = d.get("Status of Booking", "")
            nf_cls  = " non-final" if status.strip() in NON_FINAL_STATUSES else ""
            alt_cls = " alt" if alt else ""
            row_id  = f' id="{first_row_id}"' if di == 0 else ""
            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}"{row_id}>'
                f'<td class="ded-name">{name}</td><td>{tan}</td>'
                f'<td class="sec">{sec}</td><td>{txdate}</td><td>{bkdate}</td>'
                f'<td>{_rmk_badge(rmk)}</td>'
                f'{_td_num(amt)}{_td_num(tds)}{_td_num(dep)}</tr>'
            )
            alt = not alt

        s_amt = _fmt_num(ded.get("Total Amount Paid / Credited(Rs.)", ""))
        s_tds = _fmt_num(ded.get("Total Tax Deducted(Rs.)", ""))
        s_dep = _fmt_num(ded.get("Total TDS Deposited(Rs.)", ""))
        rows_html.append(
            f'<tr class="subtotal">'
            f'<td colspan="5" style="text-align:right;padding-right:8px">Sr. {sr}&nbsp;·&nbsp;Subtotal — {name}</td>'
            f'<td></td>{_td_num(s_amt)}{_td_num(s_tds)}{_td_num(s_dep)}</tr>'
        )

    g_amt = sum((_fmt_num(d.get("Total Amount Paid / Credited(Rs.)", "")) or 0) for d in pdata["rows"])
    g_tds = sum((_fmt_num(d.get("Total Tax Deducted(Rs.)", "")) or 0)           for d in pdata["rows"])
    g_dep = sum((_fmt_num(d.get("Total TDS Deposited(Rs.)", "")) or 0)          for d in pdata["rows"])
    rows_html.append(
        f'<tr class="grandtotal"><td colspan="5" style="text-align:right;padding-right:8px">GRAND TOTAL — Part-II</td>'
        f'<td></td>{_td_num(g_amt)}{_td_num(g_tds)}{_td_num(g_dep)}</tr>'
    )
    return (
        f'<div class="note-bar">ℹ Part-II — Interest/dividend where 15G/15H submitted (no TDS deducted)</div>'
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_part_III(pdata: dict, row_ids: dict) -> str:
    hdrs = ["Deductor Name", "TAN", "Section", "Transaction Date",
            "Booking Status", "Remarks", "Amount Paid/Credited (₹)"]
    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    sr = 0
    alt = False

    for ded in pdata["rows"]:
        sr += 1
        name = ded.get("Name of Deductor", "")
        tan  = ded.get("TAN of Deductor", "")
        details = ded.get("_details", [])
        first_row_id = f"pIIIr{sr}"
        row_ids["III"][sr] = {"name": name, "tan": tan}

        for di, d in enumerate(details):
            sec    = d.get("Section", "")
            txdate = d.get("Transaction Date", "")
            status = d.get("Status of Booking", "")
            rmk    = d.get("Remarks", "")
            amt    = _fmt_num(d.get("Amount Paid / Credited(Rs.)", ""))
            nf_cls  = " non-final" if status.strip() in NON_FINAL_STATUSES else ""
            alt_cls = " alt" if alt else ""
            row_id  = f' id="{first_row_id}"' if di == 0 else ""
            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}"{row_id}>'
                f'<td class="ded-name">{name}</td><td>{tan}</td>'
                f'<td class="sec">{sec}</td><td>{txdate}</td>'
                f'<td>{_st_badge(status)}</td><td>{_rmk_badge(rmk)}</td>'
                f'{_td_num(amt)}</tr>'
            )
            alt = not alt

        s_amt = _fmt_num(ded.get("Total Amount Paid / Credited(Rs.)", ""))
        rows_html.append(
            f'<tr class="subtotal">'
            f'<td colspan="5" style="text-align:right;padding-right:8px">Sr. {sr}&nbsp;·&nbsp;Subtotal — {name}</td>'
            f'<td></td>{_td_num(s_amt)}</tr>'
        )

    g_amt = sum((_fmt_num(d.get("Total Amount Paid / Credited(Rs.)", "")) or 0) for d in pdata["rows"])
    rows_html.append(
        f'<tr class="grandtotal"><td colspan="5" style="text-align:right;padding-right:8px">GRAND TOTAL — Part-III</td>'
        f'<td></td>{_td_num(g_amt)}</tr>'
    )
    return (
        f'<div class="note-bar">ℹ Part-III — Tax paid before release (winnings/benefits/VDA). No separate TDS column.</div>'
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_part_IV_VIII(roman: str, pdata: dict, row_ids: dict) -> str:
    """Parts IV and VIII: property/VDA TDS — flat row per transaction."""
    is_VIII = roman == "VIII"
    if is_VIII:
        hdrs = ["Name of Deductee (Seller/Landlord)", "PAN of Deductee",
                "Ack. No.", "TDS Cert. No.", "Section", "Transaction Date",
                "Date of Deposit", "Booking Status", "Demand Payment",
                "Total Transaction Amount (₹)", "TDS Deposited (₹)", "Other Amount (₹)"]
        name_key = "Name of Deductee"
        pan_key  = "PAN of Deductee"
        note = "ℹ Part-VIII — TDS deducted <strong>by you</strong> as buyer/tenant · Informational only — not included in Summary"
    else:
        hdrs = ["Name of Deductor (Buyer/Tenant)", "PAN of Deductor",
                "Ack. No.", "TDS Cert. No.", "Section", "Transaction Date",
                "Date of Deposit", "Booking Status", "Demand Payment",
                "Total Transaction Amount (₹)", "TDS Deposited (₹)"]
        name_key = "Name of Deductor"
        pan_key  = "PAN of Deductor"
        note = None

    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    sr = 0
    alt = False

    for ded in pdata["rows"]:
        sr += 1
        name = ded.get(name_key) or ded.get("Name of Deductor") or ""
        pan  = ded.get(pan_key)  or ded.get("PAN of Deductor")  or ""
        ack  = ded.get("Acknowledgement Number", "")
        details = ded.get("_details", [])
        first_row_id = f"p{roman}r{sr}"
        row_ids[roman][sr] = {"name": name, "tan": pan}

        for di, d in enumerate(details):
            cert   = d.get("TDS Certificate Number", "")
            sec    = d.get("Section", "")
            dep_dt = d.get("Date of Deposit", "")
            status = d.get("Status of Booking", "")
            demand = d.get("Demand Payment", "")
            tds    = _fmt_num(d.get("TDS Deposited(Rs.)", ""))
            other  = _fmt_num(d.get("Total Amount Deposited other than TDS(Rs.)", "")) if is_VIII else None
            txdate = ded.get("Transaction Date", "")
            txamt  = _fmt_num(ded.get("Total Transaction Amount(Rs.)", ""))

            nf_cls  = " non-final" if status.strip() in NON_FINAL_STATUSES else ""
            alt_cls = " alt" if alt else ""
            row_id  = f' id="{first_row_id}"' if di == 0 else ""

            other_td = _td_num(other) if is_VIII else ""
            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}"{row_id}>'
                f'<td class="ded-name">{name}</td><td>{pan}</td>'
                f'<td>{ack}</td><td>{cert}</td>'
                f'<td class="sec">{sec}</td><td>{txdate}</td><td>{dep_dt}</td>'
                f'<td>{_st_badge(status)}</td><td>{demand}</td>'
                f'{_td_num(txamt)}{_td_num(tds)}{other_td}</tr>'
            )
            alt = not alt

        s_txamt = _fmt_num(ded.get("Total Transaction Amount(Rs.)", ""))
        s_tds   = _fmt_num(ded.get("Total TDS Deposited(Rs.)", ""))
        other_td2 = _td_num(0) if is_VIII else ""
        rows_html.append(
            f'<tr class="subtotal">'
            f'<td colspan="8" style="text-align:right;padding-right:8px">Sr. {sr}&nbsp;·&nbsp;Subtotal — {name}</td>'
            f'<td></td>{_td_num(s_txamt)}{_td_num(s_tds)}{other_td2}</tr>'
        )

    g_tx  = sum((_fmt_num(d.get("Total Transaction Amount(Rs.)", "")) or 0) for d in pdata["rows"])
    g_tds = sum((_fmt_num(d.get("Total TDS Deposited(Rs.)", "")) or 0)      for d in pdata["rows"])
    other_td3 = _td_num(0) if is_VIII else ""
    rows_html.append(
        f'<tr class="grandtotal">'
        f'<td colspan="8" style="text-align:right;padding-right:8px">GRAND TOTAL — Part-{roman}</td>'
        f'<td></td>{_td_num(g_tx)}{_td_num(g_tds)}{other_td3}</tr>'
    )
    prefix = f'<div class="note-bar">{note}</div>' if note else ""
    return (
        prefix +
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_part_V(pdata: dict, row_ids: dict) -> str:
    hdrs = ["Name of Buyer", "PAN of Buyer", "Ack. No.", "Transaction Date",
            "Total Transaction Amount (₹)", "BSR Code", "Date of Deposit",
            "Challan Serial No.", "Total Tax Amount (₹)", "Booking Status",
            "Demand Payment"]
    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    sr = 0
    alt = False

    for ded in pdata["rows"]:
        sr += 1
        name = ded.get("Name of Buyer", "")
        pan  = ded.get("PAN of Buyer", "")
        ack  = ded.get("Acknowledgement Number", "")
        txdate = ded.get("Transaction Date", "")
        txamt  = _fmt_num(ded.get("Total Transaction Amount(Rs.)", ""))
        details = ded.get("_details", [])
        first_row_id = f"pVr{sr}"
        row_ids["V"][sr] = {"name": name, "tan": pan}

        for di, d in enumerate(details):
            bsr    = d.get("BSR Code", "")
            dep_dt = d.get("Date of Deposit", "")
            challan= d.get("Challan Serial Number", "") or d.get("Challan Serial No.", "")
            tax    = _fmt_num(d.get("Total Tax Amount(Rs.)", ""))
            status = d.get("Status of Booking", "")
            demand = d.get("Demand Payment", "")

            nf_cls  = " non-final" if status.strip() in NON_FINAL_STATUSES else ""
            alt_cls = " alt" if alt else ""
            row_id  = f' id="{first_row_id}"' if di == 0 else ""
            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}"{row_id}>'
                f'<td class="ded-name">{name}</td><td>{pan}</td><td>{ack}</td>'
                f'<td>{txdate}</td>{_td_num(txamt)}'
                f'<td>{bsr}</td><td>{dep_dt}</td><td>{challan}</td>'
                f'{_td_num(tax)}<td>{_st_badge(status)}</td><td>{demand}</td></tr>'
            )
            alt = not alt

        s_txamt = _fmt_num(ded.get("Total Transaction Amount(Rs.)", ""))
        rows_html.append(
            f'<tr class="subtotal">'
            f'<td colspan="3" style="text-align:right;padding-right:8px">Sr. {sr}&nbsp;·&nbsp;Subtotal — {name}</td>'
            f'{_td_num(s_txamt)}<td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>'
        )

    g_tx = sum((_fmt_num(d.get("Total Transaction Amount(Rs.)", "")) or 0) for d in pdata["rows"])
    rows_html.append(
        f'<tr class="grandtotal">'
        f'<td colspan="3" style="text-align:right;padding-right:8px">GRAND TOTAL — Part-V</td>'
        f'{_td_num(g_tx)}<td></td><td></td><td></td><td></td><td></td><td></td><td></td></tr>'
    )
    return (
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_part_VII(pdata: dict) -> str:
    hdrs = ["Assessment Year", "Mode", "Refund Issued", "Nature of Refund",
            "Amount of Refund (₹)", "Interest (₹)", "Date of Payment", "Remarks"]
    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    alt = False

    for ded in pdata["rows"]:
        # Part VII has no deductor/detail split — each row is a refund record
        # The summary row itself carries the fields
        ay     = ded.get("Assessment Year", "")
        mode   = ded.get("Mode", "")
        issued = ded.get("Refund Issued", "")
        nature = ded.get("Nature of Refund", "")
        amt    = _fmt_num(ded.get("Amount of Refund(Rs.)", ""))
        intr   = _fmt_num(ded.get("Interest(Rs.)", ""))
        dt     = ded.get("Date of Payment", "")
        rmk    = ded.get("Remarks", "")
        alt_cls = " alt" if alt else ""
        rows_html.append(
            f'<tr class="{alt_cls}"><td>{ay}</td><td>{mode}</td><td>{issued}</td>'
            f'<td>{nature}</td>{_td_num(amt)}{_td_num(intr)}<td>{dt}</td>'
            f'<td>{_rmk_badge(rmk)}</td></tr>'
        )
        alt = not alt

    if not rows_html:
        rows_html = ['<tr class="empty-note"><td colspan="8">No Transactions Present</td></tr>']

    return (
        f'<div class="note-bar">ℹ Part-VII — Refunds paid to assessee · Informational only — not included in Summary</div>'
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_part_IX(pdata: dict, row_ids: dict) -> str:
    hdrs = ["Name of Seller", "PAN of Seller", "Ack. No.", "Transaction Date",
            "Total Transaction Amount (₹)", "BSR Code", "Date of Deposit",
            "Challan Serial No.", "Total Tax Amount (₹)", "Booking Status",
            "Demand Payment", "Other Amount Deposited (₹)"]
    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    sr = 0
    alt = False

    for ded in pdata["rows"]:
        sr += 1
        name = ded.get("Name of Seller", "")
        pan  = ded.get("PAN of Seller", "")
        ack  = ded.get("Acknowledgement Number", "")
        txdate = ded.get("Transaction Date", "")
        txamt  = _fmt_num(ded.get("Total Transaction Amount(Rs.)", ""))
        details = ded.get("_details", [])
        first_row_id = f"pIXr{sr}"
        row_ids["IX"][sr] = {"name": name, "tan": pan}

        for di, d in enumerate(details):
            bsr    = d.get("BSR Code", "")
            dep_dt = d.get("Date of Deposit", "")
            challan= d.get("Challan Serial Number", "") or d.get("Challan Serial No.", "")
            tax    = _fmt_num(d.get("Total Tax Amount(Rs.)", ""))
            status = d.get("Status of Booking", "")
            demand = d.get("Demand Payment", "")
            other  = _fmt_num(d.get("Total Amount Deposited other than TDS(Rs.)", ""))

            nf_cls  = " non-final" if status.strip() in NON_FINAL_STATUSES else ""
            alt_cls = " alt" if alt else ""
            row_id  = f' id="{first_row_id}"' if di == 0 else ""
            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}"{row_id}>'
                f'<td class="ded-name">{name}</td><td>{pan}</td><td>{ack}</td>'
                f'<td>{txdate}</td>{_td_num(txamt)}'
                f'<td>{bsr}</td><td>{dep_dt}</td><td>{challan}</td>'
                f'{_td_num(tax)}<td>{_st_badge(status)}</td><td>{demand}</td>{_td_num(other)}</tr>'
            )
            alt = not alt

    if not rows_html:
        rows_html = ['<tr class="empty-note"><td colspan="12">No Transactions Present</td></tr>']

    return (
        f'<div class="note-bar">ℹ Part-IX — 26QE transactions as Buyer of VDA · Informational only — not included in Summary</div>'
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


def _html_summary(parsed: dict, row_ids: dict) -> str:
    header = parsed["header"]
    parts  = parsed["parts"]

    hdrs = ["Part", "Part Title", "Deductor / Payer / Collector Name", "TAN / PAN",
            "Total Amount (₹)", "Tax Deducted/Collected (₹)", "Tax Deposited (₹)", "⚠ Non-Final Rows"]
    th = "".join(f"<th>{h}</th>" for h in hdrs)
    rows_html = []
    alt = False

    grand_amt = grand_tds = grand_dep = 0.0

    credit_parts = [r for r in ["I","II","III","IV","V","VI"] if r in parts and not parts[r]["empty"]]

    for roman in credit_parts:
        pdata = parts[roman]
        meta  = PART_META.get(roman, {})
        part_label = f"Part-{roman}"
        part_title = meta.get("title", "")
        part_id = f"p{roman}"

        for sr, info in row_ids.get(roman, {}).items():
            name = info["name"]
            tan  = info["tan"]
            row_id = f"p{roman}r{sr}"

            # Find deductor row to get amounts
            ded_rows = [d for d in pdata["rows"]
                        if (d.get("Name of Deductor") or d.get("Name of Collector") or
                            d.get("Name of Buyer") or d.get("Name of Deductee") or "") == name]
            amt = tds = dep = 0.0
            has_nf = False
            nf_count = 0
            if ded_rows:
                dr = ded_rows[0]
                amt = _fmt_num(dr.get("Total Amount Paid / Credited(Rs.)") or
                               dr.get("Total Amount Paid / Debited(Rs.)") or
                               dr.get("Total Transaction Amount(Rs.)") or "") or 0
                tds = _fmt_num(dr.get("Total Tax Deducted(Rs.)") or
                               dr.get("Total Tax Collected(Rs.)") or
                               dr.get("Total TDS Deposited(Rs.)") or "") or 0
                dep = _fmt_num(dr.get("Total TDS Deposited(Rs.)") or
                               dr.get("Total TCS Deposited(Rs.)") or "") or 0
                for d in dr.get("_details", []):
                    st = (d.get("Status of Booking") or "").strip()
                    if st in NON_FINAL_STATUSES:
                        has_nf = True
                        nf_count += 1

            grand_amt += amt
            grand_tds += tds
            grand_dep += dep

            nf_cls = " non-final" if has_nf else ""
            alt_cls = " alt" if alt else ""
            nf_cell = (f'<td class="num" style="color:#c00;font-weight:700">{nf_count}</td>'
                       if nf_count else '<td class="num">0</td>')
            nf_name = f" ⚠" if has_nf else ""

            # Part-III has only amount column, no TDS
            if roman == "III":
                tds_cell = '<td class="num">—</td><td class="num">—</td>'
            else:
                tds_cell = f'{_td_num(tds)}{_td_num(dep)}'

            onclick = f"show('{part_id}');jumpTo('{row_id}')"
            rows_html.append(
                f'<tr class="{alt_cls}{nf_cls}">'
                f'<td class="pbadge">{part_label}</td>'
                f'<td>{part_title}</td>'
                f'<td class="link"><a onclick="{onclick}">{name}{nf_name}</a></td>'
                f'<td>{tan}</td>'
                f'{_td_num(amt)}{tds_cell}{nf_cell}</tr>'
            )
            alt = not alt

        # Part subtotal
        rows_html.append(
            f'<tr class="subtotal">'
            f'<td colspan="3" style="text-align:right;padding-right:8px">'
            f'{part_label} Subtotal ({len(row_ids.get(roman, {}))} entries)</td>'
            f'<td></td><td></td><td></td><td></td><td></td></tr>'
        )

    rows_html.append(
        f'<tr class="grandtotal">'
        f'<td colspan="3" style="text-align:right;padding-right:8px">GRAND TOTAL (Parts I–VI, credit only)</td>'
        f'<td></td>{_td_num(grand_amt)}{_td_num(grand_tds)}{_td_num(grand_dep)}<td></td></tr>'
    )

    info_parts = [r for r in ["VII","VIII","IX"] if r in parts and not parts[r]["empty"]]
    if info_parts:
        rows_html.append(
            f'<tr><td colspan="8" style="text-align:center;padding:5px;font-style:italic;'
            f'color:#888;font-size:10px">Parts '
            + " · ".join(f"VII" if p == "VII" else p for p in info_parts)
            + f' present but informational only (no tax credit)</td></tr>'
        )

    return (
        f'<div class="freeze-bar">↕ Row 1 frozen · Click name to jump to that deductor\'s row · '
        f'Red = non-Final rows present · Parts VII, VIII &amp; IX excluded (informational)</div>'
        f'<div class="tbl-wrap"><table>'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table></div>'
    )


_LEGENDS_HTML = """
<div class="tbl-wrap">
<table>
<tr><td colspan="3" class="leg-title" style="font-size:13px;font-weight:700;padding:8px 10px;background:#1a5c32;color:#fff">Status of Booking — Legend</td></tr>
<tr><td colspan="3" style="padding:4px 10px;font-size:10px;color:#555;background:#f9f9f9">Only <strong>F (Final)</strong> entries are eligible for tax credit. All other statuses are highlighted red in Part sheets.</td></tr>
<tr><th style="width:60px">Code</th><th style="width:120px">Description</th><th>Definition</th></tr>
<tr class="alt"><td style="text-align:center"><span class="st st-U">U</span></td><td>Unmatched</td><td>Deductor has not deposited taxes or furnished incorrect particulars.</td></tr>
<tr><td style="text-align:center"><span class="st st-M">M</span></td><td>Matched</td><td>Challan details in TDS statement match OLTAS records.</td></tr>
<tr class="alt"><td style="text-align:center"><span class="st st-P">P</span></td><td>Provisional</td><td>Provisional credit for Government deductors — changes to F after PAO verification.</td></tr>
<tr><td style="text-align:center"><span class="st st-F">F</span></td><td>Final ✅</td><td>Payment matched and verified. <strong>Eligible for tax credit.</strong></td></tr>
<tr class="alt"><td style="text-align:center"><span class="st st-O">O</span></td><td>Overbooked</td><td>Amount over-claimed. Credit becomes Final only after deductor corrects.</td></tr>
<tr><td style="text-align:center"><span class="st st-Z">Z</span></td><td>Mismatch</td><td>Challan details do not match OLTAS. Updated to M once corrected.</td></tr>
<tr><td colspan="3" class="leg-title" style="font-size:13px;font-weight:700;padding:8px 10px;background:#1a5c32;color:#fff">Remarks — Legend</td></tr>
<tr><th>Code</th><th colspan="2">Description</th></tr>
<tr class="alt"><td style="text-align:center"><span class="rmk">A</span></td><td colspan="2">Rectification of error in challan uploaded by bank</td></tr>
<tr><td style="text-align:center"><span class="rmk rmk-B">B</span></td><td colspan="2">Rectification of error in statement uploaded by deductor</td></tr>
<tr class="alt"><td style="text-align:center"><span class="rmk">D</span></td><td colspan="2">Rectification of error in Form 24G filed by Accounts Officer</td></tr>
<tr><td style="text-align:center"><span class="rmk">E</span></td><td colspan="2">Rectification of error in challan by Assessing Officer</td></tr>
<tr class="alt"><td style="text-align:center"><span class="rmk">F</span></td><td colspan="2">Lower / No deduction certificate u/s 197</td></tr>
<tr><td style="text-align:center"><span class="rmk rmk-G">G</span></td><td colspan="2">Reprocessing of statement — reversal / correction entry (amounts typically negative)</td></tr>
<tr class="alt"><td style="text-align:center"><span class="rmk">T</span></td><td colspan="2">Transporter</td></tr>
</table>
</div>
"""


def _build_html(parsed: dict, row_ids: dict, report_ts: str = "") -> str:
    header = parsed["header"]
    parts  = parsed["parts"]

    pan  = header.get("Permanent Account Number (PAN)", "")
    name = header.get("Name of Assessee", "")
    fy   = header.get("Financial Year", "")
    ay   = header.get("Assessment Year", "")
    pan_status = header.get("Current Status of PAN", "")
    addr_parts = [header.get(f"Address Line {i}", "") for i in range(1, 6)]
    state = header.get("Statecode", "")
    pin   = header.get("Pin Code", "")
    gen_date   = header.get("File Creation Date", datetime.today().strftime("%d-%m-%Y"))
    if not report_ts:
        report_ts = datetime.now().strftime("%d-%b-%Y %I:%M %p")

    addr_html = "<br>".join(a for a in addr_parts if a)
    if state:
        addr_html += f"<br>{state}"
    if pin:
        addr_html += f" — {pin}"

    # Build tabs
    active_parts = [r for r in ["I","II","III","IV","V","VI","VII","VIII","IX"]
                    if r in parts and not parts[r]["empty"]]

    tab_items = [('assessee', '📋 Assessee Details')]
    for roman in active_parts:
        tab_items.append((f"p{roman}", f"Part-{roman}"))
    tab_items.append(('summary', '🗂 Summary'))
    tab_items.append(('legends', '📖 Legends'))

    tabs_html = ""
    for i, (tid, tlabel) in enumerate(tab_items):
        active = ' active' if i == 0 else ''
        tabs_html += f'<div class="tab{active}" onclick="show(\'{tid}\')">{tlabel}</div>\n'

    tab_ids_js = "[" + ",".join(f'"{tid}"' for tid, _ in tab_items) + "]"

    # Assessee sheet
    assessee_html = f"""
<div id="sheet-assessee" class="sheet active">
  <div class="brand">
    <div><div class="brand-title">AayDoc Capio</div>
    <div class="brand-sub">Annual Tax Statement — Form 26AS · AY {ay}</div></div>
    <div class="brand-badge">Generated by AayDoc Capio · {report_ts}</div>
  </div>
  <div class="tbl-wrap">
    <table class="ass-tbl">
      <tr class="sec-hdr"><td colspan="2">ASSESSEE INFORMATION</td></tr>
      <tr><td class="label">Permanent Account Number (PAN)</td><td>{pan}</td></tr>
      <tr class="alt"><td class="label">Current Status of PAN</td>
        <td style="color:#1a5c32;font-weight:700">✅ {pan_status}</td></tr>
      <tr><td class="label">Name of Assessee</td><td>{name}</td></tr>
      <tr class="alt"><td class="label">Financial Year</td><td>{fy}</td></tr>
      <tr><td class="label">Assessment Year</td><td>{ay}</td></tr>
      <tr class="alt"><td class="label">Data Updated Till</td><td>{gen_date}</td></tr>
      <tr><td class="label">Report Generated On</td><td>{report_ts}</td></tr>
      <tr class="sec-hdr"><td colspan="2">ADDRESS</td></tr>
      <tr><td class="label">Address</td><td>{addr_html}</td></tr>
      <tr class="sec-hdr"><td colspan="2">NOTES</td></tr>
      <tr><td colspan="2" style="color:#555;font-size:11px;padding:6px 10px">
        · All amounts in INR<br>
        · Only <strong>Final (F)</strong> status bookings are eligible for tax credit.
          Rows with other statuses (U/M/O/Z/P) are highlighted red.<br>
        · Parts VII, VIII and IX are informational only — not included in Summary.<br>
        · Figures in brackets represent reversal entries (Remarks G/B).
      </td></tr>
    </table>
  </div>
</div>"""

    # Part sheets
    part_sheets_html = ""
    part_builders = {
        "I":   lambda pd: _html_part_I_VI("I",   pd, row_ids),
        "II":  lambda pd: _html_part_II(pd, row_ids),
        "III": lambda pd: _html_part_III(pd, row_ids),
        "IV":  lambda pd: _html_part_IV_VIII("IV",  pd, row_ids),
        "V":   lambda pd: _html_part_V(pd, row_ids),
        "VI":  lambda pd: _html_part_I_VI("VI",  pd, row_ids),
        "VII": lambda pd: _html_part_VII(pd),
        "VIII":lambda pd: _html_part_IV_VIII("VIII", pd, row_ids),
        "IX":  lambda pd: _html_part_IX(pd, row_ids),
    }
    for roman in active_parts:
        pdata   = parts[roman]
        meta    = PART_META.get(roman, {})
        title   = meta.get("title", f"Part-{roman}")
        sheet_id = f"p{roman}"
        body = part_builders[roman](pdata)
        part_sheets_html += (
            f'\n<div id="sheet-{sheet_id}" class="sheet">\n'
            f'  <div class="freeze-bar">Part-{roman} — {title}</div>\n'
            f'  {body}\n'
            f'</div>\n'
        )

    # Summary sheet
    summary_body = _html_summary(parsed, row_ids)
    summary_sheet = f'\n<div id="sheet-summary" class="sheet">\n{summary_body}\n</div>\n'

    # Legends sheet
    legends_sheet = f'\n<div id="sheet-legends" class="sheet">\n{_LEGENDS_HTML}\n</div>\n'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Form 26AS — {name} — AY {ay}</title>
<style>{_CSS}</style>
</head>
<body>
<div class="tab-bar">
{tabs_html}</div>
{assessee_html}
{part_sheets_html}
{summary_sheet}
{legends_sheet}
<div class="footer">
  <span class="footer-copy">© 2026 AayDoc Capio™. All rights reserved. Developed by CA. Deepak Bhholusaria.</span>
  <a href="https://www.linkedin.com/in/bhholusaria/" target="_blank">🔗 linkedin.com/in/bhholusaria</a>
  <a href="mailto:deepak@ailearrning.guru">✉ deepak@ailearrning.guru</a>
</div>
<script>
const TAB_IDS = {tab_ids_js};
{_JS}
</script>
</body>
</html>"""


# ── Excel writer ──────────────────────────────────────────────────────────────

def _safe_save(wb, xlsx_path: str) -> str:
    """Save workbook to xlsx_path. If the file is locked (open in Excel),
    fall back to a timestamped filename so the user still gets their file.
    Returns the actual path written.
    """
    import tempfile, shutil
    # Write to a temp file first — this always succeeds regardless of lock
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx",
                                        dir=os.path.dirname(xlsx_path) or ".")
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
    except Exception:
        os.unlink(tmp_path)
        raise

    # Try to move temp → target
    try:
        shutil.move(tmp_path, xlsx_path)
        return xlsx_path
    except PermissionError:
        # Target locked — save alongside with timestamp suffix
        base, ext = os.path.splitext(xlsx_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = f"{base}_{ts}{ext}"
        shutil.move(tmp_path, alt_path)
        return alt_path


def _write_xlsx(parsed: dict, row_ids: dict, xlsx_path: str, report_ts: str = "") -> str:
    """Returns the actual path the file was saved to (may differ if original was locked)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header = parsed["header"]
    parts  = parsed["parts"]

    # ── Colour / style constants ───────────────────────────────────────────
    C_NAVY   = "FF0A1628"
    C_GREEN  = "FF1a5c32"
    C_SUBTTL = "FFd0e8c8"
    C_RED_BG = "FFFFF0F0"
    C_RED_FG = "FF880000"
    C_LABEL  = "FFF0F4F0"

    _thin = Side(style="thin", color="FFD0D0D0")
    _thick_top = Side(style="medium", color="FF1a5c32")

    def _border():
        return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _border_subtotal():
        return Border(left=_thin, right=_thin, top=_thick_top, bottom=_thin)

    def _f(bold=False, size=10, color="FF000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _fill(color):
        return PatternFill("solid", fgColor=color)

    def _al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _style_hdr(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font  = _f(bold=True, color="FFFFFFFF")
            cell.fill  = _fill(C_GREEN)
            cell.alignment = _al("center")
            cell.border = _border()

    def _style_subtotal(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font  = _f(bold=True, color="FF1a3a22")
            cell.fill  = _fill(C_SUBTTL)
            cell.border = _border_subtotal()

    def _style_grandtotal(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font  = _f(bold=True, color="FFFFFFFF")
            cell.fill  = _fill(C_NAVY)
            cell.border = _border()

    def _style_nonfinal(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill  = _fill(C_RED_BG)
            cell.font  = _f(color=C_RED_FG)

    def _write_str(ws, row, col, val, bold=False, color="FF000000", fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=str(val) if val is not None else "")
        cell.font   = _f(bold=bold, color=color)
        cell.fill   = fill or PatternFill()
        cell.alignment = align or _al()
        cell.border = _border()

    def _write_num(ws, row, col, val):
        v = _fmt_num(str(val)) if not isinstance(val, (int, float)) else val
        cell = ws.cell(row=row, column=col, value=v if v is not None else 0)
        cell.number_format = '#,##0.00'
        cell.alignment = _al("right")
        cell.font  = _f()
        cell.border = _border()

    def _write_formula(ws, row, col, formula, number_format='#,##0.00'):
        cell = ws.cell(row=row, column=col, value=formula)
        cell.number_format = number_format
        cell.alignment = _al("right")
        cell.font  = _f(bold=True, color="FF1a3a22")
        cell.border = _border_subtotal()

    def _autofit(ws):
        col_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    txt = str(cell.value)
                    # strip Excel formula for width calc
                    if txt.startswith("="):
                        txt = txt.split('"')[1] if '"' in txt else txt
                    col_widths[cell.column] = min(
                        max(col_widths.get(cell.column, 8), len(txt) + 2), 52)
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    def _brand_row(ws, ncols, ay, assessee=""):
        """Row 1: left cell = '26AS — Assessee — AY', right cell = AayDoc branding.
        Both on navy background."""
        last_col = get_column_letter(ncols)

        # Left: title spanning first 4 cols (or all if ncols <= 4)
        split_col = min(4, ncols - 1) if ncols > 1 else 1
        split_letter = get_column_letter(split_col)
        if split_col > 1:
            ws.merge_cells(f"A1:{split_letter}1")
        left = ws["A1"]
        left.value = f"Form 26AS  —  {assessee}  —  AY {ay}"
        left.font  = _f(bold=True, size=10, color="FFFFFFFF")
        left.fill  = _fill(C_NAVY)
        left.alignment = _al("left")

        # Right: branding spanning remaining cols
        right_start = get_column_letter(split_col + 1)
        if split_col + 1 <= ncols:
            if split_col + 1 < ncols:
                ws.merge_cells(f"{right_start}1:{last_col}1")
            right = ws[f"{right_start}1"]
            right.value = f"AayDoc Capio™  ·  {report_ts}  ·  © 2026  ·  CA. Deepak Bhholusaria"
            right.font  = _f(bold=False, size=8, color="FF94A3B8")
            right.fill  = _fill(C_NAVY)
            right.alignment = _al("right")

        ws.row_dimensions[1].height = 16

    def _group_rows(ws, start, end):
        """Add outline grouping so detail rows can be collapsed under subtotal."""
        for r in range(start, end + 1):
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = False
        ws.sheet_properties.outlinePr.summaryBelow = True

    # ── Assessee Details ───────────────────────────────────────────────────
    ay            = header.get("Assessment Year", "")
    assessee_name = header.get("Name of Assessee", "")
    ws_ass = wb.create_sheet("Assessee Details")
    ws_ass.sheet_properties.tabColor = "1a5c32"

    # Title row
    ws_ass.merge_cells("A1:B1")
    ws_ass["A1"] = f"Form 26AS  —  {assessee_name}  —  AY {ay}"
    ws_ass["A1"].font  = _f(bold=True, size=13, color="FFFFFFFF")
    ws_ass["A1"].fill  = _fill(C_NAVY)
    ws_ass["A1"].alignment = _al("center")
    ws_ass["A1"].border = _border()
    ws_ass.row_dimensions[1].height = 24

    # Branding sub-line
    ws_ass.merge_cells("A2:B2")
    ws_ass["A2"] = "AayDoc Capio™  ·  © 2026  ·  Developed by CA. Deepak Bhholusaria"
    ws_ass["A2"].font  = _f(bold=False, size=8, color="FF94A3B8")
    ws_ass["A2"].fill  = _fill(C_NAVY)
    ws_ass["A2"].alignment = _al("center")
    ws_ass["A2"].border = _border()
    ws_ass.row_dimensions[2].height = 14

    fields_map = [
        ("ASSESSEE INFORMATION", None),
        ("Permanent Account Number (PAN)", "Permanent Account Number (PAN)"),
        ("Current Status of PAN",          "Current Status of PAN"),
        ("Name of Assessee",               "Name of Assessee"),
        ("Financial Year",                 "Financial Year"),
        ("Assessment Year",                "Assessment Year"),
        ("Data Updated Till",              "File Creation Date"),
        ("Report Generated On",            "__report_ts__"),
        ("ADDRESS", None),
        ("Address Line 1",  "Address Line 1"),
        ("Address Line 2",  "Address Line 2"),
        ("Address Line 3",  "Address Line 3"),
        ("Address Line 4",  "Address Line 4"),
        ("Address Line 5",  "Address Line 5"),
        ("State",           "Statecode"),
        ("PIN Code",        "Pin Code"),
    ]
    r = 3
    for label, key in fields_map:
        if key is None:
            ws_ass.merge_cells(f"A{r}:B{r}")
            ws_ass[f"A{r}"] = label
            ws_ass[f"A{r}"].font   = _f(bold=True, color="FFFFFFFF")
            ws_ass[f"A{r}"].fill   = _fill(C_GREEN)
            ws_ass[f"A{r}"].alignment = _al()
            ws_ass[f"A{r}"].border = _border()
        else:
            ws_ass[f"A{r}"] = label
            ws_ass[f"A{r}"].font   = _f(bold=True)
            ws_ass[f"A{r}"].fill   = _fill(C_LABEL)
            ws_ass[f"A{r}"].alignment = _al()
            ws_ass[f"A{r}"].border = _border()
            ws_ass[f"B{r}"] = report_ts if key == "__report_ts__" else header.get(key, "")
            ws_ass[f"B{r}"].font   = _f()
            ws_ass[f"B{r}"].alignment = _al()
            ws_ass[f"B{r}"].border = _border()
        r += 1

    ws_ass.column_dimensions["A"].width = 35
    ws_ass.column_dimensions["B"].width = 48

    # ── Notes section ──────────────────────────────────────────────────────
    ws_ass.merge_cells(f"A{r}:B{r}")
    ws_ass[f"A{r}"] = "NOTES"
    ws_ass[f"A{r}"].font   = _f(bold=True, color="FFFFFFFF")
    ws_ass[f"A{r}"].fill   = _fill(C_GREEN)
    ws_ass[f"A{r}"].alignment = _al()
    ws_ass[f"A{r}"].border = _border()
    r += 1

    notes = (
        "· All amounts in INR\n"
        "· Only Final (F) status bookings are eligible for tax credit. "
        "Rows with other statuses (U/M/O/Z/P) are highlighted red.\n"
        "· Parts VII, VIII and IX are informational only — not included in Summary.\n"
        "· Figures in brackets represent reversal entries (Remarks G/B).\n\n"
        "© 2026 AayDoc Capio™. All rights reserved. Developed by CA. Deepak Bhholusaria.\n"
        "LinkedIn: https://www.linkedin.com/in/bhholusaria/\n"
        "Email: deepak@ailearrning.guru"
    )
    ws_ass.merge_cells(f"A{r}:B{r}")
    cell = ws_ass[f"A{r}"]
    cell.value = notes
    cell.font  = _f(size=10)
    cell.fill  = _fill("FFFFF9F0")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = _border()
    ws_ass.row_dimensions[r].height = 72

    # ── Generic deductor+detail part writer (Parts I, II, III, VI) ────────
    def write_deductor_part(roman, cols, num_cols,
                            get_name, get_tan,
                            get_detail_vals,     # fn(ded, detail) -> list of (col_idx, val, is_num)
                            get_sum_vals,        # fn(ded) -> list of (col_idx, sum_val)
                            get_grand_vals,      # fn(rows) -> list of (col_idx, sum_val)
                            detail_status_key="Status of Booking"):
        meta  = PART_META.get(roman, {})
        pdata = parts[roman]
        ncols = len(cols)
        ws    = wb.create_sheet(f"Part-{roman}")
        ws.sheet_properties.tabColor = "1a5c32"

        _brand_row(ws, ncols, ay, assessee=assessee_name)
        r = 2
        for c, h in enumerate(cols, 1):
            _write_str(ws, r, c, h)
        _style_hdr(ws, r, ncols)
        ws.freeze_panes = f"A{r+1}"
        r += 1

        sr = 0
        for ded in pdata["rows"]:
            sr += 1
            name    = get_name(ded)
            tan     = get_tan(ded)
            details = ded.get("_details", [])
            row_ids[roman][sr] = {"name": name, "tan": tan, "xl_row": r}
            detail_start = r

            for d in details:
                status = d.get(detail_status_key, "").strip()
                vals = get_detail_vals(ded, d)
                for col_idx, val, is_num in vals:
                    if is_num:
                        _write_num(ws, r, col_idx, val or 0)
                    else:
                        _write_str(ws, r, col_idx, val or "")
                if status in NON_FINAL_STATUSES:
                    _style_nonfinal(ws, r, ncols)
                r += 1

            # Subtotal row with SUM formulas
            detail_end = r - 1
            nf_warn = " ⚠" if _non_final_class(details) else ""
            merge_to = get_column_letter(min(ncols - len(num_cols), ncols))
            ws.merge_cells(f"A{r}:{merge_to}{r}")
            _write_str(ws, r, 1,
                       f"Sr. {sr}  ·  Subtotal — {name}{nf_warn}",
                       bold=True)
            ws[f"A{r}"].alignment = _al("right")
            for col_idx, _ in get_sum_vals(ded):
                col_letter = get_column_letter(col_idx)
                if detail_start <= detail_end:
                    _write_formula(ws, r, col_idx,
                                   f"=SUM({col_letter}{detail_start}:{col_letter}{detail_end})")
                else:
                    _write_num(ws, r, col_idx, 0)
            _style_subtotal(ws, r, ncols)
            _group_rows(ws, detail_start, detail_end)
            r += 1

        # Grand total row
        subtotal_rows = [row_ids[roman][s]["xl_row"] + len(parts[roman]["rows"][s-1].get("_details",[]))
                         for s in row_ids[roman]]
        ws.merge_cells(f"A{r}:{merge_to}{r}")
        _write_str(ws, r, 1, f"GRAND TOTAL — Part-{roman}", bold=True,
                   color="FFFFFFFF", fill=_fill(C_NAVY))
        ws[f"A{r}"].alignment = _al("right")
        # Collect all subtotal row numbers for grand total SUM
        sub_rows = []
        sr2 = 0
        for ded in pdata["rows"]:
            sr2 += 1
            info = row_ids[roman].get(sr2, {})
            det_count = len(ded.get("_details", []))
            sub_row = info.get("xl_row", 3) + det_count
            sub_rows.append(sub_row)
        for col_idx, _ in get_grand_vals(pdata["rows"]):
            col_letter = get_column_letter(col_idx)
            refs = "+".join(f"{col_letter}{sr}" for sr in sub_rows)
            if refs:
                cell = ws.cell(row=r, column=col_idx,
                               value=f"={refs}")
                cell.number_format = '#,##0.00'
                cell.alignment = _al("right")
                cell.font  = _f(bold=True, color="FF7fff8a")
                cell.border = _border()
        _style_grandtotal(ws, r, ncols)

        _autofit(ws)

    # ── Part-I / VI ────────────────────────────────────────────────────────
    for roman in ["I", "VI"]:
        if roman not in parts or parts[roman]["empty"]:
            continue
        is_VI = roman == "VI"
        if is_VI:
            cols = ["Collector Name", "TAN", "Section", "Transaction Date",
                    "Booking Status", "Remarks",
                    "Total Amount Paid/Debited (Rs.)", "Tax Collected (Rs.)", "TCS Deposited (Rs.)"]
        else:
            cols = ["Deductor Name", "TAN", "Section", "Transaction Date",
                    "Booking Status", "Remarks",
                    "Amount Paid/Credited (Rs.)", "Tax Deducted (Rs.)", "TDS Deposited (Rs.)"]

        def _dv(ded, d, _is_VI=is_VI):
            status = d.get("Status of Booking", "").strip()
            rmk    = d.get("Remarks", "")
            amt_k  = "Amount Paid / Debited(Rs.)" if _is_VI else "Amount Paid / Credited(Rs.)"
            tds_k  = "Tax Collected(Rs.)" if _is_VI else "Tax Deducted(Rs.)"
            dep_k  = "TCS Deposited(Rs.)" if _is_VI else "TDS Deposited(Rs.)"
            name = ded.get("Name of Deductor") or ded.get("Name of Collector") or ""
            tan  = ded.get("TAN of Deductor")  or ded.get("TAN of Collector")  or ""
            return [
                (1, name, False), (2, tan, False),
                (3, d.get("Section",""), False),
                (4, d.get("Transaction Date",""), False),
                (5, STATUS_FULL.get(status, status), False),
                (6, rmk if rmk and rmk != "-" else "", False),
                (7, _fmt_num(d.get(amt_k,"")) , True),
                (8, _fmt_num(d.get(tds_k,"")) , True),
                (9, _fmt_num(d.get(dep_k,"")) , True),
            ]

        def _sv(ded, _is_VI=is_VI):
            return [(7, None), (8, None), (9, None)]

        def _gv(rows, _is_VI=is_VI):
            return [(7, None), (8, None), (9, None)]

        write_deductor_part(
            roman, cols, [7, 8, 9],
            get_name=lambda d: d.get("Name of Deductor") or d.get("Name of Collector") or "",
            get_tan =lambda d: d.get("TAN of Deductor")  or d.get("TAN of Collector")  or "",
            get_detail_vals=_dv, get_sum_vals=_sv, get_grand_vals=_gv,
        )

    # ── Part-II ────────────────────────────────────────────────────────────
    if "II" in parts and not parts["II"]["empty"]:
        cols2 = ["Deductor Name", "TAN", "Section", "Transaction Date",
                 "Date of Booking", "Remarks",
                 "Amount Paid/Credited (Rs.)", "Tax Deducted (Rs.)", "TDS Deposited (Rs.)"]
        def _dv2(ded, d):
            status = d.get("Status of Booking", "").strip()
            rmk = d.get("Remarks", "")
            return [
                (1, ded.get("Name of Deductor",""), False),
                (2, ded.get("TAN of Deductor",""), False),
                (3, d.get("Section",""), False),
                (4, d.get("Transaction Date",""), False),
                (5, d.get("Date of Booking",""), False),
                (6, rmk if rmk and rmk != "-" else "", False),
                (7, _fmt_num(d.get("Amount Paid / Credited(Rs.)","")) , True),
                (8, _fmt_num(d.get("Tax Deducted(Rs.)","")) , True),
                (9, _fmt_num(d.get("TDS Deposited(Rs.)","")) , True),
            ]
        write_deductor_part("II", cols2, [7,8,9],
            get_name=lambda d: d.get("Name of Deductor",""),
            get_tan =lambda d: d.get("TAN of Deductor",""),
            get_detail_vals=_dv2,
            get_sum_vals=lambda d: [(7,None),(8,None),(9,None)],
            get_grand_vals=lambda rows: [(7,None),(8,None),(9,None)])

    # ── Part-III ───────────────────────────────────────────────────────────
    if "III" in parts and not parts["III"]["empty"]:
        cols3 = ["Deductor Name", "TAN", "Section", "Transaction Date",
                 "Booking Status", "Remarks", "Amount Paid/Credited (Rs.)"]
        def _dv3(ded, d):
            status = d.get("Status of Booking","").strip()
            rmk = d.get("Remarks","")
            return [
                (1, ded.get("Name of Deductor",""), False),
                (2, ded.get("TAN of Deductor",""), False),
                (3, d.get("Section",""), False),
                (4, d.get("Transaction Date",""), False),
                (5, STATUS_FULL.get(status, status), False),
                (6, rmk if rmk and rmk != "-" else "", False),
                (7, _fmt_num(d.get("Amount Paid / Credited(Rs.)","")) , True),
            ]
        write_deductor_part("III", cols3, [7],
            get_name=lambda d: d.get("Name of Deductor",""),
            get_tan =lambda d: d.get("TAN of Deductor",""),
            get_detail_vals=_dv3,
            get_sum_vals=lambda d: [(7,None)],
            get_grand_vals=lambda rows: [(7,None)])

    # ── Part-IV / VIII ─────────────────────────────────────────────────────
    for roman in ["IV", "VIII"]:
        if roman not in parts or parts[roman]["empty"]:
            continue
        is_VIII = roman == "VIII"
        if is_VIII:
            cols = ["Name of Deductee", "PAN of Deductee", "Ack. No.", "TDS Cert. No.",
                    "Section", "Transaction Date", "Date of Deposit", "Booking Status",
                    "Demand Payment", "Total Transaction Amount (Rs.)",
                    "TDS Deposited (Rs.)", "Other Amount (Rs.)"]
            name_key, pan_key = "Name of Deductee", "PAN of Deductee"
        else:
            cols = ["Name of Deductor (Buyer)", "PAN of Deductor", "Ack. No.", "TDS Cert. No.",
                    "Section", "Transaction Date", "Date of Deposit", "Booking Status",
                    "Demand Payment", "Total Transaction Amount (Rs.)", "TDS Deposited (Rs.)"]
            name_key, pan_key = "Name of Deductor", "PAN of Deductor"

        ncols = len(cols)
        pdata = parts[roman]
        ws    = wb.create_sheet(f"Part-{roman}")
        ws.sheet_properties.tabColor = "1a5c32"
        _brand_row(ws, ncols, ay, assessee=assessee_name)
        r = 2
        for c, h in enumerate(cols, 1): _write_str(ws, r, c, h)
        _style_hdr(ws, r, ncols)
        ws.freeze_panes = f"A{r+1}"
        r += 1

        sr = 0
        for ded in pdata["rows"]:
            sr += 1
            name   = (ded.get(name_key) or ded.get("Name of Deductor") or "").strip()
            pan    = (ded.get(pan_key)  or ded.get("PAN of Deductor")  or "").strip()
            ack    = ded.get("Acknowledgement Number", "")
            txdate = ded.get("Transaction Date", "")
            txamt  = _fmt_num(ded.get("Total Transaction Amount(Rs.)","")) or 0
            details = ded.get("_details", [])
            row_ids[roman][sr] = {"name": name, "tan": pan, "xl_row": r}
            detail_start = r

            for d in details:
                status = d.get("Status of Booking","").strip()
                _write_str(ws, r, 1, name, bold=True)
                _write_str(ws, r, 2, pan)
                _write_str(ws, r, 3, ack)
                _write_str(ws, r, 4, d.get("TDS Certificate Number",""))
                _write_str(ws, r, 5, d.get("Section",""))
                _write_str(ws, r, 6, txdate)
                _write_str(ws, r, 7, d.get("Date of Deposit",""))
                _write_str(ws, r, 8, STATUS_FULL.get(status, status))
                _write_str(ws, r, 9, d.get("Demand Payment",""))
                _write_num(ws, r, 10, txamt)
                _write_num(ws, r, 11, _fmt_num(d.get("TDS Deposited(Rs.)","")) or 0)
                if is_VIII:
                    _write_num(ws, r, 12, _fmt_num(d.get("Total Amount Deposited other than TDS(Rs.)","")) or 0)
                if status in NON_FINAL_STATUSES:
                    _style_nonfinal(ws, r, ncols)
                r += 1

            detail_end = r - 1
            merge_end = get_column_letter(9)
            ws.merge_cells(f"A{r}:{merge_end}{r}")
            _write_str(ws, r, 1, f"Sr. {sr}  ·  Subtotal — {name}", bold=True)
            ws[f"A{r}"].alignment = _al("right")
            for ci in ([10, 11, 12] if is_VIII else [10, 11]):
                cl = get_column_letter(ci)
                if detail_start <= detail_end:
                    _write_formula(ws, r, ci, f"=SUM({cl}{detail_start}:{cl}{detail_end})")
                else:
                    _write_num(ws, r, ci, 0)
            _style_subtotal(ws, r, ncols)
            _group_rows(ws, detail_start, detail_end)
            r += 1

        # grand total
        sub_rows2 = []
        sr2 = 0
        for ded in pdata["rows"]:
            sr2 += 1
            info = row_ids[roman].get(sr2, {})
            sub_rows2.append(info.get("xl_row", 3) + len(ded.get("_details",[])))
        ws.merge_cells(f"A{r}:I{r}")
        _write_str(ws, r, 1, f"GRAND TOTAL — Part-{roman}", bold=True,
                   color="FFFFFFFF", fill=_fill(C_NAVY))
        ws[f"A{r}"].alignment = _al("right")
        for ci in ([10, 11, 12] if is_VIII else [10, 11]):
            cl = get_column_letter(ci)
            refs = "+".join(f"{cl}{sr}" for sr in sub_rows2)
            if refs:
                cell = ws.cell(row=r, column=ci, value=f"={refs}")
                cell.number_format = '#,##0.00'
                cell.alignment = _al("right")
                cell.font = _f(bold=True, color="FF7fff8a")
                cell.border = _border()
        _style_grandtotal(ws, r, ncols)
        _autofit(ws)

    # ── Part-V ─────────────────────────────────────────────────────────────
    if "V" in parts and not parts["V"]["empty"]:
        pdata = parts["V"]
        cols5 = ["Name of Buyer", "PAN of Buyer", "Ack. No.", "Transaction Date",
                 "Total Transaction Amount (Rs.)", "BSR Code", "Date of Deposit",
                 "Challan Serial No.", "Total Tax Amount (Rs.)", "Booking Status", "Demand Payment"]
        ncols5 = len(cols5)
        ws5 = wb.create_sheet("Part-V")
        ws5.sheet_properties.tabColor = "1a5c32"
        _brand_row(ws5, ncols5, ay, assessee=assessee_name)
        r = 2
        for c, h in enumerate(cols5, 1): _write_str(ws5, r, c, h)
        _style_hdr(ws5, r, ncols5)
        ws5.freeze_panes = f"A{r+1}"
        r += 1

        sr = 0
        for ded in pdata["rows"]:
            sr += 1
            name   = ded.get("Name of Buyer","")
            pan    = ded.get("PAN of Buyer","")
            ack    = ded.get("Acknowledgement Number","")
            txdate = ded.get("Transaction Date","")
            txamt  = _fmt_num(ded.get("Total Transaction Amount(Rs.)","")) or 0
            details = ded.get("_details",[])
            row_ids["V"][sr] = {"name": name, "tan": pan, "xl_row": r}
            detail_start = r

            for d in details:
                status = d.get("Status of Booking","").strip()
                _write_str(ws5, r, 1, name, bold=True)
                _write_str(ws5, r, 2, pan)
                _write_str(ws5, r, 3, ack)
                _write_str(ws5, r, 4, txdate)
                _write_num(ws5, r, 5, txamt)
                _write_str(ws5, r, 6, d.get("BSR Code",""))
                _write_str(ws5, r, 7, d.get("Date of Deposit",""))
                _write_str(ws5, r, 8, d.get("Challan Serial Number","") or d.get("Challan Serial No.",""))
                _write_num(ws5, r, 9, _fmt_num(d.get("Total Tax Amount(Rs.)","")) or 0)
                _write_str(ws5, r, 10, STATUS_FULL.get(status, status))
                _write_str(ws5, r, 11, d.get("Demand Payment",""))
                if status in NON_FINAL_STATUSES:
                    _style_nonfinal(ws5, r, ncols5)
                r += 1

            detail_end = r - 1
            ws5.merge_cells(f"A{r}:D{r}")
            _write_str(ws5, r, 1, f"Sr. {sr}  ·  Subtotal — {name}", bold=True)
            ws5[f"A{r}"].alignment = _al("right")
            _write_formula(ws5, r, 5, f"=SUM(E{detail_start}:E{detail_end})")
            _style_subtotal(ws5, r, ncols5)
            _group_rows(ws5, detail_start, detail_end)
            r += 1

        sub_rows5 = [row_ids["V"][s]["xl_row"] + len(pdata["rows"][s-1].get("_details",[])) for s in row_ids["V"]]
        ws5.merge_cells(f"A{r}:D{r}")
        _write_str(ws5, r, 1, "GRAND TOTAL — Part-V", bold=True, color="FFFFFFFF", fill=_fill(C_NAVY))
        ws5[f"A{r}"].alignment = _al("right")
        refs5 = "+".join(f"E{sr}" for sr in sub_rows5)
        if refs5:
            cell = ws5.cell(row=r, column=5, value=f"={refs5}")
            cell.number_format = '#,##0.00'; cell.alignment = _al("right")
            cell.font = _f(bold=True, color="FF7fff8a"); cell.border = _border()
        _style_grandtotal(ws5, r, ncols5)
        _autofit(ws5)

    # ── Part-VII ───────────────────────────────────────────────────────────
    if "VII" in parts and not parts["VII"]["empty"]:
        pdata7 = parts["VII"]
        cols7  = ["Assessment Year", "Mode", "Refund Issued", "Nature of Refund",
                  "Amount of Refund (Rs.)", "Interest (Rs.)", "Date of Payment", "Remarks"]
        ncols7 = len(cols7)
        ws7 = wb.create_sheet("Part-VII")
        ws7.sheet_properties.tabColor = "1a5c32"
        _brand_row(ws7, ncols7, ay, assessee=assessee_name)
        r = 2
        for c, h in enumerate(cols7, 1): _write_str(ws7, r, c, h)
        _style_hdr(ws7, r, ncols7)
        ws7.freeze_panes = f"A{r+1}"
        r += 1
        for ded in pdata7["rows"]:
            _write_str(ws7, r, 1, ded.get("Assessment Year",""))
            _write_str(ws7, r, 2, ded.get("Mode",""))
            _write_str(ws7, r, 3, ded.get("Refund Issued",""))
            _write_str(ws7, r, 4, ded.get("Nature of Refund",""))
            _write_num(ws7, r, 5, _fmt_num(ded.get("Amount of Refund(Rs.)","")) or 0)
            _write_num(ws7, r, 6, _fmt_num(ded.get("Interest(Rs.)","")) or 0)
            _write_str(ws7, r, 7, ded.get("Date of Payment",""))
            rmk = ded.get("Remarks",""); _write_str(ws7, r, 8, rmk if rmk != "-" else "")
            r += 1
        _autofit(ws7)

    # ── Part-IX ────────────────────────────────────────────────────────────
    if "IX" in parts and not parts["IX"]["empty"]:
        pdata9 = parts["IX"]
        cols9  = ["Name of Seller", "PAN of Seller", "Ack. No.", "Transaction Date",
                  "Total Transaction Amount (Rs.)", "BSR Code", "Date of Deposit",
                  "Challan Serial No.", "Total Tax Amount (Rs.)", "Booking Status",
                  "Demand Payment", "Other Amount Deposited (Rs.)"]
        ncols9 = len(cols9)
        ws9 = wb.create_sheet("Part-IX")
        ws9.sheet_properties.tabColor = "1a5c32"
        _brand_row(ws9, ncols9, ay, assessee=assessee_name)
        r = 2
        for c, h in enumerate(cols9, 1): _write_str(ws9, r, c, h)
        _style_hdr(ws9, r, ncols9)
        ws9.freeze_panes = f"A{r+1}"
        r += 1
        sr = 0
        for ded in pdata9["rows"]:
            sr += 1
            name = ded.get("Name of Seller",""); pan = ded.get("PAN of Seller","")
            ack = ded.get("Acknowledgement Number","")
            txdate = ded.get("Transaction Date","")
            txamt  = _fmt_num(ded.get("Total Transaction Amount(Rs.)","")) or 0
            row_ids["IX"][sr] = {"name": name, "tan": pan, "xl_row": r}
            for d in ded.get("_details",[]):
                status = d.get("Status of Booking","").strip()
                _write_str(ws9, r, 1, name, bold=True); _write_str(ws9, r, 2, pan)
                _write_str(ws9, r, 3, ack); _write_str(ws9, r, 4, txdate)
                _write_num(ws9, r, 5, txamt)
                _write_str(ws9, r, 6, d.get("BSR Code","")); _write_str(ws9, r, 7, d.get("Date of Deposit",""))
                _write_str(ws9, r, 8, d.get("Challan Serial Number","") or d.get("Challan Serial No.",""))
                _write_num(ws9, r, 9, _fmt_num(d.get("Total Tax Amount(Rs.)","")) or 0)
                _write_str(ws9, r, 10, STATUS_FULL.get(status, status))
                _write_str(ws9, r, 11, d.get("Demand Payment",""))
                _write_num(ws9, r, 12, _fmt_num(d.get("Total Amount Deposited other than TDS(Rs.)","")) or 0)
                if status in NON_FINAL_STATUSES: _style_nonfinal(ws9, r, ncols9)
                r += 1
        _autofit(ws9)

    # ── Summary sheet ──────────────────────────────────────────────────────
    sum_cols = ["Part", "Part Title", "Deductor / Payer / Collector Name", "TAN / PAN",
                "Total Amount (Rs.)", "Tax Deducted/Collected (Rs.)", "Tax Deposited (Rs.)",
                "Non-Final Rows"]
    ncols_sum = len(sum_cols)
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "1a5c32"
    _brand_row(ws_sum, ncols_sum, ay, assessee=assessee_name)
    r = 2
    for c, h in enumerate(sum_cols, 1): _write_str(ws_sum, r, c, h)
    _style_hdr(ws_sum, r, ncols_sum)
    ws_sum.freeze_panes = f"A{r+1}"
    r += 1

    credit_parts = [p for p in ["I","II","III","IV","V","VI"]
                    if p in parts and not parts[p]["empty"]]
    for roman in credit_parts:
        pdata = parts[roman]
        meta  = PART_META.get(roman, {})
        part_label = f"Part-{roman}"
        sheet_name = f"Part-{roman}"

        for sr, info in row_ids.get(roman, {}).items():
            name   = info["name"]
            tan    = info["tan"]
            xl_row = info.get("xl_row", 2)

            ded_rows = [d for d in pdata["rows"]
                        if (d.get("Name of Deductor") or d.get("Name of Collector") or
                            d.get("Name of Buyer") or d.get("Name of Deductee") or "").strip() == name.strip()]
            amt = tds = dep = 0.0
            nf_count = 0
            if ded_rows:
                dr = ded_rows[0]
                amt = _fmt_num(dr.get("Total Amount Paid / Credited(Rs.)") or
                               dr.get("Total Amount Paid / Debited(Rs.)") or
                               dr.get("Total Transaction Amount(Rs.)") or "") or 0
                tds = _fmt_num(dr.get("Total Tax Deducted(Rs.)") or
                               dr.get("Total Tax Collected(Rs.)") or
                               dr.get("Total TDS Deposited(Rs.)") or "") or 0
                dep = _fmt_num(dr.get("Total TDS Deposited(Rs.)") or
                               dr.get("Total TCS Deposited(Rs.)") or "") or 0
                for d in dr.get("_details", []):
                    if (d.get("Status of Booking") or "").strip() in NON_FINAL_STATUSES:
                        nf_count += 1

            _write_str(ws_sum, r, 1, part_label, bold=True, color="FFFFFFFF", fill=_fill(C_NAVY))
            ws_sum.cell(r, 1).alignment = _al("center")
            _write_str(ws_sum, r, 2, meta.get("title",""))
            # Hyperlink to exact deductor row in Part sheet
            link = f'=HYPERLINK("#\'{sheet_name}\'!A{xl_row}","{name}")'
            cell = ws_sum.cell(row=r, column=3, value=link)
            cell.font = Font(name="Calibri", color="FF1155CC", underline="single", size=10)
            cell.alignment = _al(); cell.border = _border()
            _write_str(ws_sum, r, 4, tan)
            _write_num(ws_sum, r, 5, amt)
            if roman == "III":
                _write_str(ws_sum, r, 6, "—"); _write_str(ws_sum, r, 7, "—")
            else:
                _write_num(ws_sum, r, 6, tds); _write_num(ws_sum, r, 7, dep)
            _write_str(ws_sum, r, 8, str(nf_count) if nf_count else "0")
            if nf_count: _style_nonfinal(ws_sum, r, ncols_sum)
            r += 1

        # Part subtotal
        ws_sum.merge_cells(f"A{r}:C{r}")
        _write_str(ws_sum, r, 1,
                   f"{part_label} Subtotal — {len(row_ids.get(roman, {}))} entries",
                   bold=True)
        ws_sum.cell(r, 1).alignment = _al("right")
        _style_subtotal(ws_sum, r, ncols_sum)
        r += 1

    # Grand total
    ws_sum.merge_cells(f"A{r}:C{r}")
    _write_str(ws_sum, r, 1, "GRAND TOTAL  (Parts I–VI, credit only)",
               bold=True, color="FFFFFFFF", fill=_fill(C_NAVY))
    ws_sum.cell(r, 1).alignment = _al("right")
    _style_grandtotal(ws_sum, r, ncols_sum)
    _autofit(ws_sum)

    # ── File properties ────────────────────────────────────────────────────
    assessee_name = parsed["header"].get("Name of Assessee", "")
    assessee_pan  = parsed["header"].get("Permanent Account Number (PAN)", "")
    ay_label      = parsed["header"].get("Assessment Year", "")
    wb.properties.creator  = "AayDoc Capio"
    wb.properties.lastModifiedBy = "AayDoc Capio"
    wb.properties.title    = f"Form 26AS — {assessee_name} — AY {ay_label}"
    wb.properties.subject  = f"Annual Tax Statement | PAN: {assessee_pan} | AY: {ay_label}"
    wb.properties.keywords = f"26AS, TDS, TCS, {assessee_pan}, {ay_label}"
    wb.properties.description = f"Generated by AayDoc Capio on {report_ts}"

    # ── Set Assessee Details as active sheet ───────────────────────────────
    wb.active = wb["Assessee Details"]
    return _safe_save(wb, xlsx_path)


# ── Public API ────────────────────────────────────────────────────────────────

def convert_26as_txt(txt_path: str, log_callback=None) -> tuple[str, str]:
    """Parse txt_path and write .xlsx + .html alongside it.
    Returns (xlsx_path, html_path).
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    base = os.path.splitext(txt_path)[0]
    xlsx_path = base + ".xlsx"
    html_path = base + ".html"

    log(f"[26AS] Parsing: {os.path.basename(txt_path)}")
    parsed = _parse(txt_path)
    report_ts = datetime.now().strftime("%d-%b-%Y %I:%M %p")

    # row_ids is filled in during HTML building, then reused for Excel
    row_ids = {r: {} for r in ["I","II","III","IV","V","VI","VII","VIII","IX"]}

    log(f"[26AS] Building HTML...")
    html = _build_html(parsed, row_ids, report_ts)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"[26AS] HTML saved: {os.path.basename(html_path)}")

    # row_ids now populated with xl_row stubs (0); Excel writer fills xl_row properly
    row_ids = {r: {} for r in ["I","II","III","IV","V","VI","VII","VIII","IX"]}

    log(f"[26AS] Building Excel workbook...")
    try:
        actual_xlsx = _write_xlsx(parsed, row_ids, xlsx_path, report_ts)
        if actual_xlsx != xlsx_path:
            log(f"[Warning] '{os.path.basename(xlsx_path)}' was open in Excel — "
                f"saved as '{os.path.basename(actual_xlsx)}' instead.")
        log(f"[26AS] Excel saved: {os.path.basename(actual_xlsx)}")
        xlsx_path = actual_xlsx
    except PermissionError as pe:
        log(f"[Error] Excel save failed: {pe}")
        raise

    return xlsx_path, html_path
